# -*- coding: utf-8 -*-
"""
===============================================================================
 OPERAÇÃO ARGOS - Sistema Multi-Grupo
 Sistema de Download e Análise Exploratória
 Receita Estadual de Santa Catarina

 Grupos suportados: GESSUPER, GESMAC
===============================================================================
"""

# ============================================================
# IMPORTS PRINCIPAIS
# ============================================================
import streamlit as st
import pandas as pd
import numpy as np
import math
import time
import os
import gc
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots
from sqlalchemy import create_engine
import warnings
import ssl
import re
from datetime import datetime, timedelta
from io import BytesIO
import zipfile
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.workbook.properties import CalcProperties
import threading
import concurrent.futures

# Para salvar na rede
try:
    import smbclient
    SMB_AVAILABLE = True
except ImportError:
    SMB_AVAILABLE = False

# =============================================================================
# CONSTANTES GLOBAIS
# =============================================================================
MAX_ROWS_PER_EXCEL = 1000000
LARGE_FILE_WARNING = 200000
CACHE_TTL_SECONDS = 1800
SESSION_TIMEOUT_MINUTES = 30
LARGE_DATASET_THRESHOLD = 200000
REDE_PATH = r"\\sef.sc.gov.br\DFS\Fiscalizacao\NIAT\ARGOS\ARGOS_EXPORT"
RANKING_CACHE_TTL = 86400

# =============================================================================
# CONFIGURAÇÃO DOS GRUPOS
# =============================================================================
GRUPOS_CONFIG = {
    "GESSUPER": {
        "nome": "GESSUPER",
        "titulo": "Infrações GESSUPER",
        "cor_primaria": "#1565C0",
        "tabelas": {
            "nfce": "niat.infracoes_gessuper_nfce_3M",
            "cupons": "niat.infracoes_gessuper_cupons_3M"
        },
        "tem_nfe": False,
        "modelos_exportacao": ["padrao"],  # Só tem modelo padrão
        "colunas_extras": {}
    },
    "GESMAC": {
        "nome": "GESMAC",
        "titulo": "Infrações GESMAC",
        "cor_primaria": "#2E7D32",
        "tabelas": {
            "nfce": "niat.infracoes_gesmac_nfce_3m",
            "cupons": "niat.infracoes_gesmac_cupons_3m",
            "nfe": "niat.infracoes_gesmac_nfe_3m"
        },
        "tem_nfe": True,
        "modelos_exportacao": ["nfe", "nfce_cupom"],
        "colunas_extras": {
            "ie_emitente": "ie_emitente",
            "ie_destinatario": "ie_destinatario",
            "cnpj_destinatario": "cnpj_destinatario",
            "razao_destinatario": "razao_destinatario",
            "estado_destinatario": "estado_destinatario",
            "uf_entrega": "uf_entrega",
            "origem_prod": "origem_prod",
            "ind_final": "ind_final",
            "cst": "cst",
            "valor_total": "valor_total",
            "valor_do_frete": "valor_do_frete",
            "valor_do_seguro": "valor_do_seguro",
            "valor_outras_despesas": "valor_outras_despesas",
            "valor_do_desconto": "valor_do_desconto",
            "regime_destinatario": "regime_destinatario",
            "cnae_destinatario": "cnae_destinatario",
            "ttd_importacao": "ttd_importacao"
        }
    }
}

# Índices para cada modelo de exportação
INDICE_NFE_GESMAC = [
    ("Chave de acesso", "Indica do número da chave de acesso das Notas Fiscais. Não é aplicável para as informações da ECF."),
    ("URL", "Link para acessar o documento fiscal (apenas Notas Fiscais)."),
    ("Tipo Documento", "Indica a fonte da informação. Podia variar entre Nfe (Nota Fiscal Eletrônica), NFCe (Nota Fiscal do Consumidor Eletrônica) ou ECF (Emissor de Cupom Fiscal)"),
    ("Data de emissão", "Data de emissão do documento. (No caso de Cupom Fiscal, é a data da Redução Z)"),
    ("Entrada ou saída", "Indica se a operação é de entrada ou saída de mercadorias."),
    ("ECF-FAB", "Indica o número de série do Emissor de Cupom Fiscal (ECF). Não aplicável para operações com Notas Fiscais"),
    ("GTIN", "Código GTIN da mercadoria."),
    ("NCM", "Código NCM da mercadoria."),
    ("No. Nota", "Número da Nota Fiscal. Não é aplicável para informações da ECF."),
    ("No. Item", "Número do item dentro da Nota Fiscal. Não aplicável a Cupons."),
    ("Origem do Produto", "Informação de Origem do Produto retirado da Nota Fiscal. Não aplicável a ECF (Cupons) - Indica se o produto é nacional ou estrangeiro."),
    ("Ind Final e Tipo de Operação Final", "Informação de Ind Final retirado da Nota Fiscal. Não aplicável a ECF (Cupons). Indica se o destinatário receberá o produto para revenda/industrialização ou consumo final."),
    ("TTD 409/410/411", "Indica se o TTD 409, 410 ou 411 estava ativo para o contribuinte no respectivo período da Nota Fiscal. (Aplicável somente para Nfe)"),
    ("Código do produto", "Código do produto declarado pelo contribuinte para a operação. Válido apenas para Cupons Fiscais"),
    ("Cód. Tot. Par", "Código totalizador. Informação presenta apenas nas operações ECF."),
    ("Alíquota Destacada", "Alíquota de ICMS destacada no documento fiscal pelo contribuinte"),
    ("ICMS Destacado", "ICMS destacado no documento fiscal pelo contribuinte"),
    ("Valor da operação", "Valor da Base de Cálculo calculada pelo fisco, sem considerar reduções da base de cálculo. As reduções da BC serão aplicadas na alíquota efetiva correta. Para as notas fiscais (NF-e e NFC-e inclui frete, seguro, despesas adicionais , descontado os descontos concedidos). Para os Cupons leva-se em conta apenas o valor declarado na EFD que é o valor efetivo da operação."),
    ("Alíquota Efetiva Correta (FISCO)", "Alíquota de ICMS considerada pelo fisco para a operação. Aqui considerando eventuais reduções da Base de Cálculo. Para os Cupons fiscais é a alíquota retirada do COD TOT PAR."),
    ("Alíquota Efetiva destacada pelo Contribuinte", "Alíquota efetiva destacada pelo Contribuinte, que é calculada dividindo o ICMS destacado pelo Valor da Operação sem considerar redução da base de cálculo"),
    ("ICMS devido", "Valor do ICMS considerado como correto pelo fisco."),
    ("ICMS não-recolhido", "Valor do ICMS a ser recolhido como diferença pelo contribuinte. Trata-se da dedução do valor de \"ICMS devido\" pelo valor do campo \"ICMS destacado\"")
]

INDICE_NFCE_CUPOM_GESMAC = [
    ("Chave de acesso", "Indica do número da chave de acesso das Notas Fiscais. Não é aplicável para as informações da ECF."),
    ("URL", "Link para acessar o documento fiscal (apenas Notas Fiscais)."),
    ("Tipo Documento", "Indica a fonte da informação. Podia variar entre Nfe (Nota Fiscal Eletrônica), NFCe (Nota Fiscal do Consumidor Eletrônica) ou ECF (Emissor de Cupom Fiscal)"),
    ("Data de emissão", "Data de emissão do documento. (No caso de Cupom Fiscal, é a data da Redução Z)"),
    ("Entrada ou saída", "Indica se a operação é de entrada ou saída de mercadorias."),
    ("ECF-FAB", "Indica o número de série do Emissor de Cupom Fiscal (ECF). Não aplicável para operações com Notas Fiscais"),
    ("GTIN", "Código GTIN da mercadoria."),
    ("NCM", "Código NCM da mercadoria."),
    ("No. Nota", "Número da Nota Fiscal. Não é aplicável para informações da ECF."),
    ("No. Item", "Número do item dentro da Nota Fiscal. Não aplicável a Cupons."),
    ("Origem do Produto", "Informação de Origem do Produto retirado da Nota Fiscal. Não aplicável a ECF (Cupons) - Indica se o produto é nacional ou estrangeiro."),
    ("Ind Final e Tipo de Operação Final", "Informação de Ind Final retirado da Nota Fiscal. Não aplicável a ECF (Cupons). Indica se o destinatário receberá o produto para revenda/industrialização ou consumo final."),
    ("TTD 409/410/411", "Indica se o TTD 409, 410 ou 411 estava ativo para o contribuinte no respectivo período da Nota Fiscal. (Aplicável somente para Nfe)"),
    ("Código do produto", "Código do produto declarado pelo contribuinte para a operação. Válido apenas para Cupons Fiscais"),
    ("Cód. Tot. Par", "Código totalizador. Informação presenta apenas nas operações ECF."),
    ("Alíquota Destacada", "Alíquota de ICMS destacada no documento fiscal pelo contribuinte"),
    ("ICMS Destacado", "ICMS destacado no documento fiscal pelo contribuinte"),
    ("Valor da operação", "Valor da Base de Cálculo calculada pelo fisco, sem considerar reduções da base de cálculo. As reduções da BC serão aplicadas na alíquota efetiva correta. Para as notas fiscais (NF-e e NFC-e inclui frete, seguro, despesas adicionais , descontado os descontos concedidos). Para os Cupons leva-se em conta apenas o valor declarado na EFD que é o valor efetivo da operação."),
    ("Alíquota Efetiva Correta (FISCO)", "Alíquota de ICMS considerada pelo fisco para a operação. Aqui considerando eventuais reduções da Base de Cálculo. Para os Cupons fiscais é a alíquota retirada do COD TOT PAR."),
    ("Alíquota Efetiva destacada pelo Contribuinte", "Alíquota efetiva destacada pelo Contribuinte, que é calculada dividindo o ICMS destacado pelo Valor da Operação sem considerar redução da base de cálculo"),
    ("ICMS devido", "Valor do ICMS considerado como correto pelo fisco."),
    ("ICMS não-recolhido", "Valor do ICMS a ser recolhido como diferença pelo contribuinte. Trata-se da dedução do valor de \"ICMS devido\" pelo valor do campo \"ICMS destacado\"")
]

INDICE_GESSUPER = [
    ("Chave de acesso", "Número da chave de acesso das Notas Fiscais. Não aplicável para ECF."),
    ("URL", "Link para acessar o documento fiscal (apenas Notas Fiscais)."),
    ("Tipo Documento", "Fonte da informação: NFe, NFCe ou ECF."),
    ("Data de emissão", "Data de emissão do documento (Cupom Fiscal: data da Redução Z)."),
    ("Entrada ou saída", "Indica se a operação é de entrada ou saída."),
    ("ECF-FAB", "Número de série do Emissor de Cupom Fiscal."),
    ("GTIN", "Código GTIN da mercadoria."),
    ("NCM", "Código NCM da mercadoria."),
    ("No. Nota", "Número da Nota Fiscal."),
    ("No. Item", "Número do item dentro da Nota Fiscal."),
    ("Código do produto", "Código do produto declarado pelo contribuinte."),
    ("Cód. Tot. Par", "Código totalizador (apenas ECF)."),
    ("ICMS Destacado", "ICMS destacado no documento fiscal pelo contribuinte."),
    ("Valor da operação", "Base de Cálculo calculada pelo fisco."),
    ("Alíquota Efetiva Correta", "Alíquota de ICMS considerada pelo fisco."),
    ("Alíquota Efetiva destacada", "Alíquota efetiva destacada pelo Contribuinte."),
    ("ICMS devido", "Valor do ICMS considerado como correto pelo fisco."),
    ("ICMS não-recolhido", "Diferença entre ICMS devido e ICMS destacado.")
]

# =============================================================================
# CONFIGURAÇÕES INICIAIS
# =============================================================================
try:
    createunverified_https_context = ssl._create_unverified_context
except AttributeError:
    pass
else:
    ssl._create_default_https_context = createunverified_https_context

warnings.filterwarnings('ignore')

# Configuração da página
st.set_page_config(
    page_title="ARGOS - Sistema Multi-Grupo",
    page_icon="🎯",
    layout="wide",
    initial_sidebar_state="expanded"
)

# CSS customizado
st.markdown("""
<style>
    .main-header {
        font-size: 2.5rem;
        font-weight: bold;
        color: #1565C0;
        text-align: center;
        padding: 1rem 0;
    }

    div[data-testid="stMetric"] {
        background-color: #ffffff;
        border: 2px solid #2c3e50;
        border-radius: 10px;
        padding: 15px;
        box-shadow: 0 2px 4px rgba(0,0,0,0.1);
    }

    div[data-testid="stMetric"] > label {
        font-weight: 600;
        color: #2c3e50;
    }

    div[data-testid="stMetricValue"] {
        font-size: 1.8rem;
        font-weight: bold;
        color: #1f77b4;
    }

    .card-kpi {
        background: white;
        padding: 1.5rem;
        border-radius: 10px;
        border: 2px solid #e1e5f0;
        box-shadow: 0 2px 4px rgba(0,0,0,0.05);
    }

    .card-kpi-baixa { border-left: 5px solid #f44336 !important; }
    .card-kpi-media { border-left: 5px solid #FF9800 !important; }
    .card-kpi-alta { border-left: 5px solid #4CAF50 !important; }

    .status-badge {
        padding: 0.25rem 0.75rem;
        border-radius: 20px;
        font-size: 0.8rem;
        font-weight: 600;
    }

    .badge-success { background-color: #e8f5e9; color: #2e7d32; }
    .badge-warning { background-color: #fff3e0; color: #ef6c00; }
    .badge-danger { background-color: #ffebee; color: #c62828; }

    .info-box {
        background-color: #e3f2fd;
        border-left: 4px solid #1976d2;
        padding: 1rem;
        border-radius: 0 8px 8px 0;
        margin: 1rem 0;
    }

    .stTabs [data-baseweb="tab-list"] {
        gap: 24px;
    }

    .stTabs [data-baseweb="tab"] {
        height: 50px;
        padding-left: 20px;
        padding-right: 20px;
    }

    .grupo-gessuper {
        border-left: 5px solid #1565C0 !important;
    }

    .grupo-gesmac {
        border-left: 5px solid #2E7D32 !important;
    }
</style>
""", unsafe_allow_html=True)

# =============================================================================
# CREDENCIAIS E CONEXÃO
# =============================================================================
IMPALA_HOST = 'bdaworkernode02.sef.sc.gov.br'
IMPALA_PORT = 21050
DATABASE = 'niat'

try:
    IMPALA_USER = st.secrets["impala_credentials"]["user"]
    IMPALA_PASSWORD = st.secrets["impala_credentials"]["password"]
except:
    st.error("Credenciais não configuradas. Configure secrets.toml")
    st.info("""
    Crie o arquivo `.streamlit/secrets.toml` com:
    ```
    [impala_credentials]
    user = "seu_usuario"
    password = "sua_senha"
    ```
    """)
    st.stop()

# =============================================================================
# FUNÇÕES AUXILIARES
# =============================================================================

def is_table_unavailable_error(error_msg: str) -> bool:
    """Verifica se o erro é relacionado a tabela indisponível."""
    error_lower = str(error_msg).lower()
    table_error_patterns = [
        "could not resolve table reference",
        "table not found",
        "does not exist",
        "analysisexception",
        "no such table",
        "invalid table",
        "table or view not found",
        "relation.*does not exist",
        "unknown table"
    ]
    return any(pattern in error_lower for pattern in table_error_patterns)

TABLE_UNAVAILABLE_MSG = "Tabelas em atualização. Favor tentar novamente mais tarde."


def check_tables_available(engine, grupo: str = "GESSUPER") -> bool:
    """Verifica se as tabelas principais estão disponíveis para o grupo."""
    try:
        config = GRUPOS_CONFIG.get(grupo, GRUPOS_CONFIG["GESSUPER"])
        tabela = config["tabelas"].get("nfce", config["tabelas"].get("nfe"))
        query = f"SELECT 1 FROM {tabela} LIMIT 1"
        pd.read_sql(query, engine)
        return True
    except Exception as e:
        error_msg = str(e)
        if is_table_unavailable_error(error_msg):
            return False
        return True


def sanitize_identificador(raw: str) -> str:
    """Remove qualquer coisa que não seja dígito."""
    return re.sub(r"\D+", "", raw or "")


def format_currency_br(value) -> str:
    """Formata número como moeda brasileira."""
    if value is None:
        return "R$ 0,00"
    try:
        v = float(value)
    except (TypeError, ValueError):
        return "R$ 0,00"
    s = f"{v:,.2f}"
    s = s.replace(",", "X").replace(".", ",").replace("X", ".")
    return f"R$ {s}"


def format_number_br(value) -> str:
    """Formata número brasileiro."""
    if value is None:
        return "0"
    try:
        v = int(value)
    except (TypeError, ValueError):
        return "0"
    return f"{v:,}".replace(",", ".")


def nivel_config(nivel_str: str):
    """Retorna mapeamento de colunas para o nível escolhido."""
    nivel = (nivel_str or "").upper()
    if nivel == "MEDIA":
        return {
            "nivel": "MEDIA",
            "label": "MÉDIA",
            "col_aliquota": "aliquota_media",
            "col_legislacao": "legislacao_media",
            "col_infracao": "infracao_media",
            "cor": "#FF9800",
            "emoji": "🟡"
        }
    elif nivel == "ALTA":
        return {
            "nivel": "ALTA",
            "label": "ALTA",
            "col_aliquota": "aliquota_alta",
            "col_legislacao": "legislacao_alta",
            "col_infracao": "infracao_alta",
            "cor": "#4CAF50",
            "emoji": "🟢"
        }
    else:
        return {
            "nivel": "BAIXA",
            "label": "BAIXA",
            "col_aliquota": "aliquota_baixa",
            "col_legislacao": "legislacao_baixa",
            "col_infracao": "infracao_baixa",
            "cor": "#f44336",
            "emoji": "🔴"
        }


# =============================================================================
# CONEXÃO COM BANCO DE DADOS
# =============================================================================

@st.cache_resource
def get_engine():
    """Cria engine de conexão."""
    try:
        engine = create_engine(
            f'impala://{IMPALA_HOST}:{IMPALA_PORT}/{DATABASE}',
            connect_args={
                'user': IMPALA_USER,
                'password': IMPALA_PASSWORD,
                'auth_mechanism': 'LDAP',
                'use_ssl': True
            }
        )
        return engine
    except Exception as e:
        st.error(f"Erro de conexão: {str(e)[:100]}")
        return None


# =============================================================================
# FUNÇÕES DE CARREGAMENTO DE DADOS
# =============================================================================

@st.cache_data(ttl=CACHE_TTL_SECONDS, show_spinner="Buscando dados do contribuinte...")
def get_contribuinte_info(_engine, identificador_digits: str):
    """Busca informações do contribuinte."""
    query_cnpj = f"""
        SELECT
            nu_cnpj,
            nm_razao_social,
            nu_ie,
            nm_fantasia,
            nm_munic,
            cd_gerfe,
            nm_gerfe
        FROM usr_sat_ods.vw_ods_contrib
        WHERE regexp_replace(nu_cnpj, '[^0-9]', '') = '{identificador_digits}'
        LIMIT 1
    """

    query_ie = f"""
        SELECT
            nu_cnpj,
            nm_razao_social,
            nu_ie,
            nm_fantasia,
            nm_munic,
            cd_gerfe,
            nm_gerfe
        FROM usr_sat_ods.vw_ods_contrib
        WHERE regexp_replace(nu_ie, '[^0-9]', '') = '{identificador_digits}'
        LIMIT 1
    """

    try:
        df = pd.read_sql(query_cnpj, _engine)
        if df.empty:
            df = pd.read_sql(query_ie, _engine)

        if not df.empty:
            return {
                'cnpj': df['nu_cnpj'].iloc[0] if pd.notna(df['nu_cnpj'].iloc[0]) else '',
                'razao_social': df['nm_razao_social'].iloc[0] if pd.notna(df['nm_razao_social'].iloc[0]) else '',
                'ie': df['nu_ie'].iloc[0] if pd.notna(df['nu_ie'].iloc[0]) else '',
                'fantasia': df['nm_fantasia'].iloc[0] if pd.notna(df['nm_fantasia'].iloc[0]) else '',
                'municipio': df['nm_munic'].iloc[0] if pd.notna(df['nm_munic'].iloc[0]) else '',
                'gerfe': df['nm_gerfe'].iloc[0] if pd.notna(df['nm_gerfe'].iloc[0]) else ''
            }
        return None
    except Exception as e:
        error_msg = str(e)
        if is_table_unavailable_error(error_msg):
            st.session_state.tabela_indisponivel = True
        return None


@st.cache_data(ttl=86400, show_spinner=False)
def get_ncm_descricoes(_engine, ncm_list: list) -> dict:
    """Busca descrições dos NCMs."""
    if not ncm_list:
        return {}

    try:
        ncm_clean = [str(n).strip() for n in ncm_list if pd.notna(n) and str(n).strip()]
        if not ncm_clean:
            return {}

        ncm_str = "', '".join(ncm_clean)
        query = f"""
            SELECT ncm, descricao
            FROM niat.tabela_ncm
            WHERE ncm IN ('{ncm_str}')
        """
        df = pd.read_sql(query, _engine)
        return dict(zip(df['ncm'].astype(str), df['descricao']))
    except Exception as e:
        return {}


@st.cache_data(ttl=86400, show_spinner=False)
def get_cfop_descricoes(_engine, cfop_list: list) -> dict:
    """Busca descrições dos CFOPs."""
    if not cfop_list:
        return {}

    try:
        cfop_clean = [str(c).strip() for c in cfop_list if pd.notna(c) and str(c).strip()]
        if not cfop_clean:
            return {}

        cfop_str = "', '".join(cfop_clean)
        query = f"""
            SELECT cfop, descricaocfop
            FROM niat.tabela_cfop
            WHERE cfop IN ('{cfop_str}')
        """
        df = pd.read_sql(query, _engine)
        return dict(zip(df['cfop'].astype(str), df['descricaocfop']))
    except Exception as e:
        return {}


# =============================================================================
# FUNÇÕES DE CARREGAMENTO ESPECÍFICAS POR GRUPO
# =============================================================================

@st.cache_data(ttl=CACHE_TTL_SECONDS, show_spinner="Carregando dados base...")
def get_base_df_gessuper(_engine, identificador_digits: str, nivel: str = "BAIXA"):
    """Carrega DataFrame base para GESSUPER (NFC-e e Cupons)."""
    nivel_upper = (nivel or "BAIXA").upper()

    if nivel_upper == "ALTA":
        col_legislacao = "legislacao_alta"
        col_aliquota = "aliquota_alta"
        col_infracao = "infracao_alta"
    elif nivel_upper == "MEDIA":
        col_legislacao = "legislacao_media"
        col_aliquota = "aliquota_media"
        col_infracao = "infracao_media"
    else:
        col_legislacao = "legislacao_baixa"
        col_aliquota = "aliquota_baixa"
        col_infracao = "infracao_baixa"

    filtro_nivel = f"""
        {col_infracao} IS NOT NULL
        AND CAST({col_infracao} AS STRING) != 'EXCLUIR'
        AND CAST({col_aliquota} AS STRING) != 'EXCLUIR'
        AND CAST({col_legislacao} AS STRING) != 'EXCLUIR'
    """

    query = f"""
        SELECT
            data_emissao,
            periodo,
            tipo_doc,
            chave,
            NULL AS link_acesso,
            NULL AS modelo_ecf,
            entrada_ou_saida,
            cnpj_emitente,
            razao_emitente,
            numero_nota,
            gtin,
            ncm,
            CAST(numero_item AS STRING) AS numero_item,
            descricao,
            CAST(cfop AS STRING) AS cfop,
            icms_emitente,
            NULL AS cod_prod,
            NULL AS cod_tot_par,
            {col_legislacao} AS legislacao_ia,
            bc_fisco,
            {col_aliquota} AS aliquota_ia,
            NULL AS aliq_efetiva,
            {col_infracao} AS infracao_ia,
            NULL AS ie_emitente,
            NULL AS ie_destinatario,
            NULL AS cnpj_destinatario,
            NULL AS razao_destinatario,
            NULL AS estado_destinatario,
            NULL AS uf_entrega,
            NULL AS origem_prod,
            NULL AS ind_final,
            NULL AS cst,
            NULL AS valor_total,
            NULL AS valor_do_frete,
            NULL AS valor_do_seguro,
            NULL AS valor_outras_despesas,
            NULL AS valor_do_desconto,
            NULL AS regime_destinatario,
            NULL AS cnae_destinatario,
            NULL AS ttd_importacao
        FROM niat.infracoes_gessuper_nfce_3M
        WHERE regexp_replace(cnpj_emitente, '[^0-9]', '') = '{identificador_digits}'
        AND {filtro_nivel}

        UNION ALL

        SELECT
            data_emissao,
            periodo,
            tipo_doc,
            NULL AS chave,
            NULL AS link_acesso,
            modelo_ecf,
            NULL AS entrada_ou_saida,
            cnpj_emitente,
            razao_emitente,
            NULL AS numero_nota,
            gtin,
            ncm,
            CAST(NULL AS STRING) AS numero_item,
            descricao,
            CAST(cfop AS STRING) AS cfop,
            icms_emitente,
            cod_prod,
            cod_tot_par,
            {col_legislacao} AS legislacao_ia,
            bc_fisco,
            {col_aliquota} AS aliquota_ia,
            NULL AS aliq_efetiva,
            {col_infracao} AS infracao_ia,
            NULL AS ie_emitente,
            NULL AS ie_destinatario,
            NULL AS cnpj_destinatario,
            NULL AS razao_destinatario,
            NULL AS estado_destinatario,
            NULL AS uf_entrega,
            NULL AS origem_prod,
            NULL AS ind_final,
            NULL AS cst,
            NULL AS valor_total,
            NULL AS valor_do_frete,
            NULL AS valor_do_seguro,
            NULL AS valor_outras_despesas,
            NULL AS valor_do_desconto,
            NULL AS regime_destinatario,
            NULL AS cnae_destinatario,
            NULL AS ttd_importacao
        FROM niat.infracoes_gessuper_cupons_3M
        WHERE regexp_replace(cnpj_emitente, '[^0-9]', '') = '{identificador_digits}'
        AND {filtro_nivel}
    """
    try:
        df = pd.read_sql(query, _engine)
        return df
    except Exception as e:
        error_msg = str(e)
        if is_table_unavailable_error(error_msg):
            st.session_state.tabela_indisponivel = True
        return pd.DataFrame()


@st.cache_data(ttl=CACHE_TTL_SECONDS, show_spinner="Carregando dados base GESMAC...")
def get_base_df_gesmac(_engine, identificador_digits: str, nivel: str = "BAIXA"):
    """Carrega DataFrame base para GESMAC (NFCe, Cupons e NFe)."""
    nivel_upper = (nivel or "BAIXA").upper()

    if nivel_upper == "ALTA":
        col_legislacao = "legislacao_alta"
        col_aliquota = "aliquota_alta"
        col_infracao = "infracao_alta"
    elif nivel_upper == "MEDIA":
        col_legislacao = "legislacao_media"
        col_aliquota = "aliquota_media"
        col_infracao = "infracao_media"
    else:
        col_legislacao = "legislacao_baixa"
        col_aliquota = "aliquota_baixa"
        col_infracao = "infracao_baixa"

    filtro_nivel = f"""
        {col_infracao} IS NOT NULL
        AND CAST({col_infracao} AS STRING) != 'EXCLUIR'
        AND CAST({col_aliquota} AS STRING) != 'EXCLUIR'
        AND CAST({col_legislacao} AS STRING) != 'EXCLUIR'
    """

    # Query para NFCe GESMAC
    query = f"""
        SELECT
            data_emissao,
            periodo,
            tipo_doc,
            chave,
            NULL AS link_acesso,
            NULL AS modelo_ecf,
            entrada_ou_saida,
            ie_emitente,
            cnpj_emitente,
            razao_emitente,
            ie_destinatario,
            cnpj_destinatario,
            razao_destinatario,
            estado_destinatario,
            NULL as uf_entrega,
            numero_nota,
            numero_item,
            origem_prod,
            NULL as ind_final,
            NULL as cod_prod,
            gtin,
            ncm,
            descricao,
            CAST(cfop AS STRING) AS cfop,
            cst,
            valor_total,
            valor_do_frete,
            valor_do_seguro,
            valor_outras_despesas,
            valor_do_desconto,
            NULL as cod_tot_par,
            aliquota_emitente,
            icms_emitente,
            NULL as regime_destinatario,
            cnae_destinatario,
            NULL as ttd_importacao,
            bc_fisco,
            {col_aliquota} AS aliquota_ia,
            {col_legislacao} AS legislacao_ia,
            NULL AS aliq_efetiva,
            NULL AS icms_devido,
            {col_infracao} AS infracao_ia
        FROM niat.infracoes_gesmac_nfce_3m
        WHERE regexp_replace(cnpj_emitente, '[^0-9]', '') = '{identificador_digits}'
        AND {filtro_nivel}

        UNION ALL

        SELECT
            data_emissao,
            periodo,
            tipo_doc,
            NULL AS chave,
            NULL AS link_acesso,
            modelo_ecf,
            NULL as entrada_ou_saida,
            ie_emitente,
            cnpj_emitente,
            razao_emitente,
            NULL as ie_destinatario,
            NULL as cnpj_destinatario,
            NULL as razao_destinatario,
            NULL AS estado_destinatario,
            NULL AS uf_entrega,
            NULL as numero_nota,
            NULL as numero_item,
            NULL as origem_prod,
            NULL as ind_final,
            cod_prod,
            gtin,
            ncm,
            descricao,
            CAST(cfop AS STRING) AS cfop,
            NULL as cst,
            bc_fisco as valor_total,
            NULL as valor_do_frete,
            NULL as valor_do_seguro,
            NULL as valor_outras_despesas,
            NULL as valor_do_desconto,
            cod_tot_par,
            aliquota_emitente,
            icms_emitente,
            NULL as regime_destinatario,
            NULL as cnae_destinatario,
            NULL as ttd_importacao,
            bc_fisco,
            {col_aliquota} AS aliquota_ia,
            {col_legislacao} AS legislacao_ia,
            NULL AS aliq_efetiva,
            NULL AS icms_devido,
            {col_infracao} AS infracao_ia
        FROM niat.infracoes_gesmac_cupons_3m
        WHERE regexp_replace(cnpj_emitente, '[^0-9]', '') = '{identificador_digits}'
        AND {filtro_nivel}

        UNION ALL

        SELECT
            data_emissao,
            periodo,
            tipo_doc,
            chave,
            NULL AS link_acesso,
            NULL AS modelo_ecf,
            entrada_ou_saida,
            ie_emitente,
            cnpj_emitente,
            razao_emitente,
            ie_destinatario,
            cnpj_destinatario,
            razao_destinatario,
            estado_destinatario,
            uf_entrega,
            numero_nota,
            numero_item,
            origem_prod,
            ind_final,
            NULL as cod_prod,
            gtin,
            ncm,
            descricao,
            CAST(cfop AS STRING) AS cfop,
            cst,
            valor_total,
            valor_do_frete,
            valor_do_seguro,
            valor_outras_despesas,
            valor_do_desconto,
            NULL as cod_tot_par,
            aliquota_emitente,
            icms_emitente,
            regime_destinatario,
            cnae_destinatario,
            ttd_importacao,
            bc_fisco_red as bc_fisco,
            {col_aliquota} AS aliquota_ia,
            {col_legislacao} AS legislacao_ia,
            NULL AS aliq_efetiva,
            NULL AS icms_devido,
            {col_infracao} AS infracao_ia
        FROM niat.infracoes_gesmac_nfe_3m
        WHERE regexp_replace(cnpj_emitente, '[^0-9]', '') = '{identificador_digits}'
        AND {filtro_nivel}
    """
    try:
        df = pd.read_sql(query, _engine)
        return df
    except Exception as e:
        error_msg = str(e)
        if is_table_unavailable_error(error_msg):
            st.session_state.tabela_indisponivel = True
        return pd.DataFrame()


def get_base_df(_engine, identificador_digits: str, nivel: str = "BAIXA", grupo: str = "GESSUPER"):
    """Função wrapper que chama a função correta baseada no grupo."""
    if grupo == "GESMAC":
        return get_base_df_gesmac(_engine, identificador_digits, nivel)
    else:
        return get_base_df_gessuper(_engine, identificador_digits, nivel)


def calcular_totais(df: pd.DataFrame, nivel_str: str):
    """Retorna total_nivel, cfg (dict do nível), has_rows (bool)."""
    cfg = nivel_config(nivel_str)

    if df.empty:
        return 0.0, cfg, False

    if 'infracao_ia' in df.columns:
        col_infracao = 'infracao_ia'
    else:
        col_infracao = cfg['col_infracao']

    df_calc = df.copy()
    df_calc['infracao_valor'] = pd.to_numeric(df_calc[col_infracao], errors='coerce').fillna(0)
    total_nivel = df_calc['infracao_valor'].sum()

    return float(total_nivel), cfg, True


# =============================================================================
# FUNÇÕES DE EXPORTAÇÃO
# =============================================================================

def build_export_df(df: pd.DataFrame, nivel_str: str, grupo: str = "GESSUPER"):
    """Monta o DataFrame pronto para exportar."""
    cfg = nivel_config(nivel_str)

    if df.empty:
        return None

    df_export = df.copy()

    if 'infracao_ia' in df.columns:
        df_export['legislacao_ia_icms'] = df_export['legislacao_ia']
        df_export['aliquota_ia_icms'] = df_export['aliquota_ia']
        df_export['icms_devido'] = pd.to_numeric(df_export['infracao_ia'], errors='coerce').fillna(0)
    else:
        col_legislacao = cfg['col_legislacao']
        col_aliquota = cfg['col_aliquota']
        col_infracao = cfg['col_infracao']
        df_export['legislacao_ia_icms'] = df_export[col_legislacao]
        df_export['aliquota_ia_icms'] = df_export[col_aliquota]
        df_export['icms_devido'] = pd.to_numeric(df_export[col_infracao], errors='coerce').fillna(0)

    return df_export


def export_to_csv(df: pd.DataFrame, identificador: str, nivel: str) -> bytes:
    """Exporta DataFrame para CSV no formato brasileiro."""
    csv_str = df.to_csv(index=False, sep=";", decimal=",")
    return csv_str.encode("latin-1", errors="replace")


def get_export_filename(contrib_info: dict, nivel: str, extension: str, grupo: str = "GESSUPER", modelo: str = None) -> str:
    """Gera o nome do arquivo."""
    if contrib_info:
        cnpj = sanitize_identificador(contrib_info.get('cnpj', ''))
        razao = contrib_info.get('razao_social', 'EMPRESA')
        razao_clean = re.sub(r'[<>:"/\\|?*]', '', razao)[:50]
        if modelo:
            return f"{cnpj} - {razao_clean} - {modelo}.{extension}"
        return f"{cnpj} - {razao_clean}.{extension}"
    modelo_str = f"_{modelo}" if modelo else ""
    return f"infracoes_{grupo.lower()}{modelo_str}_{nivel.lower()}.{extension}"


def export_to_excel_template_gessuper(df: pd.DataFrame, contrib_info: dict, nivel: str, parte_atual: int = None, total_partes: int = None, progress_callback=None) -> bytes:
    """Exporta DataFrame para Excel usando template GESSUPER (Anexo J)."""
    def report_progress(pct, msg):
        if progress_callback:
            progress_callback(pct, msg)

    report_progress(5, "Criando estrutura do arquivo")

    buffer = BytesIO()
    wb = Workbook()

    # ABA 1: ANEXO J1 - NOTAS DE SAÍDAS
    ws1 = wb.active
    ws1.title = "ANEXO J1 - NOTAS DE SAÍDAS"

    # Estilos
    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
    header_fill_fisco = PatternFill(start_color="C62828", end_color="C62828", fill_type="solid")
    title_font = Font(bold=True, size=14, color="1565C0")
    subtitle_font = Font(bold=True, size=11, color="666666")

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    titulo_j1 = "ANEXO J1"
    if parte_atual is not None and total_partes is not None:
        titulo_j1 = f"ANEXO J1 - Parte {parte_atual} de {total_partes}"
    ws1['A1'] = titulo_j1
    ws1['A1'].font = title_font

    ws1['D2'] = "INFORMAÇÕES RETIRADAS DOS DOCUMENTOS FISCAIS (Cupons Fiscais ou NFC-e)"
    ws1['D2'].font = subtitle_font
    ws1['S2'] = "INFORMAÇÕES DECLARADAS PELO FISCO"
    ws1['S2'].font = Font(bold=True, size=11, color="C62828")

    headers_j1 = [
        "Data de emissão", "Período", "Tipo Documento", "Chave de acesso", "Link de Acesso",
        "ECF-FAB", "Entrada ou saída", "CNPJ Emitente", "Razão do Emitente", "Número",
        "GTIN", "NCM", "Item", "Descrição do produto", "CFOP", "ICMS destacado",
        "Código do Produto", "Cód. Tot. Par", "Legislação", "Valor da Operação",
        "Alíquota ICMS correta", "Alíquota ICMS efetiva", "ICMS devido", "ICMS não-recolhido"
    ]

    for col_idx, header in enumerate(headers_j1, 1):
        cell = ws1.cell(row=3, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill_fisco if col_idx >= 19 else header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    ws1.row_dimensions[3].height = 30

    column_mapping = {
        'data_emissao': 'A', 'periodo': 'B', 'tipo_doc': 'C', 'chave': 'D',
        'modelo_ecf': 'F', 'entrada_ou_saida': 'G', 'cnpj_emitente': 'H',
        'razao_emitente': 'I', 'numero_nota': 'J', 'gtin': 'K', 'ncm': 'L',
        'numero_item': 'M', 'descricao': 'N', 'cfop': 'O', 'icms_emitente': 'P',
        'cod_prod': 'Q', 'cod_tot_par': 'R', 'legislacao_ia_icms': 'S',
        'bc_fisco': 'T', 'aliquota_ia_icms': 'U'
    }

    report_progress(10, "Preenchendo dados da aba J1")

    if 'data_emissao' in df.columns:
        df = df.sort_values('data_emissao', ascending=True, na_position='last').reset_index(drop=True)

    total_rows = len(df)
    progress_interval = max(1, total_rows // 20)

    for row_idx, row_data in enumerate(df.itertuples(index=False), 4):
        atual_row = row_idx - 4
        if atual_row % progress_interval == 0:
            pct = 10 + int((atual_row / total_rows) * 50)
            report_progress(pct, f"Processando linha {atual_row:,} de {total_rows:,}")

        row_dict = row_data._asdict()

        for col_name, col_letter in column_mapping.items():
            if col_name in row_dict:
                col_idx = ord(col_letter) - ord('A') + 1
                cell = ws1.cell(row=row_idx, column=col_idx)
                value = row_dict[col_name]

                if col_name == 'data_emissao' and pd.notna(value):
                    try:
                        if isinstance(value, str):
                            cell.value = pd.to_datetime(value, dayfirst=True).date()
                        else:
                            cell.value = value
                        cell.number_format = 'DD/MM/YYYY'
                    except:
                        cell.value = value
                elif col_name == 'periodo' and pd.notna(value):
                    try:
                        if isinstance(value, str):
                            cell.value = pd.to_datetime(value, dayfirst=True).date()
                        else:
                            cell.value = value
                        cell.number_format = 'DD/MM/YYYY'
                    except:
                        cell.value = value
                elif col_name in ['icms_emitente', 'bc_fisco'] and pd.notna(value):
                    try:
                        cell.value = float(value)
                        cell.number_format = '#,##0.00'
                    except:
                        cell.value = value
                elif col_name == 'aliquota_ia_icms' and pd.notna(value):
                    try:
                        cell.value = float(value) / 100
                        cell.number_format = '0.00%'
                    except:
                        cell.value = value
                else:
                    cell.value = value if pd.notna(value) else ''

                cell.border = thin_border

        # Fórmulas
        cell_e = ws1.cell(row=row_idx, column=5)
        cell_e.value = f'=IF(D{row_idx}<>"",HYPERLINK("https://sat.sef.sc.gov.br/tax.NET/Sat.NFe.Web/Consultas/Nfe_ResumoPDF.ashx?id="&D{row_idx},"Abrir DANFE"),"")'
        cell_e.border = thin_border

        cell_v = ws1.cell(row=row_idx, column=22)
        cell_v.value = f"=IF(T{row_idx}=0,0,P{row_idx}/T{row_idx})"
        cell_v.number_format = '0.00%'
        cell_v.border = thin_border

        cell_w = ws1.cell(row=row_idx, column=23)
        cell_w.value = f"=T{row_idx}*U{row_idx}"
        cell_w.number_format = '#,##0.00'
        cell_w.border = thin_border

        cell_x = ws1.cell(row=row_idx, column=24)
        cell_x.value = f"=W{row_idx}-P{row_idx}"
        cell_x.number_format = '#,##0.00'
        cell_x.border = thin_border

    # Autoajuste
    for col_idx in range(1, 25):
        col_letter = get_column_letter(col_idx)
        max_length = 0
        for row in ws1.iter_rows(min_row=1, max_row=min(ws1.max_row, 100), min_col=col_idx, max_col=col_idx):
            for cell in row:
                try:
                    if cell.value:
                        if str(cell.value).startswith('='):
                            cell_length = 12
                        else:
                            cell_length = len(str(cell.value))
                        max_length = max(max_length, cell_length)
                except:
                    pass
        adjusted_width = min(max(max_length + 2, 8), 50)
        ws1.column_dimensions[col_letter].width = adjusted_width

    ws1.freeze_panes = 'A4'
    ultima_linha = 3 + len(df)
    ws1.auto_filter.ref = f"A3:X{ultima_linha}"

    report_progress(65, "Criando aba J2 - Resumo ICMS")

    # ABA 2: ANEXO J2 - ICMS DEVIDO
    ws2 = wb.create_sheet("ANEXO J2 - ICMS DEVIDO")

    ws2.merge_cells('A1:D1')
    ws2['A1'] = "ESTADO DE SANTA CATARINA"
    ws2['A1'].font = Font(bold=True, size=14)
    ws2['A1'].alignment = Alignment(horizontal="center")

    ws2.merge_cells('A2:D2')
    ws2['A2'] = "Secretaria de Estado da Fazenda"
    ws2['A2'].alignment = Alignment(horizontal="center")

    ws2.merge_cells('A3:D3')
    ws2['A3'] = "Diretoria de Administração Tributária"
    ws2['A3'].alignment = Alignment(horizontal="center")

    ws2.merge_cells('A4:D4')
    ws2['A4'] = "Gerência de Fiscalização"
    ws2['A4'].alignment = Alignment(horizontal="center")

    ws2['A6'] = "CNPJ:"
    ws2['A6'].font = Font(bold=True)
    ws2['B6'] = contrib_info.get('cnpj', '') if contrib_info else ''

    ws2['A7'] = "Razão Social:"
    ws2['A7'].font = Font(bold=True)
    ws2['B7'] = contrib_info.get('razao_social', '') if contrib_info else ''

    ws2.merge_cells('A10:D10')
    ws2['A10'] = "APURAÇÃO MENSAL DO VALOR DO ICMS DEVIDO NAS VENDAS DE MERCADORIAS"
    ws2['A10'].font = Font(bold=True, size=12)
    ws2['A10'].alignment = Alignment(horizontal="center")

    headers_j2 = ["Período", "ICMS destacado", "ICMS apurado", "ICMS não recolhido"]
    for col_idx, header in enumerate(headers_j2, 1):
        cell = ws2.cell(row=11, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    if 'periodo' in df.columns:
        periodos_unicos = df['periodo'].dropna().unique()
        periodos = sorted(periodos_unicos, key=lambda x: pd.to_datetime(x, dayfirst=True) if isinstance(x, str) else x)
    else:
        periodos = []

    ultima_linha_dados = len(df) + 3

    for row_idx, periodo in enumerate(periodos, 12):
        cell_a = ws2.cell(row=row_idx, column=1)
        try:
            if isinstance(periodo, str):
                cell_a.value = pd.to_datetime(periodo, dayfirst=True).date()
            else:
                cell_a.value = periodo
            cell_a.number_format = 'DD/MM/YYYY'
        except:
            cell_a.value = periodo
        cell_a.border = thin_border
        cell_a.alignment = Alignment(horizontal="center")

        cell_b = ws2.cell(row=row_idx, column=2)
        cell_b.value = f"=SUMIF('ANEXO J1 - NOTAS DE SAÍDAS'!$B$4:$B${ultima_linha_dados},$A{row_idx},'ANEXO J1 - NOTAS DE SAÍDAS'!$P$4:$P${ultima_linha_dados})"
        cell_b.number_format = '#,##0.00'
        cell_b.border = thin_border

        cell_c = ws2.cell(row=row_idx, column=3)
        cell_c.value = f"=SUMIF('ANEXO J1 - NOTAS DE SAÍDAS'!$B$4:$B${ultima_linha_dados},$A{row_idx},'ANEXO J1 - NOTAS DE SAÍDAS'!$W$4:$W${ultima_linha_dados})"
        cell_c.number_format = '#,##0.00'
        cell_c.border = thin_border

        cell_d = ws2.cell(row=row_idx, column=4)
        cell_d.value = f"=C{row_idx}-B{row_idx}"
        cell_d.number_format = '#,##0.00'
        cell_d.border = thin_border

    total_row = 12 + len(periodos)
    ws2.cell(row=total_row, column=1).value = "TOTAL"
    ws2.cell(row=total_row, column=1).font = Font(bold=True)
    ws2.cell(row=total_row, column=1).border = thin_border

    for col in range(2, 5):
        cell = ws2.cell(row=total_row, column=col)
        cell.value = f"=SUM({chr(64+col)}12:{chr(64+col)}{total_row-1})"
        cell.number_format = '#,##0.00'
        cell.font = Font(bold=True)
        cell.border = thin_border
        cell.fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")

    for col_idx in range(1, 5):
        col_letter = get_column_letter(col_idx)
        ws2.column_dimensions[col_letter].width = 20

    report_progress(80, "Criando aba Índice")

    # ABA 3: Índice
    ws3 = wb.create_sheet("Índice")
    ws3['A1'] = "Campo"
    ws3['B1'] = "Descrição"
    ws3['A1'].font = header_font
    ws3['B1'].font = header_font
    ws3['A1'].fill = header_fill
    ws3['B1'].fill = header_fill

    for row_idx, (campo, desc) in enumerate(INDICE_GESSUPER, 2):
        ws3.cell(row=row_idx, column=1).value = campo
        ws3.cell(row=row_idx, column=2).value = desc

    ws3.column_dimensions['A'].width = 30
    ws3.column_dimensions['B'].width = 100

    report_progress(90, "Configurando recálculo automático")

    wb.calculation = CalcProperties(fullCalcOnLoad=True, calcMode='auto')

    report_progress(95, "Salvando arquivo Excel")
    wb.save(buffer)
    buffer.seek(0)
    report_progress(100, "Concluído!")

    return buffer.getvalue()


def export_to_excel_template_gesmac_nfe(df: pd.DataFrame, contrib_info: dict, nivel: str, parte_atual: int = None, total_partes: int = None, progress_callback=None) -> bytes:
    """Exporta DataFrame para Excel usando template GESMAC NFe."""
    def report_progress(pct, msg):
        if progress_callback:
            progress_callback(pct, msg)

    report_progress(5, "Criando estrutura do arquivo NFe")

    # Filtra apenas NFe
    df_nfe = df[df['tipo_doc'].str.upper().str.contains('NFE', na=False) & ~df['tipo_doc'].str.upper().str.contains('NFCE', na=False)].copy()

    if df_nfe.empty:
        df_nfe = df.copy()  # Se não houver filtro, usa todos

    buffer = BytesIO()
    wb = Workbook()

    ws1 = wb.active
    ws1.title = "ANEXO NFE"

    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    header_fill_fisco = PatternFill(start_color="C62828", end_color="C62828", fill_type="solid")
    title_font = Font(bold=True, size=14, color="2E7D32")
    subtitle_font = Font(bold=True, size=11, color="666666")

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    titulo = "ANEXO NFE - GESMAC"
    if parte_atual is not None and total_partes is not None:
        titulo = f"ANEXO NFE - GESMAC - Parte {parte_atual} de {total_partes}"
    ws1['A1'] = titulo
    ws1['A1'].font = title_font

    # Headers NFe GESMAC
    headers_nfe = [
        "Data de emissão", "Período", "Tipo Documento", "Chave de acesso", "Link de Acesso",
        "ECF-FAB", "Entrada ou saída", "IE Emitente", "CNPJ Emitente", "Razão do Emitente",
        "IE Destinatário", "CNPJ Destinatário", "CPF Destinatário", "Razão do Destinatário",
        "Estado do Destinatário", "Regime do Destinatário", "CNAE do Destinatário",
        "Número da Nota", "Número do Item", "Origem do Produto", "Ind Final",
        "Tipo de Operação Final", "TTD 409/410/411", "GTIN", "NCM", "Descrição do produto",
        "CFOP", "Código do Produto", "Valor Total", "Valor do Frete", "Valor do Seguro",
        "Valor de Outras Despesas", "Valor do Desconto", "Cód. Tot. Par",
        "Alíquota Destacada", "ICMS Destacado", "Valor da Operação",
        "Alíquota Efetiva Correta (FISCO)", "Legislação Aplicável",
        "Alíquota Efetiva destacada pelo Contribuinte", "ICMS devido", "ICMS não-recolhido"
    ]

    for col_idx, header in enumerate(headers_nfe, 1):
        cell = ws1.cell(row=3, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill_fisco if col_idx >= 37 else header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    ws1.row_dimensions[3].height = 30

    report_progress(10, "Preenchendo dados NFe")

    if 'data_emissao' in df_nfe.columns:
        df_nfe = df_nfe.sort_values('data_emissao', ascending=True, na_position='last').reset_index(drop=True)

    total_rows = len(df_nfe)
    progress_interval = max(1, total_rows // 20)

    for row_idx, row_data in enumerate(df_nfe.itertuples(index=False), 4):
        atual_row = row_idx - 4
        if atual_row % progress_interval == 0:
            pct = 10 + int((atual_row / total_rows) * 50)
            report_progress(pct, f"Processando linha {atual_row:,} de {total_rows:,}")

        row_dict = row_data._asdict()

        # Mapeamento das colunas
        col_data = [
            ('data_emissao', 1), ('periodo', 2), ('tipo_doc', 3), ('chave', 4),
            (None, 5),  # Link - fórmula
            ('modelo_ecf', 6), ('entrada_ou_saida', 7), ('ie_emitente', 8),
            ('cnpj_emitente', 9), ('razao_emitente', 10), ('ie_destinatario', 11),
            ('cnpj_destinatario', 12), (None, 13),  # CPF Destinatário
            ('razao_destinatario', 14), ('estado_destinatario', 15),
            ('regime_destinatario', 16), ('cnae_destinatario', 17),
            ('numero_nota', 18), ('numero_item', 19), ('origem_prod', 20),
            ('ind_final', 21), (None, 22),  # Tipo operação
            ('ttd_importacao', 23), ('gtin', 24), ('ncm', 25),
            ('descricao', 26), ('cfop', 27), ('cod_prod', 28),
            ('valor_total', 29), ('valor_do_frete', 30), ('valor_do_seguro', 31),
            ('valor_outras_despesas', 32), ('valor_do_desconto', 33),
            ('cod_tot_par', 34), ('aliquota_emitente', 35), ('icms_emitente', 36),
            ('bc_fisco', 37), ('aliquota_ia_icms', 38), ('legislacao_ia_icms', 39)
        ]

        for col_name, col_idx in col_data:
            if col_name and col_name in row_dict:
                cell = ws1.cell(row=row_idx, column=col_idx)
                value = row_dict[col_name]

                if col_name in ['data_emissao', 'periodo'] and pd.notna(value):
                    try:
                        if isinstance(value, str):
                            cell.value = pd.to_datetime(value, dayfirst=True).date()
                        else:
                            cell.value = value
                        cell.number_format = 'DD/MM/YYYY'
                    except:
                        cell.value = value
                elif col_name in ['icms_emitente', 'bc_fisco', 'valor_total', 'valor_do_frete', 'valor_do_seguro', 'valor_outras_despesas', 'valor_do_desconto'] and pd.notna(value):
                    try:
                        cell.value = float(value)
                        cell.number_format = '#,##0.00'
                    except:
                        cell.value = value
                elif col_name in ['aliquota_ia_icms', 'aliquota_emitente'] and pd.notna(value):
                    try:
                        cell.value = float(value) / 100
                        cell.number_format = '0.00%'
                    except:
                        cell.value = value
                else:
                    cell.value = value if pd.notna(value) else ''

                cell.border = thin_border

        # Fórmula Link
        cell_link = ws1.cell(row=row_idx, column=5)
        cell_link.value = f'=IF(D{row_idx}<>"",HYPERLINK("https://sat.sef.sc.gov.br/tax.NET/Sat.NFe.Web/Consultas/Nfe_ResumoPDF.ashx?id="&D{row_idx},"Abrir DANFE"),"")'
        cell_link.border = thin_border

        # Alíquota efetiva contribuinte (col 40)
        cell_40 = ws1.cell(row=row_idx, column=40)
        cell_40.value = f"=IF(AK{row_idx}=0,0,AJ{row_idx}/AK{row_idx})"
        cell_40.number_format = '0.00%'
        cell_40.border = thin_border

        # ICMS devido (col 41)
        cell_41 = ws1.cell(row=row_idx, column=41)
        cell_41.value = f"=AK{row_idx}*AL{row_idx}"
        cell_41.number_format = '#,##0.00'
        cell_41.border = thin_border

        # ICMS não-recolhido (col 42)
        cell_42 = ws1.cell(row=row_idx, column=42)
        cell_42.value = f"=AO{row_idx}-AJ{row_idx}"
        cell_42.number_format = '#,##0.00'
        cell_42.border = thin_border

    ws1.freeze_panes = 'A4'

    report_progress(65, "Criando aba ICMS DEVIDO")

    # ABA ICMS DEVIDO
    ws2 = wb.create_sheet("ICMS DEVIDO")
    headers_icms = ["Período", "ICMS destacado", "ICMS apurado", "ICMS não recolhido"]
    for col_idx, header in enumerate(headers_icms, 1):
        cell = ws2.cell(row=1, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    if 'periodo' in df_nfe.columns:
        periodos = sorted(df_nfe['periodo'].dropna().unique(), key=lambda x: pd.to_datetime(x, dayfirst=True) if isinstance(x, str) else x)
        for row_idx, periodo in enumerate(periodos, 2):
            ws2.cell(row=row_idx, column=1).value = periodo
            ws2.cell(row=row_idx, column=1).border = thin_border

    for col_idx in range(1, 5):
        ws2.column_dimensions[get_column_letter(col_idx)].width = 20

    report_progress(80, "Criando aba Índice")

    # ABA Índice
    ws3 = wb.create_sheet("Índice")
    ws3['A1'] = "Campo"
    ws3['B1'] = "Descrição"
    ws3['A1'].font = header_font
    ws3['B1'].font = header_font
    ws3['A1'].fill = header_fill
    ws3['B1'].fill = header_fill

    for row_idx, (campo, desc) in enumerate(INDICE_NFE_GESMAC, 2):
        ws3.cell(row=row_idx, column=1).value = campo
        ws3.cell(row=row_idx, column=2).value = desc

    ws3.column_dimensions['A'].width = 40
    ws3.column_dimensions['B'].width = 100

    wb.calculation = CalcProperties(fullCalcOnLoad=True, calcMode='auto')

    report_progress(95, "Salvando arquivo")
    wb.save(buffer)
    buffer.seek(0)
    report_progress(100, "Concluído!")

    return buffer.getvalue()


def export_to_excel_template_gesmac_nfce_cupom(df: pd.DataFrame, contrib_info: dict, nivel: str, parte_atual: int = None, total_partes: int = None, progress_callback=None) -> bytes:
    """Exporta DataFrame para Excel usando template GESMAC NFCe + Cupom Fiscal."""
    def report_progress(pct, msg):
        if progress_callback:
            progress_callback(pct, msg)

    report_progress(5, "Criando estrutura do arquivo NFCe + Cupom")

    # Filtra NFCe e Cupom (exclui NFe pura)
    df_nfce = df[~(df['tipo_doc'].str.upper().str.contains('NFE', na=False) & ~df['tipo_doc'].str.upper().str.contains('NFCE', na=False))].copy()

    if df_nfce.empty:
        df_nfce = df.copy()

    buffer = BytesIO()
    wb = Workbook()

    ws1 = wb.active
    ws1.title = "ANEXO NFCE-CUPOM"

    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    header_fill_fisco = PatternFill(start_color="C62828", end_color="C62828", fill_type="solid")
    title_font = Font(bold=True, size=14, color="2E7D32")

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    titulo = "ANEXO NFCe + CUPOM FISCAL - GESMAC"
    if parte_atual is not None and total_partes is not None:
        titulo = f"ANEXO NFCe + CUPOM - GESMAC - Parte {parte_atual} de {total_partes}"
    ws1['A1'] = titulo
    ws1['A1'].font = title_font

    ws1['A2'] = "INFORMAÇÕES RETIRADAS DOS DOCUMENTOS FISCAIS (Cupons Fiscais ou NFC-e)"
    ws1['A2'].font = Font(bold=True, size=11, color="666666")

    # Headers NFCe + Cupom GESMAC
    headers_nfce = [
        "Data de emissão", "Período", "Tipo Documento", "Chave de acesso", "Link de Acesso",
        "ECF-FAB", "Entrada ou saída", "IE Emitente", "CNPJ Emitente", "Razão do Emitente",
        "IE Destinatário", "CNPJ Destinatário", "CPF Destinatário", "Razão do Destinatário",
        "Estado do Destinatário", "Regime do Destinatário", "CNAE do Destinatário",
        "Número da Nota", "Número do Item", "Origem do Produto", "Ind Final",
        "Tipo de Operação Final", "TTD 409/410/411", "GTIN", "NCM", "Descrição do produto",
        "CFOP", "Código do Produto", "Valor Total", "Valor do Frete", "Valor do Seguro",
        "Valor de Outras Despesas", "Valor do Desconto", "Cód. Tot. Par",
        "Alíquota Destacada", "ICMS Destacado", "Valor da Operação",
        "Alíquota Efetiva Correta (FISCO)", "Legislação Aplicável",
        "Alíquota Efetiva destacada pelo Contribuinte", "ICMS devido", "ICMS não-recolhido"
    ]

    for col_idx, header in enumerate(headers_nfce, 1):
        cell = ws1.cell(row=3, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill_fisco if col_idx >= 37 else header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    ws1.row_dimensions[3].height = 30

    report_progress(10, "Preenchendo dados NFCe + Cupom")

    if 'data_emissao' in df_nfce.columns:
        df_nfce = df_nfce.sort_values('data_emissao', ascending=True, na_position='last').reset_index(drop=True)

    total_rows = len(df_nfce)
    progress_interval = max(1, total_rows // 20)

    for row_idx, row_data in enumerate(df_nfce.itertuples(index=False), 4):
        atual_row = row_idx - 4
        if atual_row % progress_interval == 0:
            pct = 10 + int((atual_row / total_rows) * 50)
            report_progress(pct, f"Processando linha {atual_row:,} de {total_rows:,}")

        row_dict = row_data._asdict()

        col_data = [
            ('data_emissao', 1), ('periodo', 2), ('tipo_doc', 3), ('chave', 4),
            (None, 5), ('modelo_ecf', 6), ('entrada_ou_saida', 7), ('ie_emitente', 8),
            ('cnpj_emitente', 9), ('razao_emitente', 10), ('ie_destinatario', 11),
            ('cnpj_destinatario', 12), (None, 13), ('razao_destinatario', 14),
            ('estado_destinatario', 15), ('regime_destinatario', 16), ('cnae_destinatario', 17),
            ('numero_nota', 18), ('numero_item', 19), ('origem_prod', 20),
            ('ind_final', 21), (None, 22), ('ttd_importacao', 23), ('gtin', 24),
            ('ncm', 25), ('descricao', 26), ('cfop', 27), ('cod_prod', 28),
            ('valor_total', 29), ('valor_do_frete', 30), ('valor_do_seguro', 31),
            ('valor_outras_despesas', 32), ('valor_do_desconto', 33),
            ('cod_tot_par', 34), ('aliquota_emitente', 35), ('icms_emitente', 36),
            ('bc_fisco', 37), ('aliquota_ia_icms', 38), ('legislacao_ia_icms', 39)
        ]

        for col_name, col_idx in col_data:
            if col_name and col_name in row_dict:
                cell = ws1.cell(row=row_idx, column=col_idx)
                value = row_dict[col_name]

                if col_name in ['data_emissao', 'periodo'] and pd.notna(value):
                    try:
                        if isinstance(value, str):
                            cell.value = pd.to_datetime(value, dayfirst=True).date()
                        else:
                            cell.value = value
                        cell.number_format = 'DD/MM/YYYY'
                    except:
                        cell.value = value
                elif col_name in ['icms_emitente', 'bc_fisco', 'valor_total', 'valor_do_frete', 'valor_do_seguro', 'valor_outras_despesas', 'valor_do_desconto'] and pd.notna(value):
                    try:
                        cell.value = float(value)
                        cell.number_format = '#,##0.00'
                    except:
                        cell.value = value
                elif col_name in ['aliquota_ia_icms', 'aliquota_emitente'] and pd.notna(value):
                    try:
                        cell.value = float(value) / 100
                        cell.number_format = '0.00%'
                    except:
                        cell.value = value
                else:
                    cell.value = value if pd.notna(value) else ''

                cell.border = thin_border

        # Fórmulas
        cell_link = ws1.cell(row=row_idx, column=5)
        cell_link.value = f'=IF(D{row_idx}<>"",HYPERLINK("https://sat.sef.sc.gov.br/tax.NET/Sat.NFe.Web/Consultas/Nfe_ResumoPDF.ashx?id="&D{row_idx},"Abrir DANFE"),"")'
        cell_link.border = thin_border

        cell_40 = ws1.cell(row=row_idx, column=40)
        cell_40.value = f"=IF(AK{row_idx}=0,0,AJ{row_idx}/AK{row_idx})"
        cell_40.number_format = '0.00%'
        cell_40.border = thin_border

        cell_41 = ws1.cell(row=row_idx, column=41)
        cell_41.value = f"=AK{row_idx}*AL{row_idx}"
        cell_41.number_format = '#,##0.00'
        cell_41.border = thin_border

        cell_42 = ws1.cell(row=row_idx, column=42)
        cell_42.value = f"=AO{row_idx}-AJ{row_idx}"
        cell_42.number_format = '#,##0.00'
        cell_42.border = thin_border

    ws1.freeze_panes = 'A4'

    report_progress(65, "Criando aba ICMS DEVIDO")

    # ABA ICMS DEVIDO
    ws2 = wb.create_sheet("ICMS DEVIDO")
    headers_icms = ["Período", "ICMS destacado", "ICMS apurado", "ICMS não recolhido"]
    for col_idx, header in enumerate(headers_icms, 1):
        cell = ws2.cell(row=1, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    if 'periodo' in df_nfce.columns:
        periodos = sorted(df_nfce['periodo'].dropna().unique(), key=lambda x: pd.to_datetime(x, dayfirst=True) if isinstance(x, str) else x)
        for row_idx, periodo in enumerate(periodos, 2):
            ws2.cell(row=row_idx, column=1).value = periodo
            ws2.cell(row=row_idx, column=1).border = thin_border

    for col_idx in range(1, 5):
        ws2.column_dimensions[get_column_letter(col_idx)].width = 20

    report_progress(80, "Criando aba Índice")

    # ABA Índice
    ws3 = wb.create_sheet("Índice")
    ws3['A1'] = "Campo"
    ws3['B1'] = "Descrição"
    ws3['A1'].font = header_font
    ws3['B1'].font = header_font
    ws3['A1'].fill = header_fill
    ws3['B1'].fill = header_fill

    for row_idx, (campo, desc) in enumerate(INDICE_NFCE_CUPOM_GESMAC, 2):
        ws3.cell(row=row_idx, column=1).value = campo
        ws3.cell(row=row_idx, column=2).value = desc

    ws3.column_dimensions['A'].width = 40
    ws3.column_dimensions['B'].width = 100

    wb.calculation = CalcProperties(fullCalcOnLoad=True, calcMode='auto')

    report_progress(95, "Salvando arquivo")
    wb.save(buffer)
    buffer.seek(0)
    report_progress(100, "Concluído!")

    return buffer.getvalue()


def export_to_excel_template(df: pd.DataFrame, contrib_info: dict, nivel: str, grupo: str = "GESSUPER", modelo: str = None, parte_atual: int = None, total_partes: int = None, progress_callback=None) -> bytes:
    """Função wrapper que escolhe o template correto baseado no grupo e modelo."""
    if grupo == "GESMAC":
        if modelo == "nfe":
            return export_to_excel_template_gesmac_nfe(df, contrib_info, nivel, parte_atual, total_partes, progress_callback)
        elif modelo == "nfce_cupom":
            return export_to_excel_template_gesmac_nfce_cupom(df, contrib_info, nivel, parte_atual, total_partes, progress_callback)
        else:
            return export_to_excel_template_gesmac_nfce_cupom(df, contrib_info, nivel, parte_atual, total_partes, progress_callback)
    else:
        return export_to_excel_template_gessuper(df, contrib_info, nivel, parte_atual, total_partes, progress_callback)


# =============================================================================
# ANÁLISES EXPLORATÓRIAS
# =============================================================================

def render_analise_exploratoria(df: pd.DataFrame, nivel_str: str, _engine=None):
    """Renderiza análises exploratórias dos dados."""
    if df.empty:
        st.warning("Sem dados para análise.")
        return

    cfg = nivel_config(nivel_str)

    if 'infracao_ia' in df.columns:
        col_infracao = 'infracao_ia'
    else:
        col_infracao = cfg['col_infracao']

    st.markdown("---")
    st.subheader("Análise Exploratória")

    tabs = st.tabs(["Visão Temporal", "Por NCM/CFOP", "Por Produto", "Distribuição"])

    with tabs[0]:
        col1, col2 = st.columns(2)

        with col1:
            df_temp = df.copy()
            df_temp['infracao_valor'] = pd.to_numeric(df_temp[col_infracao], errors='coerce').fillna(0)

            if 'periodo' in df_temp.columns:
                df_periodo = df_temp.groupby('periodo').agg({
                    'infracao_valor': 'sum',
                    'chave': 'count'
                }).reset_index()
                df_periodo.columns = ['Período', 'Valor Infração', 'Quantidade']
                df_periodo = df_periodo.sort_values('Período')

                fig = px.bar(
                    df_periodo,
                    x='Período',
                    y='Valor Infração',
                    title=f"Infrações por Período (Nível {cfg['label']})",
                    color_discrete_sequence=[cfg['cor']]
                )
                st.plotly_chart(fig, use_container_width=True)

        with col2:
            if 'periodo' in df_temp.columns and 'df_periodo' in dir():
                fig2 = px.line(
                    df_periodo,
                    x='Período',
                    y='Quantidade',
                    title="Quantidade de Itens por Período",
                    markers=True
                )
                st.plotly_chart(fig2, use_container_width=True)

    with tabs[1]:
        col1, col2 = st.columns(2)

        with col1:
            df_temp = df.copy()
            df_temp['infracao_valor'] = pd.to_numeric(df_temp[col_infracao], errors='coerce').fillna(0)

            if 'ncm' in df_temp.columns:
                df_ncm = df_temp.groupby('ncm').agg({
                    'infracao_valor': ['sum', 'count']
                }).reset_index()
                df_ncm.columns = ['NCM', 'Valor Total', 'Itens']
                df_ncm = df_ncm.nlargest(10, 'Valor Total')

                st.markdown("##### Top NCM por Valor")
                st.dataframe(df_ncm, use_container_width=True, hide_index=True)

        with col2:
            if 'cfop' in df_temp.columns:
                df_cfop = df_temp.groupby('cfop').agg({
                    'infracao_valor': ['sum', 'count']
                }).reset_index()
                df_cfop.columns = ['CFOP', 'Valor Total', 'Itens']
                df_cfop = df_cfop.nlargest(10, 'Valor Total')

                st.markdown("##### Top CFOP por Valor")
                st.dataframe(df_cfop, use_container_width=True, hide_index=True)

    with tabs[2]:
        df_temp = df.copy()
        df_temp['infracao_valor'] = pd.to_numeric(df_temp[col_infracao], errors='coerce').fillna(0)

        if 'descricao' in df_temp.columns:
            df_prod = df_temp.groupby('descricao').agg({
                'infracao_valor': ['sum', 'count']
            }).reset_index()
            df_prod.columns = ['Descrição', 'Valor Total', 'Itens']
            df_prod = df_prod.nlargest(15, 'Valor Total')

            st.markdown("##### Top Produtos por Valor")
            st.dataframe(df_prod, use_container_width=True, hide_index=True)

    with tabs[3]:
        df_temp = df.copy()
        df_temp['infracao_valor'] = pd.to_numeric(df_temp[col_infracao], errors='coerce').fillna(0)
        df_temp = df_temp[df_temp['infracao_valor'] > 0]

        col1, col2 = st.columns(2)

        with col1:
            fig = px.histogram(
                df_temp,
                x='infracao_valor',
                nbins=30,
                title="Distribuição dos Valores de Infração",
                color_discrete_sequence=[cfg['cor']]
            )
            st.plotly_chart(fig, use_container_width=True)

        with col2:
            fig = px.box(
                df_temp,
                y='infracao_valor',
                title="Box Plot - Valores de Infração",
                color_discrete_sequence=[cfg['cor']]
            )
            st.plotly_chart(fig, use_container_width=True)

        st.markdown("##### Estatísticas Descritivas")
        col1, col2, col3, col4, col5 = st.columns(5)

        with col1:
            st.metric("Mínimo", format_currency_br(df_temp['infracao_valor'].min()))
        with col2:
            st.metric("Máximo", format_currency_br(df_temp['infracao_valor'].max()))
        with col3:
            st.metric("Média", format_currency_br(df_temp['infracao_valor'].mean()))
        with col4:
            st.metric("Mediana", format_currency_br(df_temp['infracao_valor'].median()))
        with col5:
            st.metric("Desvio Padrão", format_currency_br(df_temp['infracao_valor'].std()))


# =============================================================================
# INTERFACE PRINCIPAL
# =============================================================================

def main():
    """Interface principal com navegação e seleção de grupo."""

    # Gerenciamento de sessão
    if 'last_activity' not in st.session_state:
        st.session_state.last_activity = datetime.now()

    time_since_activity = datetime.now() - st.session_state.last_activity
    if time_since_activity > timedelta(minutes=SESSION_TIMEOUT_MINUTES):
        if st.session_state.get('consulta_dados') is not None:
            st.session_state.consulta_dados = None
            st.cache_data.clear()
            gc.collect()

    st.session_state.last_activity = datetime.now()

    if 'consulta_dados' not in st.session_state:
        st.session_state.consulta_dados = None

    if 'tabela_indisponivel' not in st.session_state:
        st.session_state.tabela_indisponivel = False

    if 'grupo_selecionado' not in st.session_state:
        st.session_state.grupo_selecionado = "GESSUPER"

    engine = get_engine()
    if engine is None:
        st.stop()

    # Sidebar
    with st.sidebar:
        grupo_config = GRUPOS_CONFIG[st.session_state.grupo_selecionado]

        st.markdown(f"""
        <div style='text-align: center; padding: 0.5rem 0; border-bottom: 2px solid {grupo_config["cor_primaria"]}; margin-bottom: 1rem;'>
            <h2 style='color: {grupo_config["cor_primaria"]}; margin: 0;'>ARGOS</h2>
            <p style='color: #666; margin: 0; font-size: 0.8rem;'>{grupo_config["titulo"]}</p>
        </div>
        """, unsafe_allow_html=True)

        # Seletor de Grupo
        st.markdown("### Selecione o Grupo")
        grupo_novo = st.selectbox(
            "Grupo",
            options=list(GRUPOS_CONFIG.keys()),
            index=list(GRUPOS_CONFIG.keys()).index(st.session_state.grupo_selecionado),
            format_func=lambda x: f"{GRUPOS_CONFIG[x]['nome']}",
            key="grupo_selector",
            label_visibility="collapsed"
        )

        if grupo_novo != st.session_state.grupo_selecionado:
            st.session_state.grupo_selecionado = grupo_novo
            st.session_state.consulta_dados = None
            st.rerun()

        st.markdown("---")

        # Níveis de Acurácia
        st.markdown("### Níveis de Acurácia")
        st.success("**ALTA**\n\nConsenso das 3 IAs\n\n*1-2% de erros esperados*")
        st.warning("**MÉDIA**\n\nMaioria 2x1\n\n*Até 5% de erros*")
        st.error("**BAIXA**\n\nIAs divergentes\n\n*Requer avaliação manual!*")

        st.markdown("---")

        with st.expander("Sistema", expanded=False):
            st.caption(f"Cache consulta: {CACHE_TTL_SECONDS//60} min")
            st.caption(f"Grupo atual: {st.session_state.grupo_selecionado}")

            if st.button("Limpar Cache", use_container_width=True):
                st.cache_data.clear()
                st.cache_resource.clear()
                for key in list(st.session_state.keys()):
                    del st.session_state[key]
                gc.collect()
                st.rerun()

        st.markdown("---")
        st.caption("Receita Estadual de SC")

    grupo = st.session_state.grupo_selecionado
    grupo_config = GRUPOS_CONFIG[grupo]

    # Área Principal
    if st.session_state.consulta_dados is None:
        # Tela inicial
        st.markdown(f"""
        <h2 style='color: {grupo_config["cor_primaria"]}; margin: 0;'>Operação ARGOS</h2>
        <p style='color: #666; margin: 0; font-size: 0.9rem;'>Sistema de Infrações {grupo_config["nome"]} - Receita Estadual SC</p>
        """, unsafe_allow_html=True)

        st.markdown("---")

        if not check_tables_available(engine, grupo):
            st.warning(TABLE_UNAVAILABLE_MSG)
            st.stop()

        # Campo de Consulta
        col_esq, col_form, col_dir = st.columns([1, 3, 1])

        with col_form:
            st.markdown("### Consultar Empresa")

            col_input, col_nivel = st.columns([2, 1])

            with col_input:
                cnpj_ie_input = st.text_input(
                    "CNPJ ou IE",
                    placeholder="00.000.000/0000-00 ou 000000000",
                    key="cnpj_input_principal",
                    label_visibility="collapsed"
                )

            with col_nivel:
                nivel_consulta = st.selectbox(
                    "Nível",
                    options=["ALTA", "MEDIA", "BAIXA"],
                    format_func=lambda x: {"BAIXA": "BAIXA", "MEDIA": "MÉDIA", "ALTA": "ALTA"}[x],
                    key="nivel_input_principal",
                    index=0,
                    label_visibility="collapsed"
                )

            if st.button("CONSULTAR", type="primary", use_container_width=True):
                if cnpj_ie_input:
                    ident_digits = sanitize_identificador(cnpj_ie_input)
                    if ident_digits:
                        st.session_state.tabela_indisponivel = False

                        with st.status("Consultando...", expanded=True) as status:
                            st.write("Buscando contribuinte...")
                            progress_bar = st.progress(0)
                            contrib_info = get_contribuinte_info(engine, ident_digits)
                            progress_bar.progress(25)

                            if st.session_state.get('tabela_indisponivel', False):
                                status.update(label="Tabelas indisponíveis", state="error", expanded=False)
                                st.warning(TABLE_UNAVAILABLE_MSG)
                            else:
                                if contrib_info:
                                    st.write(f"**{contrib_info.get('razao_social', 'N/A')}**")
                                else:
                                    st.write("Contribuinte não encontrado")

                                st.write(f"Carregando infrações ({nivel_consulta})...")
                                progress_bar.progress(50)
                                df = get_base_df(engine, ident_digits, nivel_consulta, grupo)
                                progress_bar.progress(100)

                                if st.session_state.get('tabela_indisponivel', False):
                                    status.update(label="Tabelas indisponíveis", state="error", expanded=False)
                                    st.warning(TABLE_UNAVAILABLE_MSG)
                                elif df.empty:
                                    status.update(label="Nenhum registro", state="error", expanded=False)
                                    st.warning(f"Nenhum registro para: {cnpj_ie_input}")
                                else:
                                    status.update(label=f"{len(df):,} registros", state="complete", expanded=False)
                                    st.session_state.consulta_dados = {
                                        'df': df,
                                        'contrib_info': contrib_info,
                                        'ident_digits': ident_digits,
                                        'identificador': cnpj_ie_input,
                                        'nivel': nivel_consulta,
                                        'grupo': grupo
                                    }
                                    st.rerun()
                    else:
                        st.error("CNPJ ou IE inválido.")
                else:
                    st.warning("Digite um CNPJ ou IE.")

        # Cards de níveis
        st.markdown("---")
        col1, col2, col3 = st.columns(3)
        with col1:
            st.success("**ALTA** - Consenso 3 IAs (1-2% erros)")
        with col2:
            st.warning("**MÉDIA** - Maioria 2x1 (até 5% erros)")
        with col3:
            st.error("**BAIXA** - IAs divergentes")

    else:
        # Exibe resultados da consulta
        dados = st.session_state.consulta_dados
        df = dados['df']
        contrib_info = dados['contrib_info']
        ident_digits = dados['ident_digits']
        identificador_consulta = dados['identificador']
        nivel_consulta = dados['nivel']
        grupo_consulta = dados.get('grupo', 'GESSUPER')

        nivel_atual = nivel_consulta
        total_nivel, cfg, has_rows = calcular_totais(df, nivel_atual)

        if contrib_info:
            razao_social = contrib_info.get('razao_social', 'N/A')
            cnpj_formatado = contrib_info.get('cnpj', identificador_consulta)
            ie_formatado = contrib_info.get('ie', '')
            municipio = contrib_info.get('municipio', '')
            gerfe = contrib_info.get('gerfe', '')
        else:
            razao_social = df['razao_emitente'].iloc[0] if 'razao_emitente' in df.columns and not df['razao_emitente'].isna().all() else "N/A"
            cnpj_formatado = identificador_consulta
            ie_formatado = ''
            municipio = ''
            gerfe = ''

        # Header
        col_header, col_btn = st.columns([4, 1])
        with col_btn:
            if st.button("Nova Consulta", use_container_width=True, type="secondary"):
                st.session_state.consulta_dados = None
                st.rerun()

        grupo_config_atual = GRUPOS_CONFIG[grupo_consulta]
        st.markdown(f"""
        <div style='background: linear-gradient(135deg, {grupo_config_atual["cor_primaria"]} 0%, #0D47A1 100%);
                    padding: 1rem 1.5rem; border-radius: 10px; margin-bottom: 1rem; color: white;'>
            <div style='display: flex; justify-content: space-between; align-items: center; flex-wrap: wrap;'>
                <div>
                    <h3 style='margin: 0; color: white;'>{razao_social}</h3>
                    <p style='margin: 0.3rem 0 0 0; opacity: 0.9;'>
                        CNPJ: {cnpj_formatado} | IE: {ie_formatado} | {municipio} | GERFE: {gerfe}
                    </p>
                </div>
                <div style='text-align: right;'>
                    <span style='background: {cfg["cor"]}; padding: 0.3rem 0.8rem; border-radius: 20px;
                                 font-weight: bold;'>{cfg["emoji"]} {cfg["label"]}</span>
                    <p style='margin: 0.3rem 0 0 0; opacity: 0.9;'>{format_number_br(len(df))} itens | {format_currency_br(total_nivel)}</p>
                    <p style='margin: 0; opacity: 0.7; font-size: 0.8rem;'>Grupo: {grupo_consulta}</p>
                </div>
            </div>
        </div>
        """, unsafe_allow_html=True)

        # Tabs
        tab_resumo, tab_exportar, tab_analise = st.tabs([
            "Resumo", "Exportar", "Análise"
        ])

        with tab_resumo:
            st.markdown("### Resumo da Empresa")

            col1, col2, col3, col4 = st.columns(4)
            with col1:
                st.metric(f"{cfg['emoji']} Total Infração", format_currency_br(total_nivel))
            with col2:
                st.metric("Qtd. Itens", format_number_br(len(df)))
            with col3:
                periodos = df['periodo'].nunique() if 'periodo' in df.columns else 0
                st.metric("Períodos", periodos)
            with col4:
                if 'data_emissao' in df.columns:
                    df_datas = pd.to_datetime(df['data_emissao'], errors='coerce')
                    if not df_datas.isna().all():
                        periodo_range = f"{df_datas.min().strftime('%m/%Y')} - {df_datas.max().strftime('%m/%Y')}"
                    else:
                        periodo_range = "N/A"
                else:
                    periodo_range = "N/A"
                st.metric("Range", periodo_range)

            # Tabela de tipos de documento
            if 'tipo_doc' in df.columns:
                st.markdown("#### Por Tipo de Documento")
                df_tipo = df.groupby('tipo_doc').agg({
                    'infracao_ia': lambda x: pd.to_numeric(x, errors='coerce').sum()
                }).reset_index()
                df_tipo.columns = ['Tipo', 'Valor Total']
                df_tipo['Qtd'] = df.groupby('tipo_doc').size().values
                st.dataframe(df_tipo, use_container_width=True, hide_index=True)

        with tab_exportar:
            df_export = build_export_df(df, nivel_atual, grupo_consulta)

            if df_export is not None and not df_export.empty:
                total_rows = len(df_export)

                st.info(f"Total de {total_rows:,} linhas para exportação")

                # Para GESMAC, mostra opções de modelo
                if grupo_consulta == "GESMAC":
                    st.markdown("#### Selecione o Modelo de Exportação")

                    col1, col2 = st.columns(2)

                    with col1:
                        if st.button("Exportar NFe", use_container_width=True, type="primary"):
                            with st.spinner("Gerando arquivo NFe..."):
                                excel_data = export_to_excel_template(df_export, contrib_info, nivel_atual, grupo_consulta, "nfe")
                                filename = get_export_filename(contrib_info, nivel_atual, "xlsx", grupo_consulta, "NFe")
                                st.download_button(
                                    label="Download NFe",
                                    data=excel_data,
                                    file_name=filename,
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                                )

                    with col2:
                        if st.button("Exportar NFCe + Cupom", use_container_width=True, type="primary"):
                            with st.spinner("Gerando arquivo NFCe + Cupom..."):
                                excel_data = export_to_excel_template(df_export, contrib_info, nivel_atual, grupo_consulta, "nfce_cupom")
                                filename = get_export_filename(contrib_info, nivel_atual, "xlsx", grupo_consulta, "NFCe_Cupom")
                                st.download_button(
                                    label="Download NFCe + Cupom",
                                    data=excel_data,
                                    file_name=filename,
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                                )

                else:
                    # GESSUPER - modelo único
                    if st.button("Gerar Excel", use_container_width=True, type="primary"):
                        with st.spinner("Gerando arquivo Excel..."):
                            excel_data = export_to_excel_template(df_export, contrib_info, nivel_atual, grupo_consulta)
                            filename = get_export_filename(contrib_info, nivel_atual, "xlsx", grupo_consulta)
                            st.download_button(
                                label="Download Excel",
                                data=excel_data,
                                file_name=filename,
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )

                # CSV para ambos
                st.markdown("---")
                if st.button("Exportar CSV", use_container_width=True):
                    csv_data = export_to_csv(df_export, identificador_consulta, nivel_atual)
                    filename_csv = get_export_filename(contrib_info, nivel_atual, "csv", grupo_consulta)
                    st.download_button(
                        label="Download CSV",
                        data=csv_data,
                        file_name=filename_csv,
                        mime="text/csv"
                    )

        with tab_analise:
            render_analise_exploratoria(df, nivel_atual, engine)


if __name__ == "__main__":
    main()
