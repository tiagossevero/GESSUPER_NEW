# -*- coding: utf-8 -*-
"""
===============================================================================
 OPERAÇÃO ARGOS - Infrações GESSUPER
 Sistema de Download e Análise Exploratória
 Receita Estadual de Santa Catarina
===============================================================================
"""

# ============================================================
# IMPORTS PRINCIPAIS
# ============================================================
import streamlit as st
import pandas as pd
import numpy as np
import math
import time
import os
import gc  # Garbage Collector para limpeza de memória
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots
from sqlalchemy import create_engine
import warnings
import ssl
import re
from datetime import datetime, timedelta
from io import BytesIO
import zipfile
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.workbook.properties import CalcProperties
import threading
import concurrent.futures

# Para salvar na rede
try:
    import smbclient
    SMB_AVAILABLE = True
except ImportError:
    SMB_AVAILABLE = False

# Limite de linhas por arquivo Excel (Excel suporta 1.048.576, usamos 1.000.000 para segurança)
MAX_ROWS_PER_EXCEL = 1000000

# Limite para aviso de arquivo grande (acima disso, recomenda CSV)
LARGE_FILE_WARNING = 200000  # 200k linhas

# TTL do cache em segundos (1 hora = 3600, reduzido para economizar memória)
CACHE_TTL_SECONDS = 1800  # 30 minutos

# Tempo máximo de sessão inativa antes de limpar dados (em minutos)
SESSION_TIMEOUT_MINUTES = 30

# Limite de linhas para aplicar filtro de 12 meses (performance)
LARGE_DATASET_THRESHOLD = 200000

# Caminho da rede para salvar arquivos (evita consumo de memória)
REDE_PATH = r"\\sef.sc.gov.br\DFS\Fiscalizacao\NIAT\ARGOS\ARGOS_EXPORT"

# Cache do ranking (24 horas = 86400 segundos)
RANKING_CACHE_TTL = 86400


# =============================================================================
# CONFIGURAÇÃO DE GRUPOS (EXTENSÍVEL)
# =============================================================================
# Cada grupo tem suas próprias tabelas e configurações de exportação.
# Para adicionar um novo grupo, basta adicionar uma nova entrada neste dicionário.

GRUPOS_CONFIG = {
    "GESSUPER": {
        "nome": "GESSUPER",
        "nome_display": "Infrações GESSUPER",
        "descricao": "Sistema de Infrações GESSUPER - Receita Estadual SC",
        # Tabelas (sem NFe para GESSUPER)
        "tabelas": {
            "nfce": "niat.infracoes_gessuper_nfce_3M",
            "cupons": "niat.infracoes_gessuper_cupons_3M",
            "nfe": None  # GESSUPER não tem NFe
        },
        # Modelos de exportação
        "modelos_exportacao": ["Anexo J"],  # Modelo único para NFCe + Cupom
        # Colunas específicas para export (padrão do GESSUPER)
        "export_config": {
            "Anexo J": {
                "titulo_aba_dados": "ANEXO J1 - NOTAS DE SAÍDAS",
                "titulo_aba_icms": "ANEXO J2 - ICMS DEVIDO",
                "colunas_header": [
                    "Data de emissão", "Período", "Tipo Documento", "Chave de acesso",
                    "Link de Acesso", "ECF-FAB", "Entrada ou saída", "CNPJ Emitente",
                    "Razão do Emitente", "Número", "GTIN", "NCM", "Item",
                    "Descrição do produto", "CFOP", "ICMS destacado", "Código do Produto",
                    "Cód. Tot. Par", "Legislação", "Valor da Operação", "Alíquota ICMS correta",
                    "Alíquota ICMS efetiva", "ICMS devido", "ICMS não-recolhido"
                ]
            }
        },
        # Índice de descrição dos campos
        "indice_campos": [
            ("Chave de acesso", "Número da chave de acesso das Notas Fiscais. Não aplicável para ECF."),
            ("URL", "Link para acessar o documento fiscal (apenas Notas Fiscais)."),
            ("Tipo Documento", "Fonte da informação: NFe, NFCe ou ECF."),
            ("Data de emissão", "Data de emissão do documento (Cupom Fiscal: data da Redução Z)."),
            ("Entrada ou saída", "Indica se a operação é de entrada ou saída."),
            ("ECF-FAB", "Número de série do Emissor de Cupom Fiscal."),
            ("GTIN", "Código GTIN da mercadoria."),
            ("NCM", "Código NCM da mercadoria."),
            ("No. Nota", "Número da Nota Fiscal."),
            ("No. Item", "Número do item dentro da Nota Fiscal."),
            ("Código do produto", "Código do produto declarado pelo contribuinte."),
            ("Cód. Tot. Par", "Código totalizador (apenas ECF)."),
            ("ICMS Destacado", "ICMS destacado no documento fiscal pelo contribuinte."),
            ("Valor da operação", "Base de Cálculo calculada pelo fisco."),
            ("Alíquota Efetiva Correta", "Alíquota de ICMS considerada pelo fisco."),
            ("Alíquota Efetiva destacada", "Alíquota efetiva destacada pelo Contribuinte."),
            ("ICMS devido", "Valor do ICMS considerado como correto pelo fisco."),
            ("ICMS não-recolhido", "Diferença entre ICMS devido e ICMS destacado.")
        ]
    },
    "GESMAC": {
        "nome": "GESMAC",
        "nome_display": "Infrações GESMAC",
        "descricao": "Sistema de Infrações GESMAC - Receita Estadual SC",
        # Tabelas (com NFe para GESMAC)
        "tabelas": {
            "nfce": "niat.infracoes_gesmac_nfce_3m",
            "cupons": "niat.infracoes_gesmac_cupons_3m",
            "nfe": "niat.infracoes_gesmac_nfe_3m"  # GESMAC tem NFe
        },
        # Modelos de exportação (2 modelos para GESMAC)
        "modelos_exportacao": ["NFe", "NFCe + Cupom Fiscal"],
        # Colunas específicas para export
        "export_config": {
            "NFe": {
                "titulo_aba_dados": "ANEXO NFe - NOTAS DE SAÍDAS",
                "titulo_aba_icms": "ICMS DEVIDO - NFe",
                "colunas_header": [
                    "Data de emissão", "Período", "Tipo Documento", "Chave de acesso",
                    "Link de Acesso", "ECF-FAB", "Entrada ou saída", "IE Emitente",
                    "CNPJ Emitente", "Razão do Emitente", "IE Destinatário",
                    "CNPJ Destinatário", "CPF Destinatário", "Razão do Destinatário",
                    "Estado do Destinatário", "Regime do Destinatário", "CNAE do Destinatário",
                    "Número da Nota", "Número do Item", "Origem do Produto", "Ind Final",
                    "Tipo de Operação Final", "TTD 409/410/411", "GTIN", "NCM",
                    "Descrição do produto", "CFOP", "Código do Produto", "Valor Total",
                    "Valor do Frete", "Valor do Seguro", "Valor de Outras Despesas",
                    "Valor do Desconto", "Cod. Tot. Par", "Alíquota Destacada", "ICMS Destacado",
                    "Valor da Operação", "Alíquota Efetiva Correta (FISCO)", "Legislação Aplicável",
                    "Alíquota Efetiva destacada pelo Contribuinte", "ICMS devido", "ICMS não-recolhido"
                ]
            },
            "NFCe + Cupom Fiscal": {
                "titulo_aba_dados": "ANEXO NFCe+CF - DOCUMENTOS",
                "titulo_aba_icms": "ICMS DEVIDO - NFCe+CF",
                "colunas_header": [
                    "Data de emissão", "Período", "Tipo Documento", "Chave de acesso",
                    "Link de Acesso", "ECF-FAB", "Entrada ou saída", "IE Emitente",
                    "CNPJ Emitente", "Razão do Emitente", "IE Destinatário",
                    "CNPJ Destinatário", "CPF Destinatário", "Razão do Destinatário",
                    "Estado do Destinatário", "Regime do Destinatário", "CNAE do Destinatário",
                    "Número da Nota", "Número do Item", "Origem do Produto", "Ind Final",
                    "Tipo de Operação Final", "TTD 409/410/411", "GTIN", "NCM",
                    "Descrição do produto", "CFOP", "Código do Produto", "Valor Total",
                    "Valor do Frete", "Valor do Seguro", "Valor de Outras Despesas",
                    "Valor do Desconto", "Cod. Tot. Par", "Alíquota Destacada", "ICMS Destacado",
                    "Valor da Operação", "Alíquota Efetiva Correta (FISCO)", "Legislação Aplicável",
                    "Alíquota Efetiva destacada pelo Contribuinte", "ICMS devido", "ICMS não-recolhido"
                ]
            }
        },
        # Índice de descrição dos campos para GESMAC
        "indice_campos": [
            ("Chave de acesso", "Indica do número da chave de acesso das Notas Fiscais. Não é aplicável para as informações da ECF."),
            ("URL", "Link para acessar o documento fiscal (apenas Notas Fiscais)."),
            ("Tipo Documento", "Indica a fonte da informação. Podia variar entre NFe (Nota Fiscal Eletrônica), NFCe (Nota Fiscal do Consumidor Eletrônica) ou ECF (Emissor de Cupom Fiscal)"),
            ("Data de emissão", "Data de emissão do documento. (No caso de Cupom Fiscal, é a data da Redução Z)"),
            ("Entrada ou saída", "Indica se a operação é de entrada ou saída de mercadorias."),
            ("ECF-FAB", "Indica o número de série do Emissor de Cupom Fiscal (ECF). Não aplicável para operações com Notas Fiscais"),
            ("GTIN", "Código GTIN da mercadoria."),
            ("NCM", "Código NCM da mercadoria."),
            ("No. Nota", "Número da Nota Fiscal. Não é aplicável para informações da ECF."),
            ("No. Item", "Número do item dentro da Nota Fiscal. Não aplicável a Cupons."),
            ("Origem do Produto", "Informação de Origem do Produto retirado da Nota Fiscal. Não aplicável a ECF (Cupons) - Indica se o produto é nacional ou estrangeiro."),
            ("Ind Final e Tipo de Operação Final", "Informação de Ind Final retirado da Nota Fiscal. Não aplicável a ECF (Cupons). Indica se o destinatário receberá o produto para revenda/industrialização ou consumo final."),
            ("TTD 409/410/411", "Indica se o TTD 409, 410 ou 411 estava ativo para o contribuinte no respectivo período da Nota Fiscal. (Aplicável somente para NFe)"),
            ("Código do produto", "Código do produto declarado pelo contribuinte para a operação. Válido apenas para Cupons Fiscais"),
            ("Cód. Tot. Par", "Código totalizador. Informação presente apenas nas operações ECF."),
            ("Alíquota Destacada", "Alíquota de ICMS destacada no documento fiscal pelo contribuinte"),
            ("ICMS Destacado", "ICMS destacado no documento fiscal pelo contribuinte"),
            ("Valor da operação", "Valor da Base de Cálculo calculada pelo fisco, sem considerar reduções da base de cálculo. As reduções da BC serão aplicadas na alíquota efetiva correta. Para as notas fiscais (NF-e e NFC-e inclui frete, seguro, despesas adicionais, descontado os descontos concedidos). Para os Cupons leva-se em conta apenas o valor declarado na EFD que é o valor efetivo da operação."),
            ("Alíquota Efetiva Correta (FISCO)", "Alíquota de ICMS considerada pelo fisco para a operação. Aqui considerando eventuais reduções da Base de Cálculo. Para os Cupons fiscais é a alíquota retirada do COD TOT PAR."),
            ("Alíquota Efetiva destacada pelo Contribuinte", "Alíquota efetiva destacada pelo Contribuinte, que é calculada dividindo o ICMS destacado pelo Valor da Operação sem considerar redução da base de cálculo"),
            ("ICMS devido", "Valor do ICMS considerado como correto pelo fisco."),
            ("ICMS não-recolhido", "Valor do ICMS a ser recolhido como diferença pelo contribuinte. Trata-se da dedução do valor de 'ICMS devido' pelo valor do campo 'ICMS destacado'")
        ]
    }
}

# Grupo padrão
GRUPO_PADRAO = "GESSUPER"


# =============================================================================
# 1. CONFIGURAÇÕES INICIAIS
# =============================================================================

# Hack SSL
try:
    createunverified_https_context = ssl._create_unverified_context
except AttributeError:
    pass
else:
    ssl._create_default_https_context = createunverified_https_context

warnings.filterwarnings('ignore')

# Configuração da página
st.set_page_config(
    page_title="ARGOS - Infrações GESSUPER",
    page_icon="🎯",
    layout="wide",
    initial_sidebar_state="expanded"
)

# CSS customizado
st.markdown("""
<style>
    .main-header {
        font-size: 2.5rem;
        font-weight: bold;
        color: #1565C0;
        text-align: center;
        padding: 1rem 0;
    }
    
    /* ESTILO DOS KPIs - BORDA PRETA */
    div[data-testid="stMetric"] {
        background-color: #ffffff;
        border: 2px solid #2c3e50;
        border-radius: 10px;
        padding: 15px;
        box-shadow: 0 2px 4px rgba(0,0,0,0.1);
    }
    
    div[data-testid="stMetric"] > label {
        font-weight: 600;
        color: #2c3e50;
    }
    
    div[data-testid="stMetricValue"] {
        font-size: 1.8rem;
        font-weight: bold;
        color: #1f77b4;
    }
    
    .card-kpi {
        background: white;
        padding: 1.5rem;
        border-radius: 10px;
        border: 2px solid #e1e5f0;
        box-shadow: 0 2px 4px rgba(0,0,0,0.05);
    }
    
    .card-kpi-baixa { border-left: 5px solid #f44336 !important; }  /* Vermelho (baixa confiança) */
    .card-kpi-media { border-left: 5px solid #FF9800 !important; }
    .card-kpi-alta { border-left: 5px solid #4CAF50 !important; }   /* Verde (alta confiança) */
    
    .status-badge {
        padding: 0.25rem 0.75rem;
        border-radius: 20px;
        font-size: 0.8rem;
        font-weight: 600;
    }
    
    .badge-success { background-color: #e8f5e9; color: #2e7d32; }
    .badge-warning { background-color: #fff3e0; color: #ef6c00; }
    .badge-danger { background-color: #ffebee; color: #c62828; }
    
    .info-box {
        background-color: #e3f2fd;
        border-left: 4px solid #1976d2;
        padding: 1rem;
        border-radius: 0 8px 8px 0;
        margin: 1rem 0;
    }
    
    .stTabs [data-baseweb="tab-list"] {
        gap: 24px;
    }
    
    .stTabs [data-baseweb="tab"] {
        height: 50px;
        padding-left: 20px;
        padding-right: 20px;
    }
</style>
""", unsafe_allow_html=True)

# =============================================================================
# 2. CREDENCIAIS E CONEXÃO
# =============================================================================

IMPALA_HOST = 'bdaworkernode02.sef.sc.gov.br'
IMPALA_PORT = 21050
DATABASE = 'niat'

try:
    IMPALA_USER = st.secrets["impala_credentials"]["user"]
    IMPALA_PASSWORD = st.secrets["impala_credentials"]["password"]
except:
    st.error("⚠️ Credenciais não configuradas. Configure secrets.toml")
    st.info("""
    Crie o arquivo `.streamlit/secrets.toml` com:
    ```
    [impala_credentials]
    user = "seu_usuario"
    password = "sua_senha"
    ```
    """)
    st.stop()

# =============================================================================
# 3. FUNÇÕES AUXILIARES
# =============================================================================

def is_table_unavailable_error(error_msg: str) -> bool:
    """
    Verifica se o erro é relacionado a tabela indisponível/em atualização.
    Retorna True se for erro de tabela, False caso contrário.
    """
    error_lower = str(error_msg).lower()
    table_error_patterns = [
        "could not resolve table reference",
        "table not found",
        "does not exist",
        "analysisexception",
        "no such table",
        "invalid table",
        "table or view not found",
        "relation.*does not exist",
        "unknown table"
    ]
    return any(pattern in error_lower for pattern in table_error_patterns)

TABLE_UNAVAILABLE_MSG = "⚠️ **Tabelas em atualização.** Favor tentar novamente mais tarde."


def get_grupo_config(grupo: str = None) -> dict:
    """
    Retorna a configuração do grupo especificado.
    Se não especificado, usa o grupo do session_state ou o padrão.
    """
    if grupo is None:
        grupo = st.session_state.get('grupo_selecionado', GRUPO_PADRAO)
    return GRUPOS_CONFIG.get(grupo, GRUPOS_CONFIG[GRUPO_PADRAO])


def get_grupo_tabelas(grupo: str = None) -> dict:
    """
    Retorna o dicionário de tabelas para o grupo especificado.
    """
    config = get_grupo_config(grupo)
    return config.get('tabelas', {})


def check_tables_available(engine, grupo: str = None) -> bool:
    """
    Verifica se as tabelas principais estão disponíveis para o grupo especificado.
    Faz uma query simples (LIMIT 1) para testar a conexão.
    Retorna True se disponíveis, False se indisponíveis.
    """
    tabelas = get_grupo_tabelas(grupo)
    tabela_teste = tabelas.get('nfce') or tabelas.get('cupons') or tabelas.get('nfe')

    if not tabela_teste:
        return False

    try:
        query = f"SELECT 1 FROM {tabela_teste} LIMIT 1"
        pd.read_sql(query, engine)
        return True
    except Exception as e:
        error_msg = str(e)
        if is_table_unavailable_error(error_msg):
            return False
        # Outros erros (conexão, etc) - assumir disponível e deixar erro aparecer depois
        return True


def sanitize_identificador(raw: str) -> str:
    """Remove qualquer coisa que não seja dígito (CNPJ/IE)."""
    return re.sub(r"\D+", "", raw or "")

def format_currency_br(value) -> str:
    """Formata número como moeda brasileira: R$ 1.234.567,89"""
    if value is None:
        return "R$ 0,00"
    try:
        v = float(value)
    except (TypeError, ValueError):
        return "R$ 0,00"
    s = f"{v:,.2f}"
    s = s.replace(",", "X").replace(".", ",").replace("X", ".")
    return f"R$ {s}"

def format_number_br(value) -> str:
    """Formata número brasileiro: 1.234.567"""
    if value is None:
        return "0"
    try:
        v = int(value)
    except (TypeError, ValueError):
        return "0"
    return f"{v:,}".replace(",", ".")

def nivel_config(nivel_str: str):
    """
    Retorna mapeamento de colunas para o nível escolhido.
    nivel_str: 'BAIXA' | 'MEDIA' | 'ALTA'
    
    Cores invertidas:
    - ALTA = Verde (maior confiança)
    - BAIXA = Vermelho (menor confiança)
    """
    nivel = (nivel_str or "").upper()
    if nivel == "MEDIA":
        return {
            "nivel": "MEDIA",
            "label": "MÉDIA",
            "col_aliquota": "aliquota_media",
            "col_legislacao": "legislacao_media",
            "col_infracao": "infracao_media",
            "cor": "#FF9800",
            "emoji": "🟡"
        }
    elif nivel == "ALTA":
        return {
            "nivel": "ALTA",
            "label": "ALTA",
            "col_aliquota": "aliquota_alta",
            "col_legislacao": "legislacao_alta",
            "col_infracao": "infracao_alta",
            "cor": "#4CAF50",  # Verde (alta confiança)
            "emoji": "🟢"
        }
    else:
        return {
            "nivel": "BAIXA",
            "label": "BAIXA",
            "col_aliquota": "aliquota_baixa",
            "col_legislacao": "legislacao_baixa",
            "col_infracao": "infracao_baixa",
            "cor": "#f44336",  # Vermelho (baixa confiança)
            "emoji": "🔴"
        }

# =============================================================================
# 4. CONEXÃO COM BANCO DE DADOS
# =============================================================================

@st.cache_resource
def get_engine():
    """Cria engine de conexão (compartilhada entre sessões)."""
    try:
        engine = create_engine(
            f'impala://{IMPALA_HOST}:{IMPALA_PORT}/{DATABASE}',
            connect_args={
                'user': IMPALA_USER,
                'password': IMPALA_PASSWORD,
                'auth_mechanism': 'LDAP',
                'use_ssl': True
            }
        )
        return engine
    except Exception as e:
        st.error(f"❌ Erro de conexão: {str(e)[:100]}")
        return None

# =============================================================================
# 5. FUNÇÕES DE CARREGAMENTO DE DADOS
# =============================================================================

@st.cache_data(ttl=CACHE_TTL_SECONDS, show_spinner="Buscando dados do contribuinte...")
def get_contribuinte_info(_engine, identificador_digits: str):
    """
    Busca informações do contribuinte na tabela usr_sat_ods.vw_ods_contrib.
    Retorna CNPJ formatado e Razão Social.
    """
    # Tenta buscar por CNPJ primeiro
    query_cnpj = f"""
        SELECT 
            nu_cnpj,
            nm_razao_social,
            nu_ie,
            nm_fantasia,
            nm_munic,
            cd_gerfe,
            nm_gerfe
        FROM usr_sat_ods.vw_ods_contrib
        WHERE regexp_replace(nu_cnpj, '[^0-9]', '') = '{identificador_digits}'
        LIMIT 1
    """
    
    # Tenta buscar por IE
    query_ie = f"""
        SELECT 
            nu_cnpj,
            nm_razao_social,
            nu_ie,
            nm_fantasia,
            nm_munic,
            cd_gerfe,
            nm_gerfe
        FROM usr_sat_ods.vw_ods_contrib
        WHERE regexp_replace(nu_ie, '[^0-9]', '') = '{identificador_digits}'
        LIMIT 1
    """
    
    try:
        # Tenta por CNPJ
        df = pd.read_sql(query_cnpj, _engine)
        if df.empty:
            # Tenta por IE
            df = pd.read_sql(query_ie, _engine)
        
        if not df.empty:
            return {
                'cnpj': df['nu_cnpj'].iloc[0] if pd.notna(df['nu_cnpj'].iloc[0]) else '',
                'razao_social': df['nm_razao_social'].iloc[0] if pd.notna(df['nm_razao_social'].iloc[0]) else '',
                'ie': df['nu_ie'].iloc[0] if pd.notna(df['nu_ie'].iloc[0]) else '',
                'fantasia': df['nm_fantasia'].iloc[0] if pd.notna(df['nm_fantasia'].iloc[0]) else '',
                'municipio': df['nm_munic'].iloc[0] if pd.notna(df['nm_munic'].iloc[0]) else '',
                'gerfe': df['nm_gerfe'].iloc[0] if pd.notna(df['nm_gerfe'].iloc[0]) else ''
            }
        return None
    except Exception as e:
        error_msg = str(e)
        if is_table_unavailable_error(error_msg):
            st.session_state.tabela_indisponivel = True
        return None

@st.cache_data(ttl=86400, show_spinner=False)  # Cache de 24h para tabelas de referência
def get_ncm_descricoes(_engine, ncm_list: list) -> dict:
    """
    Busca descrições dos NCMs na tabela niat.tabela_ncm.
    Retorna dict: {ncm: descricao}
    """
    if not ncm_list:
        return {}
    
    try:
        # Limpa e formata lista de NCMs
        ncm_clean = [str(n).strip() for n in ncm_list if pd.notna(n) and str(n).strip()]
        if not ncm_clean:
            return {}
        
        ncm_str = "', '".join(ncm_clean)
        query = f"""
            SELECT ncm, descricao
            FROM niat.tabela_ncm
            WHERE ncm IN ('{ncm_str}')
        """
        df = pd.read_sql(query, _engine)
        return dict(zip(df['ncm'].astype(str), df['descricao']))
    except Exception as e:
        return {}

@st.cache_data(ttl=86400, show_spinner=False)  # Cache de 24h para tabelas de referência
def get_cfop_descricoes(_engine, cfop_list: list) -> dict:
    """
    Busca descrições dos CFOPs na tabela niat.tabela_cfop.
    Retorna dict: {cfop: descricaocfop}
    """
    if not cfop_list:
        return {}
    
    try:
        # Limpa e formata lista de CFOPs
        cfop_clean = [str(c).strip() for c in cfop_list if pd.notna(c) and str(c).strip()]
        if not cfop_clean:
            return {}
        
        cfop_str = "', '".join(cfop_clean)
        query = f"""
            SELECT cfop, descricaocfop
            FROM niat.tabela_cfop
            WHERE cfop IN ('{cfop_str}')
        """
        df = pd.read_sql(query, _engine)
        return dict(zip(df['cfop'].astype(str), df['descricaocfop']))
    except Exception as e:
        return {}

@st.cache_data(ttl=CACHE_TTL_SECONDS, show_spinner="Carregando dados base...")
def get_base_df(_engine, identificador_digits: str, nivel: str = "BAIXA", grupo: str = None, tipo_doc_filter: str = None):
    """
    Carrega o DataFrame base para o CNPJ/IE informado.
    Suporta múltiplos grupos (GESSUPER, GESMAC, etc.)

    Args:
        _engine: Engine de conexão
        identificador_digits: CNPJ ou IE (apenas dígitos)
        nivel: Nível de acurácia (BAIXA, MEDIA, ALTA)
        grupo: Grupo (GESSUPER, GESMAC). Se None, usa session_state
        tipo_doc_filter: Filtro opcional por tipo de documento ('NFe', 'NFCe', 'Cupom', None=todos)

    As colunas são renomeadas para nomes genéricos:
        - legislacao_X -> legislacao_ia
        - aliquota_X -> aliquota_ia
        - infracao_X -> infracao_ia

    Filtra apenas registros onde infracao_X IS NOT NULL e != 'EXCLUIR'
    """
    # Obtém configuração do grupo
    if grupo is None:
        grupo = st.session_state.get('grupo_selecionado', GRUPO_PADRAO)

    tabelas = get_grupo_tabelas(grupo)

    # Define as colunas baseado no nível
    nivel_upper = (nivel or "BAIXA").upper()

    if nivel_upper == "ALTA":
        col_legislacao = "legislacao_alta"
        col_aliquota = "aliquota_alta"
        col_infracao = "infracao_alta"
    elif nivel_upper == "MEDIA":
        col_legislacao = "legislacao_media"
        col_aliquota = "aliquota_media"
        col_infracao = "infracao_media"
    else:  # BAIXA (padrão)
        col_legislacao = "legislacao_baixa"
        col_aliquota = "aliquota_baixa"
        col_infracao = "infracao_baixa"

    # Filtro para excluir registros NULL ou EXCLUIR (infracao, aliquota e legislacao)
    filtro_nivel = f"""
        {col_infracao} IS NOT NULL
        AND CAST({col_infracao} AS STRING) != 'EXCLUIR'
        AND CAST({col_aliquota} AS STRING) != 'EXCLUIR'
        AND CAST({col_legislacao} AS STRING) != 'EXCLUIR'
    """

    queries = []

    # Query NFCe (comum a GESSUPER e GESMAC)
    if tabelas.get('nfce') and (tipo_doc_filter is None or tipo_doc_filter == 'NFCe'):
        if grupo == "GESMAC":
            query_nfce = f"""
                SELECT
                    data_emissao,
                    periodo,
                    tipo_doc,
                    chave,
                    NULL AS link_acesso,
                    NULL AS modelo_ecf,
                    entrada_ou_saida,
                    ie_emitente,
                    cnpj_emitente,
                    razao_emitente,
                    ie_destinatario,
                    cnpj_destinatario,
                    NULL AS cpf_destinatario,
                    razao_destinatario,
                    estado_destinatario,
                    NULL AS uf_entrega,
                    numero_nota,
                    numero_item,
                    origem_prod,
                    NULL AS ind_final,
                    NULL AS cod_prod,
                    gtin,
                    ncm,
                    descricao,
                    CAST(cfop AS STRING) AS cfop,
                    cst,
                    valor_total,
                    valor_do_frete,
                    valor_do_seguro,
                    valor_outras_despesas,
                    valor_do_desconto,
                    NULL AS cod_tot_par,
                    aliquota_emitente AS icms_emitente,
                    icms_emitente AS icms_destacado,
                    NULL AS regime_destinatario,
                    cnae_destinatario,
                    NULL AS ttd_importacao,
                    bc_fisco,
                    {col_legislacao} AS legislacao_ia,
                    {col_aliquota} AS aliquota_ia,
                    NULL AS aliq_efetiva,
                    NULL AS icms_devido,
                    {col_infracao} AS infracao_ia
                FROM {tabelas['nfce']}
                WHERE regexp_replace(cnpj_emitente, '[^0-9]', '') = '{identificador_digits}'
                AND {filtro_nivel}
            """
        else:  # GESSUPER
            query_nfce = f"""
                SELECT
                    data_emissao,
                    periodo,
                    tipo_doc,
                    chave,
                    NULL AS link_acesso,
                    NULL AS modelo_ecf,
                    entrada_ou_saida,
                    NULL AS ie_emitente,
                    cnpj_emitente,
                    razao_emitente,
                    NULL AS ie_destinatario,
                    NULL AS cnpj_destinatario,
                    NULL AS cpf_destinatario,
                    NULL AS razao_destinatario,
                    NULL AS estado_destinatario,
                    NULL AS uf_entrega,
                    numero_nota,
                    CAST(numero_item AS STRING) AS numero_item,
                    NULL AS origem_prod,
                    NULL AS ind_final,
                    NULL AS cod_prod,
                    gtin,
                    ncm,
                    descricao,
                    CAST(cfop AS STRING) AS cfop,
                    NULL AS cst,
                    NULL AS valor_total,
                    NULL AS valor_do_frete,
                    NULL AS valor_do_seguro,
                    NULL AS valor_outras_despesas,
                    NULL AS valor_do_desconto,
                    NULL AS cod_tot_par,
                    icms_emitente,
                    NULL AS icms_destacado,
                    NULL AS regime_destinatario,
                    NULL AS cnae_destinatario,
                    NULL AS ttd_importacao,
                    bc_fisco,
                    {col_legislacao} AS legislacao_ia,
                    {col_aliquota} AS aliquota_ia,
                    NULL AS aliq_efetiva,
                    NULL AS icms_devido,
                    {col_infracao} AS infracao_ia
                FROM {tabelas['nfce']}
                WHERE regexp_replace(cnpj_emitente, '[^0-9]', '') = '{identificador_digits}'
                AND {filtro_nivel}
            """
        queries.append(query_nfce)

    # Query Cupons (comum a GESSUPER e GESMAC)
    if tabelas.get('cupons') and (tipo_doc_filter is None or tipo_doc_filter == 'Cupom'):
        if grupo == "GESMAC":
            query_cupons = f"""
                SELECT
                    data_emissao,
                    periodo,
                    tipo_doc,
                    NULL AS chave,
                    NULL AS link_acesso,
                    modelo_ecf,
                    NULL AS entrada_ou_saida,
                    ie_emitente,
                    cnpj_emitente,
                    razao_emitente,
                    NULL AS ie_destinatario,
                    NULL AS cnpj_destinatario,
                    NULL AS cpf_destinatario,
                    NULL AS razao_destinatario,
                    NULL AS estado_destinatario,
                    NULL AS uf_entrega,
                    NULL AS numero_nota,
                    NULL AS numero_item,
                    NULL AS origem_prod,
                    NULL AS ind_final,
                    cod_prod,
                    gtin,
                    ncm,
                    descricao,
                    CAST(cfop AS STRING) AS cfop,
                    NULL AS cst,
                    bc_fisco AS valor_total,
                    NULL AS valor_do_frete,
                    NULL AS valor_do_seguro,
                    NULL AS valor_outras_despesas,
                    NULL AS valor_do_desconto,
                    cod_tot_par,
                    aliquota_emitente AS icms_emitente,
                    icms_emitente AS icms_destacado,
                    NULL AS regime_destinatario,
                    NULL AS cnae_destinatario,
                    NULL AS ttd_importacao,
                    bc_fisco,
                    {col_legislacao} AS legislacao_ia,
                    {col_aliquota} AS aliquota_ia,
                    NULL AS aliq_efetiva,
                    NULL AS icms_devido,
                    {col_infracao} AS infracao_ia
                FROM {tabelas['cupons']}
                WHERE regexp_replace(cnpj_emitente, '[^0-9]', '') = '{identificador_digits}'
                AND {filtro_nivel}
            """
        else:  # GESSUPER
            query_cupons = f"""
                SELECT
                    data_emissao,
                    periodo,
                    tipo_doc,
                    NULL AS chave,
                    NULL AS link_acesso,
                    modelo_ecf,
                    NULL AS entrada_ou_saida,
                    NULL AS ie_emitente,
                    cnpj_emitente,
                    razao_emitente,
                    NULL AS ie_destinatario,
                    NULL AS cnpj_destinatario,
                    NULL AS cpf_destinatario,
                    NULL AS razao_destinatario,
                    NULL AS estado_destinatario,
                    NULL AS uf_entrega,
                    NULL AS numero_nota,
                    CAST(NULL AS STRING) AS numero_item,
                    NULL AS origem_prod,
                    NULL AS ind_final,
                    cod_prod,
                    gtin,
                    ncm,
                    descricao,
                    CAST(cfop AS STRING) AS cfop,
                    NULL AS cst,
                    NULL AS valor_total,
                    NULL AS valor_do_frete,
                    NULL AS valor_do_seguro,
                    NULL AS valor_outras_despesas,
                    NULL AS valor_do_desconto,
                    cod_tot_par,
                    icms_emitente,
                    NULL AS icms_destacado,
                    NULL AS regime_destinatario,
                    NULL AS cnae_destinatario,
                    NULL AS ttd_importacao,
                    bc_fisco,
                    {col_legislacao} AS legislacao_ia,
                    {col_aliquota} AS aliquota_ia,
                    NULL AS aliq_efetiva,
                    NULL AS icms_devido,
                    {col_infracao} AS infracao_ia
                FROM {tabelas['cupons']}
                WHERE regexp_replace(cnpj_emitente, '[^0-9]', '') = '{identificador_digits}'
                AND {filtro_nivel}
            """
        queries.append(query_cupons)

    # Query NFe (apenas GESMAC)
    if tabelas.get('nfe') and (tipo_doc_filter is None or tipo_doc_filter == 'NFe'):
        query_nfe = f"""
            SELECT
                data_emissao,
                periodo,
                tipo_doc,
                chave,
                NULL AS link_acesso,
                NULL AS modelo_ecf,
                entrada_ou_saida,
                ie_emitente,
                cnpj_emitente,
                razao_emitente,
                ie_destinatario,
                cnpj_destinatario,
                NULL AS cpf_destinatario,
                razao_destinatario,
                estado_destinatario,
                uf_entrega,
                numero_nota,
                numero_item,
                origem_prod,
                ind_final,
                NULL AS cod_prod,
                gtin,
                ncm,
                descricao,
                CAST(cfop AS STRING) AS cfop,
                cst,
                valor_total,
                valor_do_frete,
                valor_do_seguro,
                valor_outras_despesas,
                valor_do_desconto,
                NULL AS cod_tot_par,
                aliquota_emitente AS icms_emitente,
                icms_emitente AS icms_destacado,
                regime_destinatario,
                cnae_destinatario,
                ttd_importacao,
                bc_fisco_red AS bc_fisco,
                {col_legislacao} AS legislacao_ia,
                {col_aliquota} AS aliquota_ia,
                NULL AS aliq_efetiva,
                NULL AS icms_devido,
                {col_infracao} AS infracao_ia
            FROM {tabelas['nfe']}
            WHERE regexp_replace(cnpj_emitente, '[^0-9]', '') = '{identificador_digits}'
            AND {filtro_nivel}
        """
        queries.append(query_nfe)

    if not queries:
        return pd.DataFrame()

    # Combina as queries com UNION ALL
    full_query = " UNION ALL ".join(queries)

    try:
        df = pd.read_sql(full_query, _engine)
        return df
    except Exception as e:
        error_msg = str(e)
        if is_table_unavailable_error(error_msg):
            st.session_state.tabela_indisponivel = True
        return pd.DataFrame()

def calcular_totais(df: pd.DataFrame, nivel_str: str):
    """
    Retorna:
        total_nivel, cfg (dict do nível), has_rows (bool)
    
    A query SQL já traz os dados do nível selecionado com colunas genéricas:
        - legislacao_ia, aliquota_ia, infracao_ia
    
    Infração = valor da coluna infracao_ia (ICMS devido pela IA)
    """
    cfg = nivel_config(nivel_str)
    
    if df.empty:
        return 0.0, cfg, False
    
    # Verifica qual coluna de infração usar
    # Nova estrutura: infracao_ia (coluna genérica)
    # Estrutura antiga: infracao_baixa, infracao_media, infracao_alta
    if 'infracao_ia' in df.columns:
        col_infracao = 'infracao_ia'
    else:
        # Fallback para estrutura antiga
        col_infracao = cfg['col_infracao']
    
    # Converte valores para numérico e soma
    # Usa COALESCE equivalente: converte para float, trata NaN como 0
    df_calc = df.copy()
    df_calc['infracao_valor'] = pd.to_numeric(df_calc[col_infracao], errors='coerce').fillna(0)
    
    # Soma os valores das infrações
    total_nivel = df_calc['infracao_valor'].sum()
    
    return float(total_nivel), cfg, True

def build_export_df(df: pd.DataFrame, nivel_str: str, grupo: str = None, modelo_export: str = None):
    """
    Monta o DataFrame pronto para exportar.
    A query SQL já traz as colunas renomeadas para nomes genéricos:
        - legislacao_ia, aliquota_ia, infracao_ia

    Este método apenas renomeia para o formato final do Excel.

    Args:
        df: DataFrame com os dados
        nivel_str: Nível de acurácia
        grupo: Grupo (GESSUPER, GESMAC). Se None, usa session_state
        modelo_export: Modelo de exportação para GESMAC ('NFe' ou 'NFCe + Cupom Fiscal')
    """
    cfg = nivel_config(nivel_str)

    if grupo is None:
        grupo = st.session_state.get('grupo_selecionado', GRUPO_PADRAO)

    if df.empty:
        return None

    # Copia o DataFrame
    df_export = df.copy()

    # Filtra por tipo de documento se modelo específico for selecionado (GESMAC)
    if grupo == "GESMAC" and modelo_export:
        if modelo_export == "NFe":
            # Filtra apenas NFe
            df_export = df_export[df_export['tipo_doc'].str.upper().str.contains('NFE', na=False) &
                                  ~df_export['tipo_doc'].str.upper().str.contains('NFCE', na=False)]
        elif modelo_export == "NFCe + Cupom Fiscal":
            # Filtra NFCe e Cupom
            df_export = df_export[df_export['tipo_doc'].str.upper().str.contains('NFCE|ECF|CUPOM', regex=True, na=False) |
                                  ~df_export['tipo_doc'].str.upper().str.contains('NFE', na=False)]

    if df_export.empty:
        return None

    # Verifica qual estrutura de colunas usar (nova ou antiga)
    if 'infracao_ia' in df_export.columns:
        # Nova estrutura com colunas genéricas
        df_export['legislacao_ia_icms'] = df_export['legislacao_ia']
        df_export['aliquota_ia_icms'] = df_export['aliquota_ia']
        df_export['icms_devido'] = pd.to_numeric(df_export['infracao_ia'], errors='coerce').fillna(0)
    else:
        # Estrutura antiga com colunas por nível
        col_legislacao = cfg['col_legislacao']
        col_aliquota = cfg['col_aliquota']
        col_infracao = cfg['col_infracao']
        df_export['legislacao_ia_icms'] = df_export[col_legislacao]
        df_export['aliquota_ia_icms'] = df_export[col_aliquota]
        df_export['icms_devido'] = pd.to_numeric(df_export[col_infracao], errors='coerce').fillna(0)

    # Calcula ICMS não recolhido (ICMS devido - ICMS destacado)
    if 'icms_destacado' in df_export.columns:
        icms_destacado = pd.to_numeric(df_export['icms_destacado'], errors='coerce').fillna(0)
    else:
        icms_destacado = pd.to_numeric(df_export['icms_emitente'], errors='coerce').fillna(0)

    df_export['icms_nao_recolhido'] = df_export['icms_devido'] - icms_destacado
    df_export['icms_nao_recolhido'] = df_export['icms_nao_recolhido'].clip(lower=0)  # Não pode ser negativo

    # Define colunas de exportação baseado no grupo
    if grupo == "GESMAC":
        # Colunas estendidas para GESMAC (NFe e NFCe + Cupom têm estrutura similar)
        colunas_export = [
            "data_emissao", "periodo", "tipo_doc", "chave", "link_acesso",
            "modelo_ecf", "entrada_ou_saida", "ie_emitente", "cnpj_emitente",
            "razao_emitente", "ie_destinatario", "cnpj_destinatario",
            "cpf_destinatario", "razao_destinatario", "estado_destinatario",
            "regime_destinatario", "cnae_destinatario", "numero_nota", "numero_item",
            "origem_prod", "ind_final", "ttd_importacao", "gtin", "ncm", "descricao",
            "cfop", "cod_prod", "valor_total", "valor_do_frete", "valor_do_seguro",
            "valor_outras_despesas", "valor_do_desconto", "cod_tot_par",
            "icms_emitente", "icms_destacado", "bc_fisco", "aliquota_ia_icms",
            "legislacao_ia_icms", "aliq_efetiva", "icms_devido", "icms_nao_recolhido"
        ]
    else:
        # Colunas padrão para GESSUPER
        colunas_export = [
            "data_emissao", "periodo", "tipo_doc", "chave", "link_acesso",
            "modelo_ecf", "entrada_ou_saida", "cnpj_emitente", "razao_emitente",
            "numero_nota", "gtin", "ncm", "numero_item", "descricao", "cfop",
            "icms_emitente", "cod_prod", "cod_tot_par", "legislacao_ia_icms",
            "bc_fisco", "aliquota_ia_icms", "aliq_efetiva", "icms_devido"
        ]

    # Filtra apenas colunas que existem no DataFrame
    colunas_existentes = [col for col in colunas_export if col in df_export.columns]

    return df_export[colunas_existentes]

# =============================================================================
# 6. FUNÇÕES DE EXPORTAÇÃO
# =============================================================================

def export_to_csv(df: pd.DataFrame, identificador: str, nivel: str) -> bytes:
    """
    Exporta DataFrame para CSV no formato brasileiro.
    - Separador: ponto e vírgula (;)
    - Encoding: latin-1 (ANSI)
    - Decimal: vírgula (,)
    """
    csv_str = df.to_csv(index=False, sep=";", decimal=",")
    return csv_str.encode("latin-1", errors="replace")

def export_to_excel_template(df: pd.DataFrame, contrib_info: dict, nivel: str, parte_atual: int = None, total_partes: int = None, progress_callback=None) -> bytes:
    """
    Exporta DataFrame para Excel usando a estrutura do template Anexo J.
    Inclui fórmulas para recálculos automáticos na aba J2.
    
    Args:
        df: DataFrame com os dados
        contrib_info: Informações do contribuinte
        nivel: Nível de acurácia (BAIXA, MEDIA, ALTA)
        parte_atual: Número da parte atual (se dividido)
        total_partes: Total de partes (se dividido)
        progress_callback: Função callback(percentual, mensagem) para reportar progresso
    """
    def report_progress(pct, msg):
        if progress_callback:
            progress_callback(pct, msg)
    
    report_progress(5, "Criando estrutura do arquivo")
    
    buffer = BytesIO()
    
    # Cria workbook
    wb = Workbook()
    
    # =========================================================================
    # ABA 1: ANEXO J1 - NOTAS DE SAÍDAS (dados detalhados)
    # =========================================================================
    ws1 = wb.active
    ws1.title = "ANEXO J1 - NOTAS DE SAÍDAS"
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
    header_fill_fisco = PatternFill(start_color="C62828", end_color="C62828", fill_type="solid")
    title_font = Font(bold=True, size=14, color="1565C0")
    subtitle_font = Font(bold=True, size=11, color="666666")
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Linha 1: Título (inclui parte se aplicável)
    titulo_j1 = "ANEXO J1"
    if parte_atual is not None and total_partes is not None:
        titulo_j1 = f"ANEXO J1 - Parte {parte_atual} de {total_partes}"
    ws1['A1'] = titulo_j1
    ws1['A1'].font = title_font
    
    # Linha 2: Subtítulos das seções
    ws1['D2'] = "INFORMAÇÕES RETIRADAS DOS DOCUMENTOS FISCAIS (Cupons Fiscais ou NFC-e)"
    ws1['D2'].font = subtitle_font
    ws1['S2'] = "INFORMAÇÕES DECLARADAS PELO FISCO"
    ws1['S2'].font = Font(bold=True, size=11, color="C62828")
    
    # Linha 3: Cabeçalhos
    headers_j1 = [
        "Data de emissão",      # A
        "Período",              # B
        "Tipo Documento",       # C
        "Chave de acesso",      # D
        "Link de Acesso",       # E
        "ECF-FAB",              # F
        "Entrada ou saída",     # G
        "CNPJ Emitente",        # H
        "Razão do Emitente",    # I
        "Número",               # J
        "GTIN",                 # K
        "NCM",                  # L
        "Item",                 # M
        "Descrição do produto", # N
        "CFOP",                 # O
        "ICMS destacado",       # P
        "Código do Produto",    # Q
        "Cód. Tot. Par",        # R
        "Legislação",           # S (FISCO)
        "Valor da Operação",    # T (FISCO)
        "Alíquota ICMS correta",# U (FISCO)
        "Alíquota ICMS efetiva",# V (FISCO)
        "ICMS devido",          # W (FISCO)
        "ICMS não-recolhido"    # X (FISCO)
    ]
    
    for col_idx, header in enumerate(headers_j1, 1):
        cell = ws1.cell(row=3, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill_fisco if col_idx >= 19 else header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
    
    # Altura da linha de cabeçalho
    ws1.row_dimensions[3].height = 30
    
    # Mapeamento de colunas do DataFrame para o template
    # Nota: link_acesso (E) será fórmula para NF-e/NFC-e
    column_mapping = {
        'data_emissao': 'A',
        'periodo': 'B',
        'tipo_doc': 'C',
        'chave': 'D',
        # 'link_acesso': 'E' - será fórmula HYPERLINK
        'modelo_ecf': 'F',
        'entrada_ou_saida': 'G',
        'cnpj_emitente': 'H',
        'razao_emitente': 'I',
        'numero_nota': 'J',
        'gtin': 'K',
        'ncm': 'L',
        'numero_item': 'M',
        'descricao': 'N',
        'cfop': 'O',
        'icms_emitente': 'P',
        'cod_prod': 'Q',
        'cod_tot_par': 'R',
        'legislacao_ia_icms': 'S',
        'bc_fisco': 'T',
        'aliquota_ia_icms': 'U'
        # Colunas V, W, X serão fórmulas
    }
    
    report_progress(10, "Preenchendo dados da aba J1")

    # Ordena dados por data_emissao (mais antigo primeiro) para ordem cronológica
    if 'data_emissao' in df.columns:
        df = df.sort_values('data_emissao', ascending=True, na_position='last').reset_index(drop=True)

    # Preenche os dados a partir da linha 4
    total_rows = len(df)
    progress_interval = max(1, total_rows // 20)  # Atualiza a cada 5%
    
    for row_idx, row_data in enumerate(df.itertuples(index=False), 4):
        # Reporta progresso a cada 5%
        atual_row = row_idx - 4
        if atual_row % progress_interval == 0:
            pct = 10 + int((atual_row / total_rows) * 50)  # 10% a 60%
            report_progress(pct, f"Processando linha {atual_row:,} de {total_rows:,}")
        
        row_dict = row_data._asdict()
        
        for col_name, col_letter in column_mapping.items():
            if col_name in row_dict:
                col_idx = ord(col_letter) - ord('A') + 1
                cell = ws1.cell(row=row_idx, column=col_idx)
                value = row_dict[col_name]
                
                # Converte datas - dayfirst=True para interpretar DD/MM/YYYY corretamente
                if col_name == 'data_emissao' and pd.notna(value):
                    try:
                        if isinstance(value, str):
                            cell.value = pd.to_datetime(value, dayfirst=True).date()
                        else:
                            cell.value = value
                        cell.number_format = 'DD/MM/YYYY'
                    except:
                        cell.value = value
                # Formata período como DD/MM/AAAA - dayfirst=True para interpretar corretamente
                elif col_name == 'periodo' and pd.notna(value):
                    try:
                        if isinstance(value, str):
                            cell.value = pd.to_datetime(value, dayfirst=True).date()
                        else:
                            cell.value = value
                        cell.number_format = 'DD/MM/YYYY'
                    except:
                        cell.value = value
                # Valores numéricos monetários
                elif col_name in ['icms_emitente', 'bc_fisco'] and pd.notna(value):
                    try:
                        cell.value = float(value)
                        cell.number_format = '#,##0.00'
                    except:
                        cell.value = value
                # Alíquota ICMS correta - dividir por 100 para formato percentual correto
                elif col_name == 'aliquota_ia_icms' and pd.notna(value):
                    try:
                        # Valor vem como 12 (12%), divide por 100 = 0.12, Excel mostra 12%
                        cell.value = float(value) / 100
                        cell.number_format = '0.00%'
                    except:
                        cell.value = value
                else:
                    cell.value = value if pd.notna(value) else ''
                
                cell.border = thin_border
        
        # Fórmula para Link de Acesso (coluna E) - HYPERLINK apenas se houver chave (NF-e/NFC-e)
        # Cupom Fiscal não tem chave, então ficará em branco
        cell_e = ws1.cell(row=row_idx, column=5)  # Coluna E
        cell_e.value = f'=IF(D{row_idx}<>"",HYPERLINK("https://sat.sef.sc.gov.br/tax.NET/Sat.NFe.Web/Consultas/Nfe_ResumoPDF.ashx?id="&D{row_idx},"Abrir DANFE"),"")'
        cell_e.border = thin_border
        
        # Fórmula para Alíquota ICMS efetiva (coluna V) = ICMS destacado / BC Fisco
        cell_v = ws1.cell(row=row_idx, column=22)  # Coluna V
        cell_v.value = f"=IF(T{row_idx}=0,0,P{row_idx}/T{row_idx})"
        cell_v.number_format = '0.00%'
        cell_v.border = thin_border
        
        # Fórmula para ICMS devido (coluna W) = BC Fisco * Alíquota ICMS correta
        cell_w = ws1.cell(row=row_idx, column=23)  # Coluna W
        cell_w.value = f"=T{row_idx}*U{row_idx}"
        cell_w.number_format = '#,##0.00'
        cell_w.border = thin_border
        
        # Fórmula para ICMS não-recolhido (coluna X) = ICMS devido - ICMS destacado
        cell_x = ws1.cell(row=row_idx, column=24)  # Coluna X
        cell_x.value = f"=W{row_idx}-P{row_idx}"
        cell_x.number_format = '#,##0.00'
        cell_x.border = thin_border
    
    # Autoajuste de largura das colunas baseado no conteúdo
    for col_idx in range(1, 25):  # Colunas A até X
        col_letter = get_column_letter(col_idx)
        max_length = 0
        
        # Verifica todas as linhas da coluna
        for row in ws1.iter_rows(min_row=1, max_row=ws1.max_row, min_col=col_idx, max_col=col_idx):
            for cell in row:
                try:
                    if cell.value:
                        # Para fórmulas, estima o tamanho do resultado
                        if str(cell.value).startswith('='):
                            cell_length = 12  # Tamanho estimado para resultados numéricos
                        else:
                            cell_length = len(str(cell.value))
                        max_length = max(max_length, cell_length)
                except:
                    pass
        
        # Define largura com mínimo de 8 e máximo de 50
        adjusted_width = min(max(max_length + 2, 8), 50)
        ws1.column_dimensions[col_letter].width = adjusted_width
    
    # Congela painéis (linha 4 em diante rola, cabeçalhos ficam fixos)
    ws1.freeze_panes = 'A4'
    
    # Adiciona filtro automático na linha de cabeçalhos (linha 3, colunas A até X)
    ultima_linha = 3 + len(df)  # Linha 3 = cabeçalhos, dados começam na linha 4
    ws1.auto_filter.ref = f"A3:X{ultima_linha}"
    
    report_progress(65, "Criando aba J2 - Resumo ICMS")
    
    # =========================================================================
    # ABA 2: ANEXO J2 - ICMS DEVIDO (resumo por período com fórmulas)
    # =========================================================================
    ws2 = wb.create_sheet("ANEXO J2 - ICMS DEVIDO")
    
    # Cabeçalho institucional
    ws2.merge_cells('A1:D1')
    ws2['A1'] = "ESTADO DE SANTA CATARINA"
    ws2['A1'].font = Font(bold=True, size=14)
    ws2['A1'].alignment = Alignment(horizontal="center")
    
    ws2.merge_cells('A2:D2')
    ws2['A2'] = "Secretaria de Estado da Fazenda"
    ws2['A2'].alignment = Alignment(horizontal="center")
    
    ws2.merge_cells('A3:D3')
    ws2['A3'] = "Diretoria de Administração Tributária"
    ws2['A3'].alignment = Alignment(horizontal="center")
    
    ws2.merge_cells('A4:D4')
    ws2['A4'] = "Gerência de Fiscalização"
    ws2['A4'].alignment = Alignment(horizontal="center")
    
    # Informações do contribuinte
    ws2['A6'] = "CNPJ:"
    ws2['A6'].font = Font(bold=True)
    ws2['B6'] = contrib_info.get('cnpj', '') if contrib_info else ''
    
    ws2['A7'] = "Razão Social:"
    ws2['A7'].font = Font(bold=True)
    ws2['B7'] = contrib_info.get('razao_social', '') if contrib_info else ''
    
    # Título da tabela
    ws2.merge_cells('A10:D10')
    ws2['A10'] = "APURAÇÃO MENSAL DO VALOR DO ICMS DEVIDO NAS VENDAS DE MERCADORIAS"
    ws2['A10'].font = Font(bold=True, size=12)
    ws2['A10'].alignment = Alignment(horizontal="center")
    
    # Cabeçalhos da tabela
    headers_j2 = ["Período", "ICMS destacado", "ICMS apurado", "ICMS não recolhido"]
    for col_idx, header in enumerate(headers_j2, 1):
        cell = ws2.cell(row=11, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
    
    # Obtém períodos únicos e ordena cronologicamente (mais antigo primeiro)
    if 'periodo' in df.columns:
        periodos_unicos = df['periodo'].dropna().unique()
        # Converte para datetime para ordenação correta, depois ordena
        periodos = sorted(periodos_unicos, key=lambda x: pd.to_datetime(x, dayfirst=True) if isinstance(x, str) else x)
    else:
        periodos = []
    
    # Dados por período com fórmulas SUMIF
    ultima_linha_dados = len(df) + 3  # Linha final dos dados na aba J1
    
    for row_idx, periodo in enumerate(periodos, 12):
        # Período - formata como DD/MM/YYYY com dayfirst=True para evitar inversão de data
        cell_a = ws2.cell(row=row_idx, column=1)
        try:
            if isinstance(periodo, str):
                cell_a.value = pd.to_datetime(periodo, dayfirst=True).date()
            else:
                cell_a.value = periodo
            cell_a.number_format = 'DD/MM/YYYY'
        except:
            cell_a.value = periodo
        cell_a.border = thin_border
        cell_a.alignment = Alignment(horizontal="center")
        
        # ICMS destacado (SUMIF na coluna P da aba J1 onde período = B)
        cell_b = ws2.cell(row=row_idx, column=2)
        cell_b.value = f"=SUMIF('ANEXO J1 - NOTAS DE SAÍDAS'!$B$4:$B${ultima_linha_dados},$A{row_idx},'ANEXO J1 - NOTAS DE SAÍDAS'!$P$4:$P${ultima_linha_dados})"
        cell_b.number_format = '#,##0.00'
        cell_b.border = thin_border
        
        # ICMS apurado/devido (SUMIF na coluna W da aba J1)
        cell_c = ws2.cell(row=row_idx, column=3)
        cell_c.value = f"=SUMIF('ANEXO J1 - NOTAS DE SAÍDAS'!$B$4:$B${ultima_linha_dados},$A{row_idx},'ANEXO J1 - NOTAS DE SAÍDAS'!$W$4:$W${ultima_linha_dados})"
        cell_c.number_format = '#,##0.00'
        cell_c.border = thin_border
        
        # ICMS não recolhido = ICMS apurado - ICMS destacado
        cell_d = ws2.cell(row=row_idx, column=4)
        cell_d.value = f"=C{row_idx}-B{row_idx}"
        cell_d.number_format = '#,##0.00'
        cell_d.border = thin_border
    
    # Linha de TOTAL
    total_row = 12 + len(periodos)
    ws2.cell(row=total_row, column=1).value = "TOTAL"
    ws2.cell(row=total_row, column=1).font = Font(bold=True)
    ws2.cell(row=total_row, column=1).border = thin_border
    
    for col in range(2, 5):
        cell = ws2.cell(row=total_row, column=col)
        cell.value = f"=SUM({chr(64+col)}12:{chr(64+col)}{total_row-1})"
        cell.number_format = '#,##0.00'
        cell.font = Font(bold=True)
        cell.border = thin_border
        cell.fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
    
    # Autoajuste de largura das colunas na aba J2
    for col_idx in range(1, 5):  # Colunas A até D
        col_letter = get_column_letter(col_idx)
        max_length = 0
        
        for row in ws2.iter_rows(min_row=1, max_row=ws2.max_row, min_col=col_idx, max_col=col_idx):
            for cell in row:
                try:
                    if cell.value:
                        if str(cell.value).startswith('='):
                            cell_length = 15  # Tamanho estimado para resultados numéricos/fórmulas
                        else:
                            cell_length = len(str(cell.value))
                        max_length = max(max_length, cell_length)
                except:
                    pass
        
        adjusted_width = min(max(max_length + 2, 12), 50)
        ws2.column_dimensions[col_letter].width = adjusted_width
    
    report_progress(80, "Criando aba Índice")
    
    # =========================================================================
    # ABA 3: Índice (descrição dos campos)
    # =========================================================================
    ws3 = wb.create_sheet("Índice")
    
    ws3['A1'] = "Campo"
    ws3['B1'] = "Descrição"
    ws3['A1'].font = header_font
    ws3['B1'].font = header_font
    ws3['A1'].fill = header_fill
    ws3['B1'].fill = header_fill
    
    indice_dados = [
        ("Chave de acesso", "Número da chave de acesso das Notas Fiscais. Não aplicável para ECF."),
        ("URL", "Link para acessar o documento fiscal (apenas Notas Fiscais)."),
        ("Tipo Documento", "Fonte da informação: NFe, NFCe ou ECF."),
        ("Data de emissão", "Data de emissão do documento (Cupom Fiscal: data da Redução Z)."),
        ("Entrada ou saída", "Indica se a operação é de entrada ou saída."),
        ("ECF-FAB", "Número de série do Emissor de Cupom Fiscal."),
        ("GTIN", "Código GTIN da mercadoria."),
        ("NCM", "Código NCM da mercadoria."),
        ("No. Nota", "Número da Nota Fiscal."),
        ("No. Item", "Número do item dentro da Nota Fiscal."),
        ("Código do produto", "Código do produto declarado pelo contribuinte."),
        ("Cód. Tot. Par", "Código totalizador (apenas ECF)."),
        ("ICMS Destacado", "ICMS destacado no documento fiscal pelo contribuinte."),
        ("Valor da operação", "Base de Cálculo calculada pelo fisco."),
        ("Alíquota Efetiva Correta", "Alíquota de ICMS considerada pelo fisco."),
        ("Alíquota Efetiva destacada", "Alíquota efetiva destacada pelo Contribuinte."),
        ("ICMS devido", "Valor do ICMS considerado como correto pelo fisco."),
        ("ICMS não-recolhido", "Diferença entre ICMS devido e ICMS destacado.")
    ]
    
    for row_idx, (campo, desc) in enumerate(indice_dados, 2):
        ws3.cell(row=row_idx, column=1).value = campo
        ws3.cell(row=row_idx, column=2).value = desc
    
    # Autoajuste de largura das colunas na aba Índice
    for col_idx in range(1, 3):  # Colunas A e B
        col_letter = get_column_letter(col_idx)
        max_length = 0
        
        for row in ws3.iter_rows(min_row=1, max_row=ws3.max_row, min_col=col_idx, max_col=col_idx):
            for cell in row:
                try:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        max_length = max(max_length, cell_length)
                except:
                    pass
        
        adjusted_width = min(max(max_length + 2, 10), 100)
        ws3.column_dimensions[col_letter].width = adjusted_width
    
    report_progress(90, "Configurando recálculo automático")
    
    # =========================================================================
    # FORÇA RECÁLCULO DE FÓRMULAS AO ABRIR O ARQUIVO
    # =========================================================================
    # Isso resolve o problema de fórmulas que aparecem em branco até o usuário
    # clicar na célula e pressionar Enter
    wb.calculation = CalcProperties(
        fullCalcOnLoad=True,  # Recálculo completo ao carregar
        calcMode='auto'       # Modo de cálculo automático
    )
    
    report_progress(95, "Salvando arquivo Excel")
    
    # Salva
    wb.save(buffer)
    buffer.seek(0)
    
    report_progress(100, "Concluído!")
    
    return buffer.getvalue()

def export_to_excel_or_zip(df: pd.DataFrame, contrib_info: dict, nivel: str, progress_callback=None) -> tuple:
    """
    Exporta DataFrame para Excel ou ZIP (se mais de 1 milhão de linhas).
    
    Args:
        df: DataFrame com os dados
        contrib_info: Informações do contribuinte
        nivel: Nível de acurácia
        progress_callback: Função callback(current, total, message) para atualizar progresso
    
    Returns:
        tuple: (bytes_data, filename, is_zip)
            - bytes_data: conteúdo do arquivo
            - filename: nome do arquivo sugerido
            - is_zip: True se for ZIP, False se for Excel único
    """
    total_rows = len(df)
    
    # Se cabe em um único arquivo Excel
    if total_rows <= MAX_ROWS_PER_EXCEL:
        if progress_callback:
            progress_callback(0, 1, "Gerando arquivo Excel...")
        excel_data = export_to_excel_template(df, contrib_info, nivel)
        if progress_callback:
            progress_callback(1, 1, "Arquivo Excel gerado!")
        filename = get_export_filename(contrib_info, nivel, "xlsx")
        return excel_data, filename, False
    
    # Precisa dividir em múltiplas partes
    total_partes = math.ceil(total_rows / MAX_ROWS_PER_EXCEL)
    
    # Cria buffer para o ZIP
    zip_buffer = BytesIO()
    
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
        for parte in range(1, total_partes + 1):
            if progress_callback:
                progress_callback(parte - 1, total_partes, f"Gerando parte {parte} de {total_partes}...")
            
            # Calcula índices de início e fim
            start_idx = (parte - 1) * MAX_ROWS_PER_EXCEL
            end_idx = min(parte * MAX_ROWS_PER_EXCEL, total_rows)
            
            # Extrai a parte do DataFrame
            df_parte = df.iloc[start_idx:end_idx].copy()
            
            # Gera o Excel para esta parte
            excel_data = export_to_excel_template(
                df_parte, 
                contrib_info, 
                nivel, 
                parte_atual=parte, 
                total_partes=total_partes
            )
            
            # Nome do arquivo da parte
            base_filename = get_export_filename(contrib_info, nivel, "xlsx")
            parte_filename = base_filename.replace(".xlsx", f" - Parte {parte} de {total_partes}.xlsx")
            
            # Adiciona ao ZIP
            zip_file.writestr(parte_filename, excel_data)
            
            if progress_callback:
                progress_callback(parte, total_partes, f"Parte {parte} de {total_partes} concluída!")
    
    zip_buffer.seek(0)
    
    # Nome do arquivo ZIP
    zip_filename = get_export_filename(contrib_info, nivel, "zip")
    
    return zip_buffer.getvalue(), zip_filename, True

def get_export_filename(contrib_info: dict, nivel: str, extension: str) -> str:
    """
    Gera o nome do arquivo no formato: CNPJ_14_DIGITOS - RAZAO_SOCIAL.extensao
    """
    if contrib_info:
        cnpj = sanitize_identificador(contrib_info.get('cnpj', ''))
        razao = contrib_info.get('razao_social', 'EMPRESA')
        # Limpa caracteres especiais da razão social
        razao_clean = re.sub(r'[<>:"/\\|?*]', '', razao)[:50]  # Limita a 50 chars
        return f"{cnpj} - {razao_clean}.{extension}"
    return f"infracoes_gessuper_{nivel.lower()}.{extension}"

def get_folder_link(path: str) -> str:
    """
    Gera um link para copiar o caminho da pasta.
    """
    return path

def save_to_network_fast(df: pd.DataFrame, contrib_info: dict, nivel: str, progress_callback=None) -> tuple:
    """
    Salva os arquivos Excel diretamente na rede usando smbclient.
    Usa a função export_to_excel_template para manter a estrutura do Anexo J.
    
    Returns:
        tuple: (success, message, file_paths, folder_path)
    """
    if not SMB_AVAILABLE:
        return False, "Biblioteca smbclient não disponível. Instale com: pip install smbprotocol", [], REDE_PATH
    
    total_rows = len(df)
    file_paths = []
    
    try:
        # Se cabe em um único arquivo
        if total_rows <= MAX_ROWS_PER_EXCEL:
            if progress_callback:
                progress_callback(0, 1, "Gerando arquivo Excel com template Anexo J...")
            
            # Callback interno para repassar progresso
            def internal_progress(pct, msg):
                if progress_callback:
                    # Converte pct (0-100) para (current, total, msg)
                    progress_callback(pct / 100 * 0.8, 1, msg)  # 0-80% para geração
            
            # Usa export_to_excel_template para manter a estrutura correta
            excel_data = export_to_excel_template(df, contrib_info, nivel, progress_callback=internal_progress)
            
            filename = get_export_filename(contrib_info, nivel, "xlsx")
            filepath = f"{REDE_PATH}\\{filename}"
            
            if progress_callback:
                progress_callback(0.85, 1, "Salvando na rede...")
            
            # Salva na rede usando smbclient
            with smbclient.open_file(filepath, mode="wb") as f:
                f.write(excel_data)
            
            file_paths.append(filepath)
            
            if progress_callback:
                progress_callback(1, 1, "Arquivo salvo!")
            
            return True, f"Arquivo salvo com sucesso!", file_paths, REDE_PATH
        
        # Precisa dividir em múltiplas partes
        total_partes = math.ceil(total_rows / MAX_ROWS_PER_EXCEL)
        
        for parte in range(1, total_partes + 1):
            if progress_callback:
                progress_callback(parte - 1, total_partes, f"Gerando parte {parte} de {total_partes}...")
            
            # Calcula índices
            start_idx = (parte - 1) * MAX_ROWS_PER_EXCEL
            end_idx = min(parte * MAX_ROWS_PER_EXCEL, total_rows)
            
            # Extrai a parte
            df_parte = df.iloc[start_idx:end_idx].copy()
            
            # Callback interno para cada parte
            def internal_progress_parte(pct, msg):
                if progress_callback:
                    base_progress = (parte - 1) / total_partes
                    part_progress = pct / 100 / total_partes * 0.9  # 90% para geração
                    progress_callback(base_progress + part_progress, 1, f"Parte {parte}: {msg}")
            
            # Usa export_to_excel_template para manter a estrutura correta
            excel_data = export_to_excel_template(
                df_parte, contrib_info, nivel,
                parte_atual=parte, total_partes=total_partes,
                progress_callback=internal_progress_parte
            )
            
            # Nome e caminho do arquivo
            base_filename = get_export_filename(contrib_info, nivel, "xlsx")
            parte_filename = base_filename.replace(".xlsx", f" - Parte {parte} de {total_partes}.xlsx")
            filepath = f"{REDE_PATH}\\{parte_filename}"
            
            # Salva na rede usando smbclient
            with smbclient.open_file(filepath, mode="wb") as f:
                f.write(excel_data)
            
            file_paths.append(filepath)
            
            # Libera memória
            del excel_data
            del df_parte
            
            if progress_callback:
                progress_callback(parte, total_partes, f"Parte {parte} de {total_partes} salva!")
        
        return True, f"{total_partes} arquivos salvos com sucesso!", file_paths, REDE_PATH
    
    except Exception as e:
        error_msg = str(e)
        # Detecta erro de autenticação Kerberos expirada
        if "Ticket expired" in error_msg or "SpnegoError" in error_msg or "authenticate" in error_msg.lower():
            return False, "🔐 **Sessão de rede expirada!** Faça logout/login no Windows ou acesse qualquer pasta de rede no Explorer para renovar.", file_paths, REDE_PATH
        return False, f"Erro ao salvar: {error_msg}", file_paths, REDE_PATH

def save_csv_to_network(df: pd.DataFrame, contrib_info: dict, nivel: str) -> tuple:
    """
    Salva CSV diretamente na rede usando smbclient.
    Formato brasileiro: separador (;), decimal (,), encoding latin-1
    
    Returns:
        tuple: (success, message, filepath, folder_path)
    """
    if not SMB_AVAILABLE:
        return False, "Biblioteca smbclient não disponível. Instale com: pip install smbprotocol", None, REDE_PATH
    
    try:
        filename = get_export_filename(contrib_info, nivel, "csv")
        filepath = f"{REDE_PATH}\\{filename}"
        
        # Gera CSV em memória no formato brasileiro
        csv_str = df.to_csv(index=False, sep=";", decimal=",")
        csv_bytes = csv_str.encode("latin-1", errors="replace")
        
        # Salva na rede usando smbclient
        with smbclient.open_file(filepath, mode="wb") as f:
            f.write(csv_bytes)
        
        return True, "CSV salvo com sucesso!", filepath, REDE_PATH
    
    except Exception as e:
        error_msg = str(e)
        # Detecta erro de autenticação Kerberos expirada
        if "Ticket expired" in error_msg or "SpnegoError" in error_msg or "authenticate" in error_msg.lower():
            return False, "🔐 **Sessão de rede expirada!** Faça logout/login no Windows ou acesse qualquer pasta de rede no Explorer para renovar.", None, REDE_PATH
        return False, f"Erro ao salvar CSV: {error_msg}", None, REDE_PATH

def save_to_network(df: pd.DataFrame, contrib_info: dict, nivel: str, progress_callback=None) -> tuple:
    """
    Salva os arquivos Excel diretamente na rede, evitando consumo de memória.
    
    Returns:
        tuple: (success, message, file_paths)
    """
    if not SMB_AVAILABLE:
        return False, "Biblioteca smbclient não disponível", []
    
    total_rows = len(df)
    file_paths = []
    
    try:
        # Cria diretório se não existir
        try:
            smbclient.makedirs(REDE_PATH, exist_ok=True)
        except:
            pass  # Diretório pode já existir
        
        # Se cabe em um único arquivo
        if total_rows <= MAX_ROWS_PER_EXCEL:
            if progress_callback:
                progress_callback(0, 1, "Gerando arquivo Excel...")
            
            excel_data = export_to_excel_template(df, contrib_info, nivel)
            filename = get_export_filename(contrib_info, nivel, "xlsx")
            filepath = f"{REDE_PATH}\\{filename}"
            
            with smbclient.open_file(filepath, mode="wb") as f:
                f.write(excel_data)
            
            file_paths.append(filepath)
            
            if progress_callback:
                progress_callback(1, 1, "Arquivo salvo na rede!")
            
            return True, f"Arquivo salvo com sucesso!", file_paths
        
        # Precisa dividir em múltiplas partes
        total_partes = math.ceil(total_rows / MAX_ROWS_PER_EXCEL)
        
        for parte in range(1, total_partes + 1):
            if progress_callback:
                progress_callback(parte - 1, total_partes, f"Gerando e salvando parte {parte} de {total_partes}...")
            
            # Calcula índices
            start_idx = (parte - 1) * MAX_ROWS_PER_EXCEL
            end_idx = min(parte * MAX_ROWS_PER_EXCEL, total_rows)
            
            # Extrai a parte
            df_parte = df.iloc[start_idx:end_idx].copy()
            
            # Gera o Excel
            excel_data = export_to_excel_template(
                df_parte, contrib_info, nivel,
                parte_atual=parte, total_partes=total_partes
            )
            
            # Nome e caminho do arquivo
            base_filename = get_export_filename(contrib_info, nivel, "xlsx")
            parte_filename = base_filename.replace(".xlsx", f" - Parte {parte} de {total_partes}.xlsx")
            filepath = f"{REDE_PATH}\\{parte_filename}"
            
            # Salva na rede
            with smbclient.open_file(filepath, mode="wb") as f:
                f.write(excel_data)
            
            file_paths.append(filepath)
            
            # Libera memória
            del excel_data
            del df_parte
            
            if progress_callback:
                progress_callback(parte, total_partes, f"Parte {parte} de {total_partes} salva!")
        
        return True, f"{total_partes} arquivos salvos com sucesso!", file_paths
    
    except Exception as e:
        error_msg = str(e)
        # Detecta erro de autenticação Kerberos expirada
        if "Ticket expired" in error_msg or "SpnegoError" in error_msg or "authenticate" in error_msg.lower():
            return False, "🔐 **Sessão de rede expirada!** Faça logout/login no Windows ou acesse qualquer pasta de rede no Explorer para renovar.", file_paths
        return False, f"Erro ao salvar: {error_msg}", file_paths

# =============================================================================
# 7. ANÁLISES EXPLORATÓRIAS
# =============================================================================

def render_analise_exploratoria(df: pd.DataFrame, nivel_str: str, _engine=None):
    """Renderiza análises exploratórias dos dados."""
    
    if df.empty:
        st.warning("Sem dados para análise.")
        return
    
    cfg = nivel_config(nivel_str)
    
    # Verifica qual coluna de infração usar (nova ou antiga estrutura)
    if 'infracao_ia' in df.columns:
        col_infracao = 'infracao_ia'
    else:
        col_infracao = cfg['col_infracao']
    
    st.markdown("---")
    st.subheader("📊 Análise Exploratória")
    
    tabs = st.tabs([
        "📈 Visão Temporal",
        "🏷️ Por NCM/CFOP", 
        "📦 Por Produto",
        "📊 Distribuição de Valores"
    ])
    
    # TAB 1: Visão Temporal
    with tabs[0]:
        col1, col2 = st.columns(2)
        
        with col1:
            # Infrações por período
            df_temp = df.copy()
            df_temp['infracao_valor'] = pd.to_numeric(df_temp[col_infracao], errors='coerce').fillna(0)
            
            if 'periodo' in df_temp.columns:
                df_periodo = df_temp.groupby('periodo').agg({
                    'infracao_valor': 'sum',
                    'chave': 'count'
                }).reset_index()
                df_periodo.columns = ['Período', 'Valor Infração', 'Quantidade']
                df_periodo = df_periodo.sort_values('Período')
                
                fig = px.bar(
                    df_periodo,
                    x='Período',
                    y='Valor Infração',
                    title=f"📅 Infrações por Período (Nível {cfg['label']})",
                    color_discrete_sequence=[cfg['cor']]
                )
                fig.update_layout(
                    xaxis_title="Período",
                    yaxis_title="Valor (R$)",
                    showlegend=False
                )
                st.plotly_chart(fig, use_container_width=True)
        
        with col2:
            # Quantidade de itens por período
            if 'periodo' in df_temp.columns and not df_periodo.empty:
                fig2 = px.line(
                    df_periodo,
                    x='Período',
                    y='Quantidade',
                    title="📊 Quantidade de Itens por Período",
                    markers=True
                )
                fig2.update_traces(line_color="#1565C0")
                fig2.update_layout(
                    xaxis_title="Período",
                    yaxis_title="Quantidade de Itens"
                )
                st.plotly_chart(fig2, use_container_width=True)
    
    # TAB 2: Por NCM/CFOP
    with tabs[1]:
        col1, col2 = st.columns(2)
        
        with col1:
            # Top 10 NCMs
            df_temp = df.copy()
            df_temp['infracao_valor'] = pd.to_numeric(df_temp[col_infracao], errors='coerce').fillna(0)
            
            if 'ncm' in df_temp.columns:
                df_ncm = df_temp.groupby('ncm').agg({
                    'infracao_valor': ['sum', 'count']
                }).reset_index()
                df_ncm.columns = ['NCM', 'Valor Total', 'Itens']
                df_ncm = df_ncm.nlargest(10, 'Valor Total')
                
                # Busca descrições dos NCMs
                if _engine is not None:
                    ncm_desc = get_ncm_descricoes(_engine, df_ncm['NCM'].tolist())
                    df_ncm['Descrição'] = df_ncm['NCM'].astype(str).map(ncm_desc).fillna('')
                else:
                    df_ncm['Descrição'] = ''
                
                # Calcula percentual para a barra de progresso
                max_valor = df_ncm['Valor Total'].max()
                df_ncm['_progress'] = df_ncm['Valor Total'] / max_valor if max_valor > 0 else 0
                
                st.markdown("##### 🏷️ Top NCM por Valor")
                st.dataframe(
                    df_ncm[['NCM', 'Descrição', 'Valor Total', 'Itens']],
                    use_container_width=True,
                    hide_index=True,
                    column_config={
                        'NCM': st.column_config.TextColumn('NCM', width='small'),
                        'Descrição': st.column_config.TextColumn('Descrição', width='medium'),
                        'Valor Total': st.column_config.ProgressColumn(
                            'Valor Total',
                            format='R$ %.2f',
                            min_value=0,
                            max_value=max_valor if max_valor > 0 else 1
                        ),
                        'Itens': st.column_config.NumberColumn('Itens', format='%d')
                    }
                )
        
        with col2:
            # Top 10 CFOPs
            if 'cfop' in df_temp.columns:
                df_cfop = df_temp.groupby('cfop').agg({
                    'infracao_valor': ['sum', 'count']
                }).reset_index()
                df_cfop.columns = ['CFOP', 'Valor Total', 'Itens']
                df_cfop = df_cfop.nlargest(10, 'Valor Total')
                
                # Busca descrições dos CFOPs
                if _engine is not None:
                    cfop_desc = get_cfop_descricoes(_engine, df_cfop['CFOP'].tolist())
                    df_cfop['Descrição'] = df_cfop['CFOP'].astype(str).map(cfop_desc).fillna('')
                else:
                    df_cfop['Descrição'] = ''
                
                # Calcula percentual para a barra de progresso
                max_valor_cfop = df_cfop['Valor Total'].max()
                df_cfop['_progress'] = df_cfop['Valor Total'] / max_valor_cfop if max_valor_cfop > 0 else 0
                
                st.markdown("##### 📊 Top CFOP por Valor")
                st.dataframe(
                    df_cfop[['CFOP', 'Descrição', 'Valor Total', 'Itens']],
                    use_container_width=True,
                    hide_index=True,
                    column_config={
                        'CFOP': st.column_config.TextColumn('CFOP', width='small'),
                        'Descrição': st.column_config.TextColumn('Descrição', width='medium'),
                        'Valor Total': st.column_config.ProgressColumn(
                            'Valor Total',
                            format='R$ %.2f',
                            min_value=0,
                            max_value=max_valor_cfop if max_valor_cfop > 0 else 1
                        ),
                        'Itens': st.column_config.NumberColumn('Itens', format='%d')
                    }
                )
    
    # TAB 3: Por Produto
    with tabs[2]:
        df_temp = df.copy()
        df_temp['infracao_valor'] = pd.to_numeric(df_temp[col_infracao], errors='coerce').fillna(0)
        
        if 'descricao' in df_temp.columns:
            # Agrupa por descrição incluindo NCM
            agg_dict = {
                'infracao_valor': ['sum', 'count']
            }
            group_cols = ['descricao']
            
            # Inclui NCM se disponível
            if 'ncm' in df_temp.columns:
                group_cols.append('ncm')
            
            df_prod = df_temp.groupby(group_cols).agg(agg_dict).reset_index()
            
            if 'ncm' in group_cols:
                df_prod.columns = ['Descrição', 'NCM', 'Valor Total', 'Itens']
            else:
                df_prod.columns = ['Descrição', 'Valor Total', 'Itens']
                df_prod['NCM'] = ''
            
            df_prod = df_prod.nlargest(15, 'Valor Total')
            
            # Calcula percentual para a barra de progresso
            max_valor_prod = df_prod['Valor Total'].max()
            
            st.markdown("##### 📦 Top Produtos por Valor")
            st.dataframe(
                df_prod[['Descrição', 'NCM', 'Valor Total', 'Itens']],
                use_container_width=True,
                hide_index=True,
                column_config={
                    'Descrição': st.column_config.TextColumn('Descrição', width='large'),
                    'NCM': st.column_config.TextColumn('NCM', width='small'),
                    'Valor Total': st.column_config.ProgressColumn(
                        'Valor Total',
                        format='R$ %.2f',
                        min_value=0,
                        max_value=max_valor_prod if max_valor_prod > 0 else 1
                    ),
                    'Itens': st.column_config.NumberColumn('Itens', format='%d')
                }
            )
    
    # TAB 4: Distribuição de Valores
    with tabs[3]:
        df_temp = df.copy()
        df_temp['infracao_valor'] = pd.to_numeric(df_temp[col_infracao], errors='coerce').fillna(0)
        df_temp = df_temp[df_temp['infracao_valor'] > 0]
        
        col1, col2 = st.columns(2)
        
        with col1:
            # Histograma
            fig = px.histogram(
                df_temp,
                x='infracao_valor',
                nbins=30,
                title="📊 Distribuição dos Valores de Infração",
                color_discrete_sequence=[cfg['cor']]
            )
            fig.update_layout(
                xaxis_title="Valor da Infração (R$)",
                yaxis_title="Frequência"
            )
            st.plotly_chart(fig, use_container_width=True)
        
        with col2:
            # Box plot
            fig = px.box(
                df_temp,
                y='infracao_valor',
                title="📈 Box Plot - Valores de Infração",
                color_discrete_sequence=[cfg['cor']]
            )
            fig.update_layout(
                yaxis_title="Valor da Infração (R$)"
            )
            st.plotly_chart(fig, use_container_width=True)
        
        # Estatísticas descritivas
        st.markdown("##### 📊 Estatísticas Descritivas")
        col1, col2, col3, col4, col5 = st.columns(5)
        
        with col1:
            st.metric("Mínimo", format_currency_br(df_temp['infracao_valor'].min()))
        with col2:
            st.metric("Máximo", format_currency_br(df_temp['infracao_valor'].max()))
        with col3:
            st.metric("Média", format_currency_br(df_temp['infracao_valor'].mean()))
        with col4:
            st.metric("Mediana", format_currency_br(df_temp['infracao_valor'].median()))
        with col5:
            st.metric("Desvio Padrão", format_currency_br(df_temp['infracao_valor'].std()))

# =============================================================================
# 8. COMPARATIVO ENTRE NÍVEIS
# =============================================================================

def render_comparativo_niveis(engine, identificador_digits: str, total_rows: int = 0, df_periodos=None, grupo: str = None):
    """
    Renderiza comparativo entre os três níveis de acurácia.

    Lógica dos níveis (hierarquia inclusiva):
    - BAIXA = todos os registros válidos (100%)
    - MÉDIA = subconjunto de BAIXA (registros mais confiáveis)
    - ALTA = subconjunto de MÉDIA (registros mais confiáveis ainda)

    Para calcular valores EXCLUSIVOS (sem sobreposição):
    - ALTA pura = válido em ALTA
    - MÉDIA pura = válido em MÉDIA mas NÃO em ALTA
    - BAIXA pura = válido em BAIXA mas NÃO em MÉDIA
    """
    if grupo is None:
        grupo = st.session_state.get('grupo_selecionado', GRUPO_PADRAO)

    tabelas = get_grupo_tabelas(grupo)

    st.markdown("---")
    st.subheader("🎯 Comparativo entre Níveis de Acurácia")

    # Verifica se precisa filtrar por período (datasets grandes)
    filtro_periodo = ""
    if total_rows > LARGE_DATASET_THRESHOLD and df_periodos is not None and len(df_periodos) > 0:
        st.warning(f"⚠️ Dataset grande ({total_rows:,} linhas). Exibindo apenas **últimos 12 meses** para melhor performance.")

        periodos_ordenados = sorted(
            df_periodos,
            key=lambda x: f"{x[3:7]}/{x[0:2]}" if len(str(x)) >= 7 else x,
            reverse=True
        )[:12]

        periodos_str = ", ".join([f"'{p}'" for p in periodos_ordenados])
        filtro_periodo = f"AND periodo IN ({periodos_str})"
        st.caption(f"📅 Períodos: {periodos_ordenados[0]} a {periodos_ordenados[-1]}")

    # Filtro base
    filtro_baixa = f"""
        regexp_replace(cnpj_emitente, '[^0-9]', '') = '{identificador_digits}'
        AND CAST(infracao_baixa AS STRING) != 'EXCLUIR'
        AND CAST(aliquota_baixa AS STRING) != 'EXCLUIR'
        AND CAST(legislacao_baixa AS STRING) != 'EXCLUIR'
        {filtro_periodo}
    """

    # Monta queries para cada tabela disponível
    union_parts = []
    select_cols = """
        infracao_alta, infracao_media, infracao_baixa,
        aliquota_alta, aliquota_media, aliquota_baixa,
        legislacao_alta, legislacao_media, legislacao_baixa,
        periodo
    """

    if tabelas.get('nfce'):
        union_parts.append(f"""
            SELECT {select_cols}
            FROM {tabelas['nfce']}
            WHERE {filtro_baixa}
        """)
    if tabelas.get('cupons'):
        union_parts.append(f"""
            SELECT {select_cols}
            FROM {tabelas['cupons']}
            WHERE {filtro_baixa}
        """)
    if tabelas.get('nfe'):
        union_parts.append(f"""
            SELECT {select_cols}
            FROM {tabelas['nfe']}
            WHERE {filtro_baixa}
        """)

    if not union_parts:
        st.warning("Nenhuma tabela disponível para este grupo.")
        return

    union_query = " UNION ALL ".join(union_parts)

    # Query com valores EXCLUSIVOS (sem sobreposição entre níveis)
    query_totais = f"""
    SELECT
        -- ALTA pura: válido em ALTA
        SUM(CASE WHEN CAST(infracao_alta AS STRING) != 'EXCLUIR'
                 AND CAST(aliquota_alta AS STRING) != 'EXCLUIR'
                 AND CAST(legislacao_alta AS STRING) != 'EXCLUIR'
                 THEN CAST(infracao_baixa AS FLOAT) ELSE 0 END) as total_alta,

        -- MÉDIA pura: válido em MÉDIA mas NÃO em ALTA
        SUM(CASE WHEN (CAST(infracao_media AS STRING) != 'EXCLUIR'
                       AND CAST(aliquota_media AS STRING) != 'EXCLUIR'
                       AND CAST(legislacao_media AS STRING) != 'EXCLUIR')
                  AND (CAST(infracao_alta AS STRING) = 'EXCLUIR'
                       OR CAST(aliquota_alta AS STRING) = 'EXCLUIR'
                       OR CAST(legislacao_alta AS STRING) = 'EXCLUIR')
                 THEN CAST(infracao_baixa AS FLOAT) ELSE 0 END) as total_media,

        -- BAIXA pura: válido em BAIXA mas NÃO em MÉDIA
        SUM(CASE WHEN (CAST(infracao_baixa AS STRING) != 'EXCLUIR'
                       AND CAST(aliquota_baixa AS STRING) != 'EXCLUIR'
                       AND CAST(legislacao_baixa AS STRING) != 'EXCLUIR')
                  AND (CAST(infracao_media AS STRING) = 'EXCLUIR'
                       OR CAST(aliquota_media AS STRING) = 'EXCLUIR'
                       OR CAST(legislacao_media AS STRING) = 'EXCLUIR')
                 THEN CAST(infracao_baixa AS FLOAT) ELSE 0 END) as total_baixa,

        -- Contagens exclusivas
        SUM(CASE WHEN CAST(infracao_alta AS STRING) != 'EXCLUIR'
                 AND CAST(aliquota_alta AS STRING) != 'EXCLUIR'
                 AND CAST(legislacao_alta AS STRING) != 'EXCLUIR'
                 THEN 1 ELSE 0 END) as qtd_alta,

        SUM(CASE WHEN (CAST(infracao_media AS STRING) != 'EXCLUIR'
                       AND CAST(aliquota_media AS STRING) != 'EXCLUIR'
                       AND CAST(legislacao_media AS STRING) != 'EXCLUIR')
                  AND (CAST(infracao_alta AS STRING) = 'EXCLUIR'
                       OR CAST(aliquota_alta AS STRING) = 'EXCLUIR'
                       OR CAST(legislacao_alta AS STRING) = 'EXCLUIR')
                 THEN 1 ELSE 0 END) as qtd_media,

        SUM(CASE WHEN (CAST(infracao_baixa AS STRING) != 'EXCLUIR'
                       AND CAST(aliquota_baixa AS STRING) != 'EXCLUIR'
                       AND CAST(legislacao_baixa AS STRING) != 'EXCLUIR')
                  AND (CAST(infracao_media AS STRING) = 'EXCLUIR'
                       OR CAST(aliquota_media AS STRING) = 'EXCLUIR'
                       OR CAST(legislacao_media AS STRING) = 'EXCLUIR')
                 THEN 1 ELSE 0 END) as qtd_baixa

    FROM (
        {union_query}
    ) t
    """
    
    try:
        df_totais = pd.read_sql(query_totais, engine)
        
        if df_totais.empty:
            st.warning("Não foi possível calcular os totais por nível.")
            return
        
        # Extrai os valores
        total_alta = float(df_totais['total_alta'].iloc[0]) if pd.notna(df_totais['total_alta'].iloc[0]) else 0
        total_media = float(df_totais['total_media'].iloc[0]) if pd.notna(df_totais['total_media'].iloc[0]) else 0
        total_baixa = float(df_totais['total_baixa'].iloc[0]) if pd.notna(df_totais['total_baixa'].iloc[0]) else 0
        
        qtd_alta = int(df_totais['qtd_alta'].iloc[0]) if pd.notna(df_totais['qtd_alta'].iloc[0]) else 0
        qtd_media = int(df_totais['qtd_media'].iloc[0]) if pd.notna(df_totais['qtd_media'].iloc[0]) else 0
        qtd_baixa = int(df_totais['qtd_baixa'].iloc[0]) if pd.notna(df_totais['qtd_baixa'].iloc[0]) else 0
        
        # Calcula total geral (soma dos 3 exclusivos)
        total_geral = total_alta + total_media + total_baixa
        qtd_total = qtd_alta + qtd_media + qtd_baixa
        
        if total_geral == 0:
            st.info("ℹ️ Nenhum registro encontrado com os filtros aplicados.")
            return
        
        # Calcula percentuais
        pct_alta = (total_alta / total_geral * 100) if total_geral > 0 else 0
        pct_media = (total_media / total_geral * 100) if total_geral > 0 else 0
        pct_baixa = (total_baixa / total_geral * 100) if total_geral > 0 else 0
        
        # KPIs lado a lado (valores EXCLUSIVOS)
        col1, col2, col3 = st.columns(3)
        
        with col1:
            st.markdown(f"""
            <div class="card-kpi card-kpi-alta">
                <h4>🟢 Nível ALTA ({pct_alta:.1f}%)</h4>
                <h2>{format_currency_br(total_alta)}</h2>
                <p>{format_number_br(qtd_alta)} itens</p>
            </div>
            """, unsafe_allow_html=True)
        
        with col2:
            st.markdown(f"""
            <div class="card-kpi card-kpi-media">
                <h4>🟡 Nível MÉDIA ({pct_media:.1f}%)</h4>
                <h2>{format_currency_br(total_media)}</h2>
                <p>{format_number_br(qtd_media)} itens</p>
            </div>
            """, unsafe_allow_html=True)
        
        with col3:
            st.markdown(f"""
            <div class="card-kpi card-kpi-baixa">
                <h4>🔴 Nível BAIXA ({pct_baixa:.1f}%)</h4>
                <h2>{format_currency_br(total_baixa)}</h2>
                <p>{format_number_br(qtd_baixa)} itens</p>
            </div>
            """, unsafe_allow_html=True)
        
        # Linha de total
        st.markdown(f"""
        <div style="text-align: center; padding: 10px; background: #f5f5f5; border-radius: 5px; margin-top: 10px;">
            <strong>💰 Total Geral:</strong> {format_currency_br(total_geral)} | 
            <strong>📋 Itens:</strong> {format_number_br(qtd_total)}
        </div>
        """, unsafe_allow_html=True)
        
        # Gráficos
        col1, col2 = st.columns(2)
        
        with col1:
            df_comp = pd.DataFrame([
                {'Nível': '🟢 ALTA', 'Valor': total_alta, 'Percentual': pct_alta},
                {'Nível': '🟡 MÉDIA', 'Valor': total_media, 'Percentual': pct_media},
                {'Nível': '🔴 BAIXA', 'Valor': total_baixa, 'Percentual': pct_baixa}
            ])
            
            fig = px.bar(
                df_comp,
                x='Nível',
                y='Valor',
                title="💰 Valor por Nível (Exclusivo)",
                color='Nível',
                color_discrete_map={
                    '🟢 ALTA': '#4CAF50',
                    '🟡 MÉDIA': '#FF9800',
                    '🔴 BAIXA': '#f44336'
                },
                text=df_comp['Percentual'].apply(lambda x: f'{x:.1f}%')
            )
            fig.update_traces(textposition='outside')
            fig.update_layout(showlegend=False)
            st.plotly_chart(fig, use_container_width=True)
        
        with col2:
            # Gráfico de pizza
            fig = px.pie(
                df_comp,
                values='Valor',
                names='Nível',
                title="📊 Distribuição por Nível",
                color='Nível',
                color_discrete_map={
                    '🟢 ALTA': '#4CAF50',
                    '🟡 MÉDIA': '#FF9800',
                    '🔴 BAIXA': '#f44336'
                },
                hole=0.4
            )
            fig.update_traces(textposition='inside', textinfo='percent+label')
            st.plotly_chart(fig, use_container_width=True)
        
        # Legenda explicativa
        st.caption("""
        **Legenda:** 
        🟢 **ALTA** = Consenso das 3 IAs (maior confiabilidade) |
        🟡 **MÉDIA** = Maioria 2x1 (confiabilidade moderada) |
        🔴 **BAIXA** = IAs divergentes (requer avaliação manual)
        """)
    
    except Exception as e:
        error_msg = str(e)
        if is_table_unavailable_error(error_msg):
            st.warning(TABLE_UNAVAILABLE_MSG)
        else:
            st.warning(f"Não foi possível carregar o comparativo entre níveis: {error_msg[:100]}")

# =============================================================================
# 9. RANKING DE EMPRESAS (CACHE DIÁRIO)
# =============================================================================

@st.cache_data(ttl=RANKING_CACHE_TTL, show_spinner=False)
def get_ranking_data(_engine, nivel: str = "ALTA", top_n: int = 100, _cache_version: int = 8, grupo: str = None):
    """
    Busca ranking agregado de empresas por valor de infração.
    Cache de 24 horas pois dados não mudam frequentemente.
    Retorna dados agregados por empresa e por ano.
    _cache_version: incrementar para invalidar cache
    grupo: grupo (GESSUPER, GESMAC). Se None, usa session_state

    """
    if grupo is None:
        grupo = st.session_state.get('grupo_selecionado', GRUPO_PADRAO)

    tabelas = get_grupo_tabelas(grupo)

    nivel_upper = (nivel or "ALTA").upper()
    col_infracao = f"infracao_{nivel_upper.lower()}"
    col_aliquota = f"aliquota_{nivel_upper.lower()}"
    col_legislacao = f"legislacao_{nivel_upper.lower()}"

    filtro = f"""
        {col_infracao} IS NOT NULL
        AND CAST({col_infracao} AS STRING) != 'EXCLUIR'
        AND CAST({col_aliquota} AS STRING) != 'EXCLUIR'
        AND CAST({col_legislacao} AS STRING) != 'EXCLUIR'
    """

    # Monta queries para cada tabela disponível
    union_parts = []
    if tabelas.get('nfce'):
        union_parts.append(f"""
            SELECT cnpj_emitente, razao_emitente, periodo, {col_infracao}
            FROM {tabelas['nfce']}
            WHERE {filtro}
        """)
    if tabelas.get('cupons'):
        union_parts.append(f"""
            SELECT cnpj_emitente, razao_emitente, periodo, {col_infracao}
            FROM {tabelas['cupons']}
            WHERE {filtro}
        """)
    if tabelas.get('nfe'):
        union_parts.append(f"""
            SELECT cnpj_emitente, razao_emitente, periodo, {col_infracao}
            FROM {tabelas['nfe']}
            WHERE {filtro}
        """)

    if not union_parts:
        return None, None, None

    union_query = " UNION ALL ".join(union_parts)

    # Query otimizada - agregação no banco de dados
    query = f"""
    SELECT
        cnpj_emitente,
        razao_emitente,
        SUBSTR(periodo, 4, 4) as ano,
        SUM(CAST({col_infracao} AS FLOAT)) as total_valor,
        COUNT(*) as qtd_itens
    FROM (
        {union_query}
    ) t
    GROUP BY cnpj_emitente, razao_emitente, SUBSTR(periodo, 4, 4)
    """
    
    try:
        df = pd.read_sql(query, _engine)
        
        if df.empty:
            return None, None, None
        
        # Pivoteia para ter anos como colunas
        df_pivot_valor = df.pivot_table(
            index=['cnpj_emitente', 'razao_emitente'],
            columns='ano',
            values='total_valor',
            aggfunc='sum',
            fill_value=0
        ).reset_index()
        
        df_pivot_qtd = df.pivot_table(
            index=['cnpj_emitente', 'razao_emitente'],
            columns='ano',
            values='qtd_itens',
            aggfunc='sum',
            fill_value=0
        ).reset_index()
        
        # Calcula totais
        anos_cols = [c for c in df_pivot_valor.columns if c not in ['cnpj_emitente', 'razao_emitente']]
        
        # Garante que as colunas de anos sejam numéricas (float64)
        for col in anos_cols:
            df_pivot_valor[col] = pd.to_numeric(df_pivot_valor[col], errors='coerce').fillna(0).astype('float64')
            df_pivot_qtd[col] = pd.to_numeric(df_pivot_qtd[col], errors='coerce').fillna(0).astype('float64')
        
        # Calcula TOTAL como soma das colunas de anos (float64)
        df_pivot_valor['TOTAL'] = df_pivot_valor[anos_cols].sum(axis=1).astype('float64')
        df_pivot_qtd['TOTAL'] = df_pivot_qtd[anos_cols].sum(axis=1).astype('float64')
        
        # =====================================================================
        # ESTATÍSTICAS GERAIS (ANTES DE LIMITAR ÀS TOP N)
        # =====================================================================
        total_geral_todas = float(df_pivot_valor['TOTAL'].sum())
        qtd_empresas_total = len(df_pivot_valor)
        total_itens_todas = int(df_pivot_qtd['TOTAL'].sum())
        
        # Estatísticas por ano (todas empresas)
        stats_por_ano = {}
        for ano in anos_cols:
            if ano in df_pivot_valor.columns:
                valor_ano = float(df_pivot_valor[ano].sum())
                qtd_ano = int(df_pivot_qtd[ano].sum()) if ano in df_pivot_qtd.columns else 0
                stats_por_ano[ano] = {
                    'valor': valor_ano,
                    'qtd': qtd_ano,
                    'pct': (valor_ano / total_geral_todas * 100) if total_geral_todas > 0 else 0,
                    'empresas_ativas': int((df_pivot_valor[ano] > 0).sum())
                }
        
        # =====================================================================
        # ORDENA E LIMITA ÀS TOP N
        # =====================================================================
        indices_ordenados = df_pivot_valor['TOTAL'].values.argsort()[::-1]  # Decrescente
        df_pivot_valor = df_pivot_valor.iloc[indices_ordenados].head(top_n).reset_index(drop=True)
        
        # Alinha df_qtd com df_valor
        df_pivot_qtd = df_pivot_qtd.set_index(['cnpj_emitente', 'razao_emitente'])
        df_pivot_qtd = df_pivot_qtd.reindex(
            df_pivot_valor.set_index(['cnpj_emitente', 'razao_emitente']).index
        ).reset_index()
        
        # Estatísticas descritivas (das top N)
        stats_descritivas = {
            'media': float(df_pivot_valor['TOTAL'].mean()),
            'mediana': float(df_pivot_valor['TOTAL'].median()),
            'std': float(df_pivot_valor['TOTAL'].std()),
            'min': float(df_pivot_valor['TOTAL'].min()),
            'max': float(df_pivot_valor['TOTAL'].max()),
            'q1': float(df_pivot_valor['TOTAL'].quantile(0.25)),
            'q3': float(df_pivot_valor['TOTAL'].quantile(0.75)),
        }
        
        # Total das top N
        total_top_n = float(df_pivot_valor['TOTAL'].sum())
        total_itens_top_n = int(df_pivot_qtd['TOTAL'].sum())
        
        return df_pivot_valor, df_pivot_qtd, {
            'total_geral': total_geral_todas,  # Total de TODAS as empresas
            'total_top_n': total_top_n,        # Total das top N
            'qtd_empresas': len(df_pivot_valor),  # Qtd no ranking (top N)
            'qtd_empresas_total': qtd_empresas_total,  # Qtd total de empresas
            'anos': sorted(anos_cols),
            'por_ano': stats_por_ano,
            'descritivas': stats_descritivas,
            'total_itens': total_itens_todas,  # Total de itens
            'total_itens_top_n': total_itens_top_n
        }
        
    except Exception as e:
        error_msg = str(e)
        if is_table_unavailable_error(error_msg):
            st.warning(TABLE_UNAVAILABLE_MSG)
        else:
            st.error(f"Erro ao buscar ranking: {error_msg[:150]}")
        return None, None, None


@st.cache_data(ttl=RANKING_CACHE_TTL, show_spinner=False)
def get_global_stats(_engine, nivel: str = "ALTA", grupo: str = None):
    """
    Busca estatísticas globais para comparação.
    Cache de 24 horas.
    """
    if grupo is None:
        grupo = st.session_state.get('grupo_selecionado', GRUPO_PADRAO)

    tabelas = get_grupo_tabelas(grupo)

    nivel_upper = (nivel or "ALTA").upper()
    col_infracao = f"infracao_{nivel_upper.lower()}"
    col_aliquota = f"aliquota_{nivel_upper.lower()}"
    col_legislacao = f"legislacao_{nivel_upper.lower()}"

    filtro = f"""
        {col_infracao} IS NOT NULL
        AND CAST({col_infracao} AS STRING) != 'EXCLUIR'
        AND CAST({col_aliquota} AS STRING) != 'EXCLUIR'
        AND CAST({col_legislacao} AS STRING) != 'EXCLUIR'
    """

    # Monta queries para cada tabela disponível
    union_parts = []
    if tabelas.get('nfce'):
        union_parts.append(f"""
            SELECT cnpj_emitente, {col_infracao}
            FROM {tabelas['nfce']}
            WHERE {filtro}
        """)
    if tabelas.get('cupons'):
        union_parts.append(f"""
            SELECT cnpj_emitente, {col_infracao}
            FROM {tabelas['cupons']}
            WHERE {filtro}
        """)
    if tabelas.get('nfe'):
        union_parts.append(f"""
            SELECT cnpj_emitente, {col_infracao}
            FROM {tabelas['nfe']}
            WHERE {filtro}
        """)

    if not union_parts:
        return None

    union_query = " UNION ALL ".join(union_parts)

    query = f"""
    SELECT
        COUNT(DISTINCT cnpj_emitente) as total_empresas,
        SUM(CAST({col_infracao} AS FLOAT)) as total_valor,
        COUNT(*) as total_itens,
        AVG(CAST({col_infracao} AS FLOAT)) as media_item
    FROM (
        {union_query}
    ) t
    """
    
    try:
        df = pd.read_sql(query, _engine)
        if df.empty:
            return None
        return {
            'total_empresas': int(df['total_empresas'].iloc[0]) if pd.notna(df['total_empresas'].iloc[0]) else 0,
            'total_valor': float(df['total_valor'].iloc[0]) if pd.notna(df['total_valor'].iloc[0]) else 0,
            'total_itens': int(df['total_itens'].iloc[0]) if pd.notna(df['total_itens'].iloc[0]) else 0,
            'media_item': float(df['media_item'].iloc[0]) if pd.notna(df['media_item'].iloc[0]) else 0
        }
    except Exception as e:
        error_msg = str(e)
        if is_table_unavailable_error(error_msg):
            st.warning(TABLE_UNAVAILABLE_MSG)
        return None


@st.cache_data(ttl=RANKING_CACHE_TTL, show_spinner=False)
def get_ranking_acuracia(_engine, top_n: int = 100, _cache_version: int = 6, grupo: str = None):
    """
    Busca ranking de empresas por qualidade de acurácia.
    Ordena por: maior % ALTA, depois % MÉDIA, depois % BAIXA, depois valor total.
    Cache de 24 horas.

    Lógica dos níveis (hierarquia inclusiva):
    - BAIXA = todos os registros válidos
    - MÉDIA = subconjunto de BAIXA (registros mais confiáveis)
    - ALTA = subconjunto de MÉDIA (registros mais confiáveis ainda)

    Para calcular valores EXCLUSIVOS:
    - ALTA pura = válido em ALTA
    - MÉDIA pura = válido em MÉDIA mas NÃO em ALTA
    - BAIXA pura = válido em BAIXA mas NÃO em MÉDIA
    """
    if grupo is None:
        grupo = st.session_state.get('grupo_selecionado', GRUPO_PADRAO)

    tabelas = get_grupo_tabelas(grupo)

    filtro_baixa = """
        CAST(infracao_baixa AS STRING) != 'EXCLUIR'
        AND CAST(aliquota_baixa AS STRING) != 'EXCLUIR'
        AND CAST(legislacao_baixa AS STRING) != 'EXCLUIR'
    """

    # Monta queries para cada tabela disponível
    union_parts = []
    select_cols = """
        cnpj_emitente, razao_emitente,
        infracao_alta, infracao_media, infracao_baixa,
        aliquota_alta, aliquota_media, aliquota_baixa,
        legislacao_alta, legislacao_media, legislacao_baixa
    """

    if tabelas.get('nfce'):
        union_parts.append(f"""
            SELECT {select_cols}
            FROM {tabelas['nfce']}
            WHERE {filtro_baixa}
        """)
    if tabelas.get('cupons'):
        union_parts.append(f"""
            SELECT {select_cols}
            FROM {tabelas['cupons']}
            WHERE {filtro_baixa}
        """)
    if tabelas.get('nfe'):
        union_parts.append(f"""
            SELECT {select_cols}
            FROM {tabelas['nfe']}
            WHERE {filtro_baixa}
        """)

    if not union_parts:
        return None

    union_query = " UNION ALL ".join(union_parts)

    query = f"""
    SELECT
        cnpj_emitente,
        razao_emitente,

        -- ALTA pura: válido em ALTA (usa infracao_baixa como valor base)
        SUM(CASE WHEN CAST(infracao_alta AS STRING) != 'EXCLUIR'
                 AND CAST(aliquota_alta AS STRING) != 'EXCLUIR'
                 AND CAST(legislacao_alta AS STRING) != 'EXCLUIR'
                 THEN CAST(infracao_baixa AS FLOAT) ELSE 0 END) as total_alta,

        -- MÉDIA pura: válido em MÉDIA mas NÃO em ALTA
        SUM(CASE WHEN (CAST(infracao_media AS STRING) != 'EXCLUIR'
                       AND CAST(aliquota_media AS STRING) != 'EXCLUIR'
                       AND CAST(legislacao_media AS STRING) != 'EXCLUIR')
                  AND (CAST(infracao_alta AS STRING) = 'EXCLUIR'
                       OR CAST(aliquota_alta AS STRING) = 'EXCLUIR'
                       OR CAST(legislacao_alta AS STRING) = 'EXCLUIR')
                 THEN CAST(infracao_baixa AS FLOAT) ELSE 0 END) as total_media,

        -- BAIXA pura: válido em BAIXA mas NÃO em MÉDIA
        SUM(CASE WHEN (CAST(infracao_baixa AS STRING) != 'EXCLUIR'
                       AND CAST(aliquota_baixa AS STRING) != 'EXCLUIR'
                       AND CAST(legislacao_baixa AS STRING) != 'EXCLUIR')
                  AND (CAST(infracao_media AS STRING) = 'EXCLUIR'
                       OR CAST(aliquota_media AS STRING) = 'EXCLUIR'
                       OR CAST(legislacao_media AS STRING) = 'EXCLUIR')
                 THEN CAST(infracao_baixa AS FLOAT) ELSE 0 END) as total_baixa,

        -- Contagens para referência
        SUM(CASE WHEN CAST(infracao_alta AS STRING) != 'EXCLUIR'
                 AND CAST(aliquota_alta AS STRING) != 'EXCLUIR'
                 AND CAST(legislacao_alta AS STRING) != 'EXCLUIR'
                 THEN 1 ELSE 0 END) as qtd_alta,

        SUM(CASE WHEN (CAST(infracao_media AS STRING) != 'EXCLUIR'
                       AND CAST(aliquota_media AS STRING) != 'EXCLUIR'
                       AND CAST(legislacao_media AS STRING) != 'EXCLUIR')
                  AND (CAST(infracao_alta AS STRING) = 'EXCLUIR'
                       OR CAST(aliquota_alta AS STRING) = 'EXCLUIR'
                       OR CAST(legislacao_alta AS STRING) = 'EXCLUIR')
                 THEN 1 ELSE 0 END) as qtd_media,

        SUM(CASE WHEN (CAST(infracao_baixa AS STRING) != 'EXCLUIR'
                       AND CAST(aliquota_baixa AS STRING) != 'EXCLUIR'
                       AND CAST(legislacao_baixa AS STRING) != 'EXCLUIR')
                  AND (CAST(infracao_media AS STRING) = 'EXCLUIR'
                       OR CAST(aliquota_media AS STRING) = 'EXCLUIR'
                       OR CAST(legislacao_media AS STRING) = 'EXCLUIR')
                 THEN 1 ELSE 0 END) as qtd_baixa

    FROM (
        {union_query}
    ) t
    GROUP BY cnpj_emitente, razao_emitente
    """
    
    try:
        df = pd.read_sql(query, _engine)
        
        if df.empty:
            return None
        
        # Calcula totais e percentuais
        df['total_valor'] = df['total_alta'] + df['total_media'] + df['total_baixa']
        df['total_qtd'] = df['qtd_alta'] + df['qtd_media'] + df['qtd_baixa']
        
        # Filtra empresas com algum valor
        df = df[df['total_valor'] > 0].copy()
        
        # Calcula percentuais
        df['pct_alta'] = (df['total_alta'] / df['total_valor'] * 100).round(2)
        df['pct_media'] = (df['total_media'] / df['total_valor'] * 100).round(2)
        df['pct_baixa'] = (df['total_baixa'] / df['total_valor'] * 100).round(2)
        
        # Ordena: maior % ALTA, depois % MÉDIA, depois % BAIXA, depois valor total
        df = df.sort_values(
            by=['pct_alta', 'pct_media', 'pct_baixa', 'total_valor'],
            ascending=[False, False, False, False]
        ).head(top_n).reset_index(drop=True)
        
        # Adiciona posição no ranking
        df['#'] = range(1, len(df) + 1)
        
        return df
        
    except Exception as e:
        error_msg = str(e)
        if is_table_unavailable_error(error_msg):
            st.warning(TABLE_UNAVAILABLE_MSG)
        else:
            st.error(f"Erro ao buscar ranking de acurácia: {error_msg[:150]}")
        return None


@st.cache_data(ttl=RANKING_CACHE_TTL, show_spinner=False)
def get_stats_acuracia_geral(_engine, _cache_version: int = 1, grupo: str = None):
    """
    Busca estatísticas gerais de acurácia (totais por nível).
    Retorna valor e quantidade para cada nível (ALTA, MÉDIA, BAIXA) com valores exclusivos.
    Cache de 24 horas.
    """
    if grupo is None:
        grupo = st.session_state.get('grupo_selecionado', GRUPO_PADRAO)

    tabelas = get_grupo_tabelas(grupo)

    filtro_baixa = """
        CAST(infracao_baixa AS STRING) != 'EXCLUIR'
        AND CAST(aliquota_baixa AS STRING) != 'EXCLUIR'
        AND CAST(legislacao_baixa AS STRING) != 'EXCLUIR'
    """

    # Monta queries para cada tabela disponível
    union_parts = []
    select_cols = """
        infracao_alta, infracao_media, infracao_baixa,
        aliquota_alta, aliquota_media, aliquota_baixa,
        legislacao_alta, legislacao_media, legislacao_baixa
    """

    if tabelas.get('nfce'):
        union_parts.append(f"""
            SELECT {select_cols}
            FROM {tabelas['nfce']}
            WHERE {filtro_baixa}
        """)
    if tabelas.get('cupons'):
        union_parts.append(f"""
            SELECT {select_cols}
            FROM {tabelas['cupons']}
            WHERE {filtro_baixa}
        """)
    if tabelas.get('nfe'):
        union_parts.append(f"""
            SELECT {select_cols}
            FROM {tabelas['nfe']}
            WHERE {filtro_baixa}
        """)

    if not union_parts:
        return None

    union_query = " UNION ALL ".join(union_parts)

    query = f"""
    SELECT
        -- ALTA pura: válido em ALTA
        SUM(CASE WHEN CAST(infracao_alta AS STRING) != 'EXCLUIR'
                 AND CAST(aliquota_alta AS STRING) != 'EXCLUIR'
                 AND CAST(legislacao_alta AS STRING) != 'EXCLUIR'
                 THEN CAST(infracao_baixa AS FLOAT) ELSE 0 END) as valor_alta,

        -- MÉDIA pura: válido em MÉDIA mas NÃO em ALTA
        SUM(CASE WHEN (CAST(infracao_media AS STRING) != 'EXCLUIR'
                       AND CAST(aliquota_media AS STRING) != 'EXCLUIR'
                       AND CAST(legislacao_media AS STRING) != 'EXCLUIR')
                  AND (CAST(infracao_alta AS STRING) = 'EXCLUIR'
                       OR CAST(aliquota_alta AS STRING) = 'EXCLUIR'
                       OR CAST(legislacao_alta AS STRING) = 'EXCLUIR')
                 THEN CAST(infracao_baixa AS FLOAT) ELSE 0 END) as valor_media,

        -- BAIXA pura: válido em BAIXA mas NÃO em MÉDIA
        SUM(CASE WHEN (CAST(infracao_baixa AS STRING) != 'EXCLUIR'
                       AND CAST(aliquota_baixa AS STRING) != 'EXCLUIR'
                       AND CAST(legislacao_baixa AS STRING) != 'EXCLUIR')
                  AND (CAST(infracao_media AS STRING) = 'EXCLUIR'
                       OR CAST(aliquota_media AS STRING) = 'EXCLUIR'
                       OR CAST(legislacao_media AS STRING) = 'EXCLUIR')
                 THEN CAST(infracao_baixa AS FLOAT) ELSE 0 END) as valor_baixa,

        -- Contagens exclusivas
        SUM(CASE WHEN CAST(infracao_alta AS STRING) != 'EXCLUIR'
                 AND CAST(aliquota_alta AS STRING) != 'EXCLUIR'
                 AND CAST(legislacao_alta AS STRING) != 'EXCLUIR'
                 THEN 1 ELSE 0 END) as qtd_alta,

        SUM(CASE WHEN (CAST(infracao_media AS STRING) != 'EXCLUIR'
                       AND CAST(aliquota_media AS STRING) != 'EXCLUIR'
                       AND CAST(legislacao_media AS STRING) != 'EXCLUIR')
                  AND (CAST(infracao_alta AS STRING) = 'EXCLUIR'
                       OR CAST(aliquota_alta AS STRING) = 'EXCLUIR'
                       OR CAST(legislacao_alta AS STRING) = 'EXCLUIR')
                 THEN 1 ELSE 0 END) as qtd_media,

        SUM(CASE WHEN (CAST(infracao_baixa AS STRING) != 'EXCLUIR'
                       AND CAST(aliquota_baixa AS STRING) != 'EXCLUIR'
                       AND CAST(legislacao_baixa AS STRING) != 'EXCLUIR')
                  AND (CAST(infracao_media AS STRING) = 'EXCLUIR'
                       OR CAST(aliquota_media AS STRING) = 'EXCLUIR'
                       OR CAST(legislacao_media AS STRING) = 'EXCLUIR')
                 THEN 1 ELSE 0 END) as qtd_baixa

    FROM (
        {union_query}
    ) t
    """
    
    try:
        df = pd.read_sql(query, _engine)
        
        if df.empty:
            return None
        
        # Extrai valores
        valor_alta = float(df['valor_alta'].iloc[0]) if pd.notna(df['valor_alta'].iloc[0]) else 0
        valor_media = float(df['valor_media'].iloc[0]) if pd.notna(df['valor_media'].iloc[0]) else 0
        valor_baixa = float(df['valor_baixa'].iloc[0]) if pd.notna(df['valor_baixa'].iloc[0]) else 0
        
        qtd_alta = int(df['qtd_alta'].iloc[0]) if pd.notna(df['qtd_alta'].iloc[0]) else 0
        qtd_media = int(df['qtd_media'].iloc[0]) if pd.notna(df['qtd_media'].iloc[0]) else 0
        qtd_baixa = int(df['qtd_baixa'].iloc[0]) if pd.notna(df['qtd_baixa'].iloc[0]) else 0
        
        # Totais
        valor_total = valor_alta + valor_media + valor_baixa
        qtd_total = qtd_alta + qtd_media + qtd_baixa
        
        # Percentuais
        pct_valor_alta = (valor_alta / valor_total * 100) if valor_total > 0 else 0
        pct_valor_media = (valor_media / valor_total * 100) if valor_total > 0 else 0
        pct_valor_baixa = (valor_baixa / valor_total * 100) if valor_total > 0 else 0
        
        pct_qtd_alta = (qtd_alta / qtd_total * 100) if qtd_total > 0 else 0
        pct_qtd_media = (qtd_media / qtd_total * 100) if qtd_total > 0 else 0
        pct_qtd_baixa = (qtd_baixa / qtd_total * 100) if qtd_total > 0 else 0
        
        return {
            'valor_alta': valor_alta,
            'valor_media': valor_media,
            'valor_baixa': valor_baixa,
            'valor_total': valor_total,
            'qtd_alta': qtd_alta,
            'qtd_media': qtd_media,
            'qtd_baixa': qtd_baixa,
            'qtd_total': qtd_total,
            'pct_valor_alta': pct_valor_alta,
            'pct_valor_media': pct_valor_media,
            'pct_valor_baixa': pct_valor_baixa,
            'pct_qtd_alta': pct_qtd_alta,
            'pct_qtd_media': pct_qtd_media,
            'pct_qtd_baixa': pct_qtd_baixa
        }
        
    except Exception as e:
        error_msg = str(e)
        if is_table_unavailable_error(error_msg):
            st.warning(TABLE_UNAVAILABLE_MSG)
        elif "TTransport" in error_msg or "timeout" in error_msg.lower() or "read 0 bytes" in error_msg:
            st.warning("⏳ **Consulta muito pesada.** A query de estatísticas gerais pode demorar. Tente novamente em alguns minutos.")
        else:
            st.error(f"Erro ao buscar estatísticas: {error_msg[:150]}")
        return None


def render_ranking(engine, nivel: str = "ALTA"):
    """Renderiza a página de Ranking de Empresas."""
    
    # Limpa flag de tabela indisponível
    st.session_state.tabela_indisponivel = False
    
    # Verificação de disponibilidade das tabelas
    if not check_tables_available(engine):
        col_title, col_btn1, col_btn2 = st.columns([3, 1, 1])
        with col_title:
            st.markdown("## 🏆 Ranking de Empresas")
        with col_btn1:
            if st.button("🔍 Consulta", use_container_width=True, type="secondary"):
                st.session_state.nav_page = "consulta"
                st.rerun()
        with col_btn2:
            if st.button("🔎 Produtos", use_container_width=True, type="secondary"):
                st.session_state.nav_page = "produtos"
                st.rerun()
        st.warning(TABLE_UNAVAILABLE_MSG)
        return
    
    # Header com botões de navegação
    col_title, col_btn1, col_btn2 = st.columns([3, 1, 1])
    
    with col_title:
        st.markdown("## 🏆 Ranking de Empresas")
    
    with col_btn1:
        if st.button("🔍 Consulta", use_container_width=True, type="secondary"):
            st.session_state.nav_page = "consulta"
            st.rerun()
    
    with col_btn2:
        if st.button("🔎 Produtos", use_container_width=True, type="secondary"):
            st.session_state.nav_page = "produtos"
            st.rerun()
    
    # Usa nível ALTA fixo (maior confiabilidade)
    # A distribuição entre níveis é mostrada na seção "Estatísticas Gerais de Acurácia"
    nivel = "ALTA"
    
    # Mostra período limite no caption
    st.caption("📊 Nível: **🟢 ALTA** (maior confiabilidade) | Dados agregados por empresa e ano | Cache: 24h")
    
    with st.spinner("Carregando ranking..."):
        df_valor, df_qtd, stats = get_ranking_data(engine, nivel, top_n=100, _cache_version=8)
    
    if df_valor is None:
        st.warning("Não foi possível carregar o ranking.")
        return
    
    # =========================================================================
    # LINHA 1: KPIs PRINCIPAIS (VISÃO GERAL)
    # =========================================================================
    st.markdown("### 📈 Visão Geral")
    
    col1, col2, col3, col4, col5 = st.columns(5)
    with col1:
        st.metric(
            "💰 Valor Total", 
            format_currency_br(stats['total_geral']),
            help="Soma de TODAS as infrações no nível selecionado"
        )
    with col2:
        pct_top100 = (stats['total_top_n'] / stats['total_geral'] * 100) if stats['total_geral'] > 0 else 0
        st.metric(
            "🏆 Top 100", 
            format_currency_br(stats['total_top_n']),
            delta=f"{pct_top100:.1f}% do total",
            help="Soma das 100 maiores empresas"
        )
    with col3:
        st.metric(
            "🏢 Empresas", 
            f"{stats['qtd_empresas_total']:,}",
            delta=f"Top 100 de {stats['qtd_empresas_total']:,}",
            delta_color="off",
            help="Total de empresas com infrações"
        )
    with col4:
        st.metric(
            "📋 Total de Itens", 
            f"{stats['total_itens']:,}",
            help="Quantidade total de itens/registros"
        )
    with col5:
        st.metric(
            "📅 Período", 
            f"{min(stats['anos'])} - {max(stats['anos'])}",
            help="Anos disponíveis nos dados"
        )
    
    st.markdown("---")
    
    # =========================================================================
    # TABELA DO RANKING (MOVIDA PARA CIMA)
    # =========================================================================
    
    # Formata tabela para exibição
    df_display = df_valor.copy()
    
    # Converte nomes de colunas para string (anos podem vir como int)
    df_display.columns = [str(col) for col in df_display.columns]
    
    # Garante que está ordenado por TOTAL (numérico) decrescente usando argsort
    indices_ordenados = df_display['TOTAL'].values.argsort()[::-1]
    df_display = df_display.iloc[indices_ordenados].reset_index(drop=True)
    
    # Adiciona posição (baseado na ordem por TOTAL)
    df_display.insert(0, '#', range(1, len(df_display) + 1))
    
    # Renomeia colunas
    df_display = df_display.rename(columns={
        'cnpj_emitente': 'CNPJ',
        'razao_emitente': 'Razão Social'
    })
    
    # Trunca razão social
    df_display['Razão Social'] = df_display['Razão Social'].apply(
        lambda x: x[:40] + '...' if pd.notna(x) and len(str(x)) > 40 else x
    )
    
    # Colunas de anos (sem TOTAL) - converte para string para garantir consistência
    anos_cols = [str(ano) for ano in stats['anos']]
    
    # Calcula percentual de cada ano em relação ao TOTAL da empresa
    for ano in anos_cols:
        if ano in df_display.columns:
            col_pct = f'{ano}%'
            df_display[col_pct] = (df_display[ano] / df_display['TOTAL'] * 100).round(1)
            # Trata divisão por zero
            df_display[col_pct] = df_display[col_pct].fillna(0)
    
    # =========================================================================
    # SELETOR DE ANO PARA ORDENAÇÃO
    # =========================================================================
    st.markdown("### 🏅 Top 100 Empresas")
    
    col_ordem, col_info = st.columns([1, 5])
    
    with col_ordem:
        opcoes_ordenacao = ["TOTAL"] + anos_cols
        ano_selecionado = st.selectbox(
            "Ordenação",
            options=opcoes_ordenacao,
            format_func=lambda x: f"📊 TOTAL" if x == "TOTAL" else f"📅 {x}",
            key="ranking_ordenar_por",
            label_visibility="collapsed"
        )
    
    with col_info:
        if ano_selecionado == "TOTAL":
            st.caption("📊 Ordenado pelo **valor total** (todos os anos)")
        else:
            st.caption(f"📅 Ordenado pelo **% em {ano_selecionado}** (maior concentração neste ano)")
    
    # Reordena o DataFrame pelo PERCENTUAL do ano selecionado
    if ano_selecionado != "TOTAL":
        col_pct_ordenar = f'{ano_selecionado}%'
        # Força a ordenação
        df_display = df_display.sort_values(by=col_pct_ordenar, ascending=False, ignore_index=True)
        # Recalcula o ranking
        df_display['#'] = range(1, len(df_display) + 1)
    
    # Reordena colunas: #, CNPJ, Razão, [Ano, Ano%, ...], TOTAL
    cols_ordenadas = ['#', 'CNPJ', 'Razão Social']
    for ano in anos_cols:
        if ano in df_display.columns:
            cols_ordenadas.append(ano)
            cols_ordenadas.append(f'{ano}%')
    cols_ordenadas.append('TOTAL')
    
    df_display = df_display[cols_ordenadas]
    
    # Configura colunas para exibição formatada MAS ordenação numérica
    column_config = {
        '#': st.column_config.NumberColumn('#', width='small'),
        'CNPJ': st.column_config.TextColumn('CNPJ', width='medium'),
        'Razão Social': st.column_config.TextColumn('Razão Social', width='large'),
        'TOTAL': st.column_config.NumberColumn('TOTAL', format="R$ %.2f"),
    }
    
    # Configura colunas de valores e percentuais
    for ano in anos_cols:
        if ano in df_display.columns:
            # Destaca a coluna do ano selecionado
            label = f"⭐{ano}" if ano == ano_selecionado else ano
            column_config[ano] = st.column_config.NumberColumn(
                label,
                format="R$ %.2f"
            )
            label_pct = f"⭐{ano}%" if ano == ano_selecionado else f'{ano}%'
            column_config[f'{ano}%'] = st.column_config.NumberColumn(
                label_pct,
                format="%.1f%%"
            )
    
    # Exibe tabela com valores numéricos (ordenação funciona corretamente)
    st.dataframe(
        df_display,
        use_container_width=True,
        hide_index=True,
        height=500,
        column_config=column_config
    )
    
    if ano_selecionado == "TOTAL":
        st.caption("💡 Clique no cabeçalho da coluna para ordenar. Colunas % mostram participação de cada ano no total da empresa.")
    else:
        st.caption(f"💡 Ranking ordenado por **% em {ano_selecionado}**. Empresas com maior concentração de infrações neste ano aparecem primeiro. ⭐ = ano selecionado.")
    
    # =========================================================================
    # EXPANDER: ESTATÍSTICAS GERAIS DE ACURÁCIA
    # =========================================================================
    with st.expander("📊 Estatísticas Gerais de Acurácia", expanded=False):
        st.markdown("""
        **Distribuição geral das infrações por nível de acurácia.**
        
        Os valores são **exclusivos** (sem sobreposição):
        - 🟢 **ALTA** = Consenso das 3 IAs (maior confiabilidade)
        - 🟡 **MÉDIA** = Maioria 2x1, não consenso total
        - 🔴 **BAIXA** = IAs divergentes (requer avaliação manual)
        """)
        
        # Tenta carregar estatísticas gerais
        stats_acur = None
        
        with st.spinner("Carregando estatísticas de acurácia..."):
            # Primeiro tenta usar os dados do ranking de acurácia (mais leve)
            df_acuracia = get_ranking_acuracia(engine, top_n=10000, _cache_version=6)
            
            if df_acuracia is not None and not df_acuracia.empty:
                # Calcula totais a partir do ranking (soma de todas as empresas)
                valor_alta = df_acuracia['total_alta'].sum()
                valor_media = df_acuracia['total_media'].sum()
                valor_baixa = df_acuracia['total_baixa'].sum()
                valor_total = valor_alta + valor_media + valor_baixa
                
                qtd_alta = df_acuracia['qtd_alta'].sum()
                qtd_media = df_acuracia['qtd_media'].sum()
                qtd_baixa = df_acuracia['qtd_baixa'].sum()
                qtd_total = qtd_alta + qtd_media + qtd_baixa
                
                if valor_total > 0 and qtd_total > 0:
                    stats_acur = {
                        'valor_alta': valor_alta,
                        'valor_media': valor_media,
                        'valor_baixa': valor_baixa,
                        'valor_total': valor_total,
                        'qtd_alta': int(qtd_alta),
                        'qtd_media': int(qtd_media),
                        'qtd_baixa': int(qtd_baixa),
                        'qtd_total': int(qtd_total),
                        'pct_valor_alta': (valor_alta / valor_total * 100),
                        'pct_valor_media': (valor_media / valor_total * 100),
                        'pct_valor_baixa': (valor_baixa / valor_total * 100),
                        'pct_qtd_alta': (qtd_alta / qtd_total * 100),
                        'pct_qtd_media': (qtd_media / qtd_total * 100),
                        'pct_qtd_baixa': (qtd_baixa / qtd_total * 100)
                    }
        
        if stats_acur is not None:
            # KPIs em cards
            st.markdown("##### 💰 Por Valor")
            col1, col2, col3, col4 = st.columns(4)
            
            with col1:
                st.metric(
                    "🟢 ALTA",
                    format_currency_br(stats_acur['valor_alta']),
                    f"{stats_acur['pct_valor_alta']:.1f}%"
                )
            with col2:
                st.metric(
                    "🟡 MÉDIA",
                    format_currency_br(stats_acur['valor_media']),
                    f"{stats_acur['pct_valor_media']:.1f}%"
                )
            with col3:
                st.metric(
                    "🔴 BAIXA",
                    format_currency_br(stats_acur['valor_baixa']),
                    f"{stats_acur['pct_valor_baixa']:.1f}%"
                )
            with col4:
                st.metric(
                    "💰 TOTAL",
                    format_currency_br(stats_acur['valor_total']),
                    "100%",
                    delta_color="off"
                )
            
            st.markdown("##### 📋 Por Quantidade de Itens")
            col1, col2, col3, col4 = st.columns(4)
            
            with col1:
                st.metric(
                    "🟢 ALTA",
                    f"{stats_acur['qtd_alta']:,}".replace(",", "."),
                    f"{stats_acur['pct_qtd_alta']:.1f}%"
                )
            with col2:
                st.metric(
                    "🟡 MÉDIA",
                    f"{stats_acur['qtd_media']:,}".replace(",", "."),
                    f"{stats_acur['pct_qtd_media']:.1f}%"
                )
            with col3:
                st.metric(
                    "🔴 BAIXA",
                    f"{stats_acur['qtd_baixa']:,}".replace(",", "."),
                    f"{stats_acur['pct_qtd_baixa']:.1f}%"
                )
            with col4:
                st.metric(
                    "📋 TOTAL",
                    f"{stats_acur['qtd_total']:,}".replace(",", "."),
                    "100%",
                    delta_color="off"
                )
            
            # Gráficos
            st.markdown("---")
            col1, col2 = st.columns(2)
            
            with col1:
                df_pie_valor = pd.DataFrame([
                    {'Nível': '🟢 ALTA', 'Valor': stats_acur['valor_alta'], 'Percentual': stats_acur['pct_valor_alta']},
                    {'Nível': '🟡 MÉDIA', 'Valor': stats_acur['valor_media'], 'Percentual': stats_acur['pct_valor_media']},
                    {'Nível': '🔴 BAIXA', 'Valor': stats_acur['valor_baixa'], 'Percentual': stats_acur['pct_valor_baixa']}
                ])
                
                fig1 = px.pie(
                    df_pie_valor,
                    values='Valor',
                    names='Nível',
                    title="💰 Distribuição por Valor",
                    color='Nível',
                    color_discrete_map={
                        '🟢 ALTA': '#4CAF50',
                        '🟡 MÉDIA': '#FF9800',
                        '🔴 BAIXA': '#f44336'
                    },
                    hole=0.4
                )
                fig1.update_traces(textposition='inside', textinfo='percent+label')
                st.plotly_chart(fig1, use_container_width=True)
            
            with col2:
                df_pie_qtd = pd.DataFrame([
                    {'Nível': '🟢 ALTA', 'Quantidade': stats_acur['qtd_alta'], 'Percentual': stats_acur['pct_qtd_alta']},
                    {'Nível': '🟡 MÉDIA', 'Quantidade': stats_acur['qtd_media'], 'Percentual': stats_acur['pct_qtd_media']},
                    {'Nível': '🔴 BAIXA', 'Quantidade': stats_acur['qtd_baixa'], 'Percentual': stats_acur['pct_qtd_baixa']}
                ])
                
                fig2 = px.pie(
                    df_pie_qtd,
                    values='Quantidade',
                    names='Nível',
                    title="📋 Distribuição por Quantidade",
                    color='Nível',
                    color_discrete_map={
                        '🟢 ALTA': '#4CAF50',
                        '🟡 MÉDIA': '#FF9800',
                        '🔴 BAIXA': '#f44336'
                    },
                    hole=0.4
                )
                fig2.update_traces(textposition='inside', textinfo='percent+label')
                st.plotly_chart(fig2, use_container_width=True)
            
            # Tabela resumo
            st.markdown("##### 📋 Tabela Resumo")
            df_resumo = pd.DataFrame([
                {
                    'Nível': '🟢 ALTA',
                    'Valor (R$)': stats_acur['valor_alta'],
                    '% Valor': stats_acur['pct_valor_alta'],
                    'Quantidade': stats_acur['qtd_alta'],
                    '% Qtd': stats_acur['pct_qtd_alta']
                },
                {
                    'Nível': '🟡 MÉDIA',
                    'Valor (R$)': stats_acur['valor_media'],
                    '% Valor': stats_acur['pct_valor_media'],
                    'Quantidade': stats_acur['qtd_media'],
                    '% Qtd': stats_acur['pct_qtd_media']
                },
                {
                    'Nível': '🔴 BAIXA',
                    'Valor (R$)': stats_acur['valor_baixa'],
                    '% Valor': stats_acur['pct_valor_baixa'],
                    'Quantidade': stats_acur['qtd_baixa'],
                    '% Qtd': stats_acur['pct_qtd_baixa']
                },
                {
                    'Nível': '💰 TOTAL',
                    'Valor (R$)': stats_acur['valor_total'],
                    '% Valor': 100.0,
                    'Quantidade': stats_acur['qtd_total'],
                    '% Qtd': 100.0
                }
            ])
            
            st.dataframe(
                df_resumo,
                use_container_width=True,
                hide_index=True,
                column_config={
                    'Nível': st.column_config.TextColumn('Nível'),
                    'Valor (R$)': st.column_config.NumberColumn('Valor (R$)', format="R$ %.2f"),
                    '% Valor': st.column_config.ProgressColumn('% Valor', format="%.1f%%", min_value=0, max_value=100),
                    'Quantidade': st.column_config.NumberColumn('Quantidade', format="%d"),
                    '% Qtd': st.column_config.ProgressColumn('% Qtd', format="%.1f%%", min_value=0, max_value=100)
                }
            )
        else:
            st.info("Não foi possível carregar as estatísticas de acurácia.")
    
    # =========================================================================
    # EXPANDER: RANKING POR ACURÁCIA (POR EMPRESA)
    # =========================================================================
    with st.expander("🎯 Ranking por Qualidade de Acurácia (por Empresa)", expanded=False):
        st.markdown("""
        **Lógica de ordenação:**
        1. 🟢 Maior % de infrações ALTA (mais confiável)
        2. 🟡 Em empate, maior % MÉDIA
        3. 🔴 Em empate, maior % BAIXA  
        4. 💰 Em empate final, maior valor total
        """)
        
        with st.spinner("Carregando ranking de acurácia..."):
            df_acuracia = get_ranking_acuracia(engine, top_n=100, _cache_version=6)
        
        if df_acuracia is not None and not df_acuracia.empty:
            # Prepara dados para exibição
            df_acur_display = df_acuracia[[
                '#', 'cnpj_emitente', 'razao_emitente', 
                'pct_alta', 'pct_media', 'pct_baixa',
                'total_alta', 'total_media', 'total_baixa', 'total_valor'
            ]].copy()
            
            df_acur_display.columns = [
                '#', 'CNPJ', 'Razão Social',
                '🟢 % ALTA', '🟡 % MÉDIA', '🔴 % BAIXA',
                '🟢 R$ ALTA', '🟡 R$ MÉDIA', '🔴 R$ BAIXA', '💰 TOTAL'
            ]
            
            # Configuração das colunas
            column_config_acur = {
                '#': st.column_config.NumberColumn('#', width='small'),
                'CNPJ': st.column_config.TextColumn('CNPJ', width='medium'),
                'Razão Social': st.column_config.TextColumn('Razão Social', width='large'),
                '🟢 % ALTA': st.column_config.ProgressColumn(
                    '🟢 % ALTA',
                    format="%.1f%%",
                    min_value=0,
                    max_value=100
                ),
                '🟡 % MÉDIA': st.column_config.ProgressColumn(
                    '🟡 % MÉDIA',
                    format="%.1f%%",
                    min_value=0,
                    max_value=100
                ),
                '🔴 % BAIXA': st.column_config.ProgressColumn(
                    '🔴 % BAIXA',
                    format="%.1f%%",
                    min_value=0,
                    max_value=100
                ),
                '🟢 R$ ALTA': st.column_config.NumberColumn('🟢 R$ ALTA', format="R$ %.2f"),
                '🟡 R$ MÉDIA': st.column_config.NumberColumn('🟡 R$ MÉDIA', format="R$ %.2f"),
                '🔴 R$ BAIXA': st.column_config.NumberColumn('🔴 R$ BAIXA', format="R$ %.2f"),
                '💰 TOTAL': st.column_config.NumberColumn('💰 TOTAL', format="R$ %.2f")
            }
            
            st.dataframe(
                df_acur_display,
                use_container_width=True,
                hide_index=True,
                height=400,
                column_config=column_config_acur
            )
            
            # Estatísticas resumidas
            st.markdown("---")
            col1, col2, col3, col4 = st.columns(4)
            
            media_pct_alta = df_acuracia['pct_alta'].mean()
            media_pct_media = df_acuracia['pct_media'].mean()
            media_pct_baixa = df_acuracia['pct_baixa'].mean()
            
            with col1:
                st.metric("📊 Média % ALTA", f"{media_pct_alta:.1f}%")
            with col2:
                st.metric("📊 Média % MÉDIA", f"{media_pct_media:.1f}%")
            with col3:
                st.metric("📊 Média % BAIXA", f"{media_pct_baixa:.1f}%")
            with col4:
                empresas_majoritaria_alta = len(df_acuracia[df_acuracia['pct_alta'] > 50])
                st.metric("🏆 Empresas >50% ALTA", f"{empresas_majoritaria_alta}")
            
            st.caption("💡 Empresas com maior % ALTA têm infrações mais confiáveis (consenso das 3 IAs).")
        else:
            st.info("Não foi possível carregar o ranking de acurácia.")
    
    # =========================================================================
    # EXPANDER: ESTATÍSTICAS DESCRITIVAS
    # =========================================================================
    with st.expander("📊 Estatísticas Descritivas", expanded=False):
        desc = stats['descritivas']
        col1, col2, col3, col4, col5 = st.columns(5)
        with col1:
            st.metric("📊 Média", format_currency_br(desc['media']))
        with col2:
            st.metric("📊 Mediana", format_currency_br(desc['mediana']))
        with col3:
            st.metric("📉 Mínimo", format_currency_br(desc['min']))
        with col4:
            st.metric("📈 Máximo", format_currency_br(desc['max']))
        with col5:
            st.metric("📏 Desvio Padrão", format_currency_br(desc['std']))
        
        # Concentração (em relação ao TOTAL GERAL de todas empresas)
        total_geral = stats['total_geral']
        if total_geral > 0:
            st.markdown("#### Concentração")
            # Calcula concentração das top N empresas (df_valor já está ordenado)
            top10_valor = df_valor.head(10)['TOTAL'].sum()
            top20_valor = df_valor.head(20)['TOTAL'].sum()
            top50_valor = df_valor.head(50)['TOTAL'].sum()
            
            pct_top10 = (top10_valor / total_geral) * 100
            pct_top20 = (top20_valor / total_geral) * 100
            pct_top50 = (top50_valor / total_geral) * 100
            
            col1, col2, col3, col4 = st.columns(4)
            with col1:
                st.metric(
                    "🔝 Top 10", 
                    f"{pct_top10:.1f}%",
                    delta=format_currency_br(top10_valor),
                    help="Concentração do valor nas 10 maiores empresas vs total geral"
                )
            with col2:
                st.metric(
                    "🔝 Top 20", 
                    f"{pct_top20:.1f}%",
                    delta=format_currency_br(top20_valor),
                    help="Concentração do valor nas 20 maiores empresas vs total geral"
                )
            with col3:
                st.metric(
                    "🔝 Top 50", 
                    f"{pct_top50:.1f}%",
                    delta=format_currency_br(top50_valor),
                    help="Concentração do valor nas 50 maiores empresas vs total geral"
                )
            with col4:
                # Média por item
                media_item = total_geral / stats['total_itens'] if stats['total_itens'] > 0 else 0
                st.metric(
                    "💵 Média/Item", 
                    format_currency_br(media_item),
                    help="Valor médio por item de infração"
                )
    
    # =========================================================================
    # EXPANDER: DISTRIBUIÇÃO POR ANO
    # =========================================================================
    with st.expander("📅 Distribuição por Ano", expanded=False):
        anos = stats['anos']
        cols = st.columns(len(anos))
        
        for i, ano in enumerate(anos):
            with cols[i]:
                ano_stats = stats['por_ano'].get(ano, {})
                valor = ano_stats.get('valor', 0)
                pct = ano_stats.get('pct', 0)
                qtd = ano_stats.get('qtd', 0)
                empresas = ano_stats.get('empresas_ativas', 0)
                
                # Cor baseada no percentual
                if pct >= 25:
                    cor = "#4CAF50"  # Verde
                elif pct >= 15:
                    cor = "#FF9800"  # Laranja
                else:
                    cor = "#9E9E9E"  # Cinza
                
                st.markdown(f"""
                <div style='background: linear-gradient(135deg, {cor}22 0%, {cor}11 100%); 
                            padding: 1rem; border-radius: 10px; border-left: 4px solid {cor};
                            text-align: center;'>
                    <h3 style='margin: 0; color: {cor};'>{ano}</h3>
                    <h2 style='margin: 0.5rem 0;'>{format_currency_br(valor)}</h2>
                    <p style='margin: 0; font-size: 1.2rem; font-weight: bold; color: {cor};'>{pct:.1f}%</p>
                    <p style='margin: 0.3rem 0 0 0; font-size: 0.8rem; color: #666;'>
                        {format_number_br(qtd)} itens | {empresas} empresas
                    </p>
                </div>
                """, unsafe_allow_html=True)
    
    # =========================================================================
    # EXPANDER: VISUALIZAÇÕES (GRÁFICOS)
    # =========================================================================
    with st.expander("📊 Visualizações", expanded=False):
        tab_dist, tab_top10 = st.tabs(["📅 Distribuição por Ano", "🏆 Top 10 Empresas"])
        
        with tab_dist:
            # Gráfico de barras por ano
            anos_data = []
            for ano in stats['anos']:
                ano_stats = stats['por_ano'].get(ano, {})
                anos_data.append({
                    'Ano': ano,
                    'Valor': ano_stats.get('valor', 0),
                    'Percentual': ano_stats.get('pct', 0),
                    'Itens': ano_stats.get('qtd', 0),
                    'Empresas': ano_stats.get('empresas_ativas', 0)
                })
            
            df_anos = pd.DataFrame(anos_data)
            
            col1, col2 = st.columns(2)
            
            with col1:
                fig1 = px.bar(
                    df_anos,
                    x='Ano',
                    y='Valor',
                    title="💰 Valor por Ano",
                    color='Valor',
                    color_continuous_scale='Blues',
                    text=df_anos['Percentual'].apply(lambda x: f'{x:.1f}%')
                )
                fig1.update_traces(textposition='outside')
                fig1.update_layout(showlegend=False)
                st.plotly_chart(fig1, use_container_width=True)
            
            with col2:
                fig2 = px.pie(
                    df_anos,
                    values='Valor',
                    names='Ano',
                    title="📊 Distribuição Percentual",
                    hole=0.4
                )
                fig2.update_traces(textposition='inside', textinfo='percent+label')
                st.plotly_chart(fig2, use_container_width=True)
        
        with tab_top10:
            df_top10 = df_valor.head(10).copy()
            df_top10['Empresa'] = df_top10['razao_emitente'].apply(
                lambda x: x[:25] + '...' if len(str(x)) > 25 else x
            )
            
            fig = px.bar(
                df_top10,
                x='Empresa',
                y='TOTAL',
                title="🏆 Top 10 Empresas por Valor Total",
                color='TOTAL',
                color_continuous_scale='Reds'
            )
            fig.update_layout(xaxis_tickangle=-45, showlegend=False)
            st.plotly_chart(fig, use_container_width=True)


# =============================================================================
# 9. PESQUISA DE PRODUTOS
# =============================================================================

def search_products_by_description(_engine, search_term: str, limit: int = 1000, grupo: str = None):
    """
    Busca produtos por descrição.
    Retorna DataFrame com produtos, empresas, alíquotas, NCM, CFOP, etc.
    Query simplificada sem GROUP BY para melhor performance.
    Suporta múltiplos grupos.
    """
    if grupo is None:
        grupo = st.session_state.get('grupo_selecionado', GRUPO_PADRAO)

    tabelas = get_grupo_tabelas(grupo)

    # Escapa aspas simples e converte para minúsculas
    search_term_safe = search_term.replace("'", "''").lower()

    filtro = f"""
        LOWER(descricao) LIKE '%{search_term_safe}%'
        AND CAST(infracao_baixa AS STRING) != 'EXCLUIR'
        AND CAST(aliquota_baixa AS STRING) != 'EXCLUIR'
    """

    # Monta queries para cada tabela disponível
    union_parts = []

    if tabelas.get('nfce'):
        union_parts.append(f"""
            SELECT
                descricao,
                ncm,
                gtin,
                CAST(cfop AS STRING) AS cfop,
                cnpj_emitente,
                razao_emitente,
                icms_emitente as aliquota_emitente,
                aliquota_alta as aliquota_ia_alta,
                aliquota_media as aliquota_ia_media,
                aliquota_baixa as aliquota_ia_baixa,
                infracao_alta,
                infracao_media,
                infracao_baixa,
                legislacao_alta,
                legislacao_media,
                legislacao_baixa,
                periodo,
                'NFC-e' as tipo_doc
            FROM {tabelas['nfce']}
            WHERE {filtro}
        """)

    if tabelas.get('cupons'):
        union_parts.append(f"""
            SELECT
                descricao,
                ncm,
                gtin,
                CAST(cfop AS STRING) AS cfop,
                cnpj_emitente,
                razao_emitente,
                icms_emitente as aliquota_emitente,
                aliquota_alta as aliquota_ia_alta,
                aliquota_media as aliquota_ia_media,
                aliquota_baixa as aliquota_ia_baixa,
                infracao_alta,
                infracao_media,
                infracao_baixa,
                legislacao_alta,
                legislacao_media,
                legislacao_baixa,
                periodo,
                'Cupom' as tipo_doc
            FROM {tabelas['cupons']}
            WHERE {filtro}
        """)

    if tabelas.get('nfe'):
        union_parts.append(f"""
            SELECT
                descricao,
                ncm,
                gtin,
                CAST(cfop AS STRING) AS cfop,
                cnpj_emitente,
                razao_emitente,
                aliquota_emitente as aliquota_emitente,
                aliquota_alta as aliquota_ia_alta,
                aliquota_media as aliquota_ia_media,
                aliquota_baixa as aliquota_ia_baixa,
                infracao_alta,
                infracao_media,
                infracao_baixa,
                legislacao_alta,
                legislacao_media,
                legislacao_baixa,
                periodo,
                'NF-e' as tipo_doc
            FROM {tabelas['nfe']}
            WHERE {filtro}
        """)

    if not union_parts:
        return pd.DataFrame()

    union_query = " UNION ALL ".join(union_parts)
    query = f"{union_query} LIMIT {limit}"

    try:
        df = pd.read_sql(query, _engine)
        return df
    except Exception as e:
        error_msg = str(e)
        if is_table_unavailable_error(error_msg):
            st.session_state.tabela_indisponivel = True
        else:
            # Mostra o erro real para debug
            st.error(f"❌ Erro na pesquisa: {error_msg[:300]}")
        return pd.DataFrame()


def render_pesquisa_produtos(engine):
    """Renderiza a página de Pesquisa de Produtos."""
    
    # Limpa flag de tabela indisponível
    st.session_state.tabela_indisponivel = False
    
    # Header com botões de navegação
    col_title, col_btn1, col_btn2 = st.columns([3, 1, 1])
    
    with col_title:
        st.markdown("## 🔎 Pesquisa de Produtos")
    
    with col_btn1:
        if st.button("🏆 Ranking", use_container_width=True, type="secondary"):
            st.session_state.nav_page = "ranking"
            st.rerun()
    
    with col_btn2:
        if st.button("🔍 Consulta", use_container_width=True, type="secondary"):
            st.session_state.nav_page = "consulta"
            st.rerun()
    
    # Verificação de disponibilidade das tabelas
    if not check_tables_available(engine):
        st.warning(TABLE_UNAVAILABLE_MSG)
        return
    
    st.caption("📦 Pesquise produtos pela descrição para analisar como estão sendo tributados")
    
    # Campo de pesquisa
    col_search, col_btn = st.columns([4, 1])
    
    with col_search:
        search_term = st.text_input(
            "Descrição do Produto",
            placeholder="Ex: CERVEJA, REFRIGERANTE, AGUA MINERAL...",
            key="search_produto",
            label_visibility="collapsed"
        )
    
    with col_btn:
        search_clicked = st.button("🔎 PESQUISAR", type="primary", use_container_width=True)
    
    # Dicas de pesquisa
    with st.expander("💡 Dicas de Pesquisa", expanded=False):
        st.markdown("""
        - Use termos simples e genéricos (ex: `CERVEJA` ao invés de `CERVEJA PILSEN 350ML`)
        - A busca não diferencia maiúsculas/minúsculas
        - Quanto mais específico o termo, menos resultados
        - Exemplos: `AGUA MINERAL`, `REFRIGERANTE`, `VINHO`, `WHISKY`, `ENERGETICO`
        """)
    
    # Executa pesquisa
    if search_clicked and search_term:
        if len(search_term) < 3:
            st.warning("⚠️ Digite pelo menos 3 caracteres para pesquisar.")
            return
        
        with st.spinner(f"🔍 Pesquisando '{search_term}'..."):
            df = search_products_by_description(engine, search_term, limit=5000)
        
        if df.empty:
            st.info(f"ℹ️ Nenhum produto encontrado com '{search_term}'.")
            return
        
        # =====================================================================
        # PRÉ-PROCESSAMENTO DOS DADOS
        # =====================================================================
        # Converte infracao_baixa para numérico
        df['valor_infracao'] = pd.to_numeric(df['infracao_baixa'], errors='coerce').fillna(0)
        
        # =====================================================================
        # RESULTADOS
        # =====================================================================
        
        st.success(f"✅ Encontrados **{len(df):,}** registros para '{search_term}'")
        
        # KPIs
        st.markdown("### 📊 Resumo")
        col1, col2, col3, col4 = st.columns(4)
        
        with col1:
            st.metric("📦 Registros", f"{len(df):,}")
        with col2:
            empresas_unicas = df['cnpj_emitente'].nunique()
            st.metric("🏢 Empresas", f"{empresas_unicas:,}")
        with col3:
            ncm_unicos = df['ncm'].nunique()
            st.metric("🏷️ NCMs", f"{ncm_unicos:,}")
        with col4:
            valor_total = df['valor_infracao'].sum()
            st.metric("💰 Valor Infrações", format_currency_br(valor_total))
        
        # Tabs de análise
        tab_aliq, tab_ncm, tab_empresas, tab_dados = st.tabs([
            "📊 Alíquotas", "🏷️ NCMs", "🏢 Empresas", "📋 Dados"
        ])
        
        # ---------------------------------------------------------------------
        # TAB: ALÍQUOTAS
        # ---------------------------------------------------------------------
        with tab_aliq:
            st.markdown("#### Distribuição de Alíquotas")
            
            # Alíquota do emitente
            st.markdown("##### 🏢 Alíquota Informada pelo Emitente")
            aliq_emitente = df.groupby('aliquota_emitente').agg({
                'valor_infracao': ['count', 'sum']
            }).reset_index()
            aliq_emitente.columns = ['Alíquota', 'Quantidade', 'Valor Infração']
            aliq_emitente = aliq_emitente.sort_values('Quantidade', ascending=False)
            
            col1, col2 = st.columns(2)
            with col1:
                fig = px.pie(
                    aliq_emitente.head(10),
                    values='Quantidade',
                    names='Alíquota',
                    title="Por Quantidade",
                    hole=0.4
                )
                st.plotly_chart(fig, use_container_width=True)
            
            with col2:
                st.dataframe(
                    aliq_emitente,
                    use_container_width=True,
                    hide_index=True,
                    column_config={
                        'Alíquota': st.column_config.TextColumn('Alíquota'),
                        'Quantidade': st.column_config.NumberColumn('Quantidade', format="%d"),
                        'Valor Infração': st.column_config.NumberColumn('Valor Infração', format="R$ %.2f")
                    }
                )
            
            # Alíquota sugerida pela IA
            st.markdown("##### 🤖 Alíquota Sugerida pela IA (Nível ALTA)")
            aliq_ia = df.groupby('aliquota_ia_alta').agg({
                'valor_infracao': ['count', 'sum']
            }).reset_index()
            aliq_ia.columns = ['Alíquota IA', 'Quantidade', 'Valor Infração']
            aliq_ia = aliq_ia.sort_values('Quantidade', ascending=False)
            
            col1, col2 = st.columns(2)
            with col1:
                fig = px.pie(
                    aliq_ia.head(10),
                    values='Quantidade',
                    names='Alíquota IA',
                    title="Por Quantidade",
                    hole=0.4
                )
                st.plotly_chart(fig, use_container_width=True)
            
            with col2:
                st.dataframe(
                    aliq_ia,
                    use_container_width=True,
                    hide_index=True
                )
            
            # Comparativo Emitente vs IA
            st.markdown("##### ⚖️ Comparativo: Alíquota Emitente vs IA")
            df_comp = df.copy()
            df_comp['divergente'] = df_comp['aliquota_emitente'].astype(str) != df_comp['aliquota_ia_alta'].astype(str)
            divergentes = df_comp['divergente'].sum()
            total = len(df_comp)
            pct_divergente = (divergentes / total * 100) if total > 0 else 0
            
            col1, col2, col3 = st.columns(3)
            with col1:
                st.metric("✅ Concordantes", f"{total - divergentes:,}")
            with col2:
                st.metric("⚠️ Divergentes", f"{divergentes:,}")
            with col3:
                st.metric("📊 % Divergência", f"{pct_divergente:.1f}%")
        
        # ---------------------------------------------------------------------
        # TAB: NCMs
        # ---------------------------------------------------------------------
        with tab_ncm:
            st.markdown("#### 🏷️ Top NCM por Valor")
            
            ncm_stats = df.groupby('ncm').agg({
                'valor_infracao': ['count', 'sum'],
                'cnpj_emitente': 'nunique',
                'descricao': 'first'
            }).reset_index()
            ncm_stats.columns = ['NCM', 'Itens', 'Valor Total', 'Empresas', 'Exemplo Descrição']
            ncm_stats = ncm_stats.sort_values('Valor Total', ascending=False)
            
            # Busca descrições dos NCMs
            ncm_list = ncm_stats['NCM'].tolist()
            ncm_desc = get_ncm_descricoes(engine, ncm_list)
            ncm_stats['Descrição'] = ncm_stats['NCM'].map(ncm_desc).fillna('')
            
            # Calcula max para barra de progresso
            max_valor_ncm = ncm_stats['Valor Total'].max()
            
            # Tabela com barras de progresso
            st.dataframe(
                ncm_stats[['NCM', 'Descrição', 'Valor Total', 'Empresas', 'Itens']],
                use_container_width=True,
                hide_index=True,
                column_config={
                    'NCM': st.column_config.TextColumn('NCM', width='small'),
                    'Descrição': st.column_config.TextColumn('Descrição', width='large'),
                    'Valor Total': st.column_config.ProgressColumn(
                        'Valor Total',
                        format='R$ %.2f',
                        min_value=0,
                        max_value=max_valor_ncm if max_valor_ncm > 0 else 1
                    ),
                    'Empresas': st.column_config.NumberColumn('Empresas', format='%d'),
                    'Itens': st.column_config.NumberColumn('Itens', format='%d')
                }
            )
        
        # ---------------------------------------------------------------------
        # TAB: EMPRESAS
        # ---------------------------------------------------------------------
        with tab_empresas:
            st.markdown("#### Empresas que Utilizam este Produto")
            
            emp_stats = df.groupby(['cnpj_emitente', 'razao_emitente']).agg({
                'valor_infracao': ['count', 'sum'],
                'aliquota_emitente': lambda x: x.mode().iloc[0] if len(x.mode()) > 0 else None,
                'aliquota_ia_alta': lambda x: x.mode().iloc[0] if len(x.mode()) > 0 else None
            }).reset_index()
            emp_stats.columns = ['CNPJ', 'Razão Social', 'Quantidade', 'Valor Infração', 'Alíq. Emitente (Moda)', 'Alíq. IA (Moda)']
            emp_stats = emp_stats.sort_values('Valor Infração', ascending=False)
            
            # Top 20 empresas
            st.markdown(f"##### 🏆 Top 20 Empresas (de {len(emp_stats):,} total)")
            
            st.dataframe(
                emp_stats.head(20),
                use_container_width=True,
                hide_index=True,
                column_config={
                    'CNPJ': st.column_config.TextColumn('CNPJ'),
                    'Razão Social': st.column_config.TextColumn('Razão Social', width='large'),
                    'Quantidade': st.column_config.NumberColumn('Quantidade', format="%d"),
                    'Valor Infração': st.column_config.NumberColumn('Valor Infração', format="R$ %.2f"),
                    'Alíq. Emitente (Moda)': st.column_config.TextColumn('Alíq. Emitente'),
                    'Alíq. IA (Moda)': st.column_config.TextColumn('Alíq. IA')
                }
            )
            
            # Possibilidade de consultar empresa
            st.markdown("---")
            st.markdown("##### 🔍 Consultar Empresa")
            
            empresas_opcoes = emp_stats.head(50)[['CNPJ', 'Razão Social']].apply(
                lambda x: f"{x['CNPJ']} - {x['Razão Social'][:40]}", axis=1
            ).tolist()
            
            if empresas_opcoes:
                empresa_selecionada = st.selectbox(
                    "Selecione uma empresa para consultar:",
                    options=[""] + empresas_opcoes,
                    key="empresa_selecao_pesquisa"
                )
                
                if empresa_selecionada and st.button("🔍 Consultar Empresa Selecionada"):
                    cnpj_selecionado = empresa_selecionada.split(" - ")[0]
                    st.session_state.cnpj_pre_preenchido = cnpj_selecionado
                    st.session_state.nav_page = "consulta"
                    st.rerun()
        
        # ---------------------------------------------------------------------
        # TAB: DADOS
        # ---------------------------------------------------------------------
        with tab_dados:
            st.markdown("#### 📋 Dados Detalhados")
            
            # Filtros
            col1, col2, col3 = st.columns(3)
            
            with col1:
                ncm_filter = st.multiselect(
                    "Filtrar por NCM",
                    options=df['ncm'].dropna().unique().tolist(),
                    key="ncm_filter_pesquisa"
                )
            
            with col2:
                aliq_filter = st.multiselect(
                    "Filtrar por Alíquota Emitente",
                    options=sorted(df['aliquota_emitente'].dropna().unique().tolist()),
                    key="aliq_filter_pesquisa"
                )
            
            with col3:
                tipo_doc_filter = st.multiselect(
                    "Filtrar por Tipo Doc",
                    options=df['tipo_doc'].unique().tolist(),
                    default=df['tipo_doc'].unique().tolist(),
                    key="tipo_doc_filter_pesquisa"
                )
            
            # Aplica filtros
            df_filtered = df.copy()
            if ncm_filter:
                df_filtered = df_filtered[df_filtered['ncm'].isin(ncm_filter)]
            if aliq_filter:
                df_filtered = df_filtered[df_filtered['aliquota_emitente'].isin(aliq_filter)]
            if tipo_doc_filter:
                df_filtered = df_filtered[df_filtered['tipo_doc'].isin(tipo_doc_filter)]
            
            st.caption(f"Exibindo {len(df_filtered):,} de {len(df):,} registros")
            
            # Tabela de dados
            df_display = df_filtered[[
                'descricao', 'ncm', 'cfop', 'cnpj_emitente', 'razao_emitente',
                'aliquota_emitente', 'aliquota_ia_alta', 'aliquota_ia_media', 'aliquota_ia_baixa',
                'valor_infracao', 'tipo_doc'
            ]].copy()
            
            df_display.columns = [
                'Descrição', 'NCM', 'CFOP', 'CNPJ', 'Razão Social',
                'Alíq. Emit.', 'Alíq. IA Alta', 'Alíq. IA Média', 'Alíq. IA Baixa',
                'Valor Infração', 'Tipo'
            ]
            
            st.dataframe(
                df_display.head(500),
                use_container_width=True,
                hide_index=True,
                height=400
            )
            
            if len(df_filtered) > 500:
                st.caption("⚠️ Exibindo apenas os primeiros 500 registros. Use os filtros para refinar.")


# =============================================================================
# 10. INTERFACE PRINCIPAL
# =============================================================================


def main():
    """Interface principal com navegação em tabs na área principal."""
    
    # =========================================================================
    # GERENCIAMENTO DE MEMÓRIA AUTOMÁTICO
    # =========================================================================
    
    if 'last_activity' not in st.session_state:
        st.session_state.last_activity = datetime.now()
    
    time_since_activity = datetime.now() - st.session_state.last_activity
    if time_since_activity > timedelta(minutes=SESSION_TIMEOUT_MINUTES):
        if st.session_state.get('consulta_dados') is not None:
            st.session_state.consulta_dados = None
            keys_to_clear = [k for k in st.session_state.keys() 
                           if k.startswith(('excel_data_', 'network_save_', 'local_save_', 'analise_'))]
            for key in keys_to_clear:
                del st.session_state[key]
            st.cache_data.clear()
            gc.collect()
    
    st.session_state.last_activity = datetime.now()
    
    if 'consulta_dados' not in st.session_state:
        st.session_state.consulta_dados = None
    
    # Flag para indicar tabelas indisponíveis
    if 'tabela_indisponivel' not in st.session_state:
        st.session_state.tabela_indisponivel = False
    
    # Variável de controle para navegação (separada do widget)
    if 'nav_page' not in st.session_state:
        st.session_state.nav_page = "ranking"  # Default: ranking

    # Grupo selecionado (GESSUPER, GESMAC, etc.)
    if 'grupo_selecionado' not in st.session_state:
        st.session_state.grupo_selecionado = GRUPO_PADRAO

    engine = get_engine()
    if engine is None:
        st.stop()
    
    # =========================================================================
    # CSS PARA SIDEBAR SEMPRE COLAPSADO
    # =========================================================================
    
    # Sidebar sempre inicia colapsado - usuário abre quando quiser
    st.markdown("""
    <style>
        /* Sidebar sempre colapsado por padrão */
        section[data-testid="stSidebar"] {
            width: 0px !important;
            min-width: 0px !important;
            transform: translateX(-100%);
            transition: transform 0.3s ease-in-out, width 0.3s ease-in-out;
        }
        section[data-testid="stSidebar"]:hover,
        section[data-testid="stSidebar"]:focus-within {
            width: 300px !important;
            min-width: 300px !important;
            transform: translateX(0);
        }
        /* Indicador visual para expandir */
        section[data-testid="stSidebar"]::before {
            content: "☰";
            position: absolute;
            right: -30px;
            top: 50%;
            transform: translateY(-50%);
            font-size: 24px;
            color: #1565C0;
            cursor: pointer;
            z-index: 1000;
        }
    </style>
    """, unsafe_allow_html=True)
    
    # =========================================================================
    # SIDEBAR - INFORMAÇÕES E SISTEMA
    # =========================================================================
    
    with st.sidebar:
        # Obtém configuração do grupo atual
        grupo_atual = st.session_state.grupo_selecionado
        grupo_config = get_grupo_config(grupo_atual)

        st.markdown(f"""
        <div style='text-align: center; padding: 0.5rem 0; border-bottom: 2px solid #1565C0; margin-bottom: 1rem;'>
            <h2 style='color: #1565C0; margin: 0;'>🎯 ARGOS</h2>
            <p style='color: #666; margin: 0; font-size: 0.8rem;'>{grupo_config['nome_display']}</p>
        </div>
        """, unsafe_allow_html=True)

        # Seletor de grupo
        st.markdown("### 🏢 Grupo")
        grupos_disponiveis = list(GRUPOS_CONFIG.keys())
        grupo_idx = grupos_disponiveis.index(grupo_atual) if grupo_atual in grupos_disponiveis else 0

        novo_grupo = st.selectbox(
            "Selecione o grupo",
            options=grupos_disponiveis,
            index=grupo_idx,
            key="grupo_selector",
            label_visibility="collapsed"
        )

        # Atualiza o grupo selecionado se mudou
        if novo_grupo != grupo_atual:
            st.session_state.grupo_selecionado = novo_grupo
            # Limpa dados da consulta anterior ao trocar de grupo
            st.session_state.consulta_dados = None
            # Limpa caches específicos do grupo
            keys_to_clear = [k for k in st.session_state.keys()
                           if k.startswith(('excel_data_', 'analise_agg_'))]
            for key in keys_to_clear:
                del st.session_state[key]
            st.cache_data.clear()
            st.rerun()

        st.markdown("---")

        # Informações sobre os níveis de acurácia
        st.markdown("### 📊 Níveis de Acurácia")
        
        st.success("**🟢 ALTA**\n\nConsenso das 3 IAs\n\n*1-2% de erros esperados*")
        st.warning("**🟡 MÉDIA**\n\nMaioria 2x1\n\n*Até 5% de erros*")
        st.error("**🔴 BAIXA**\n\nIAs divergentes\n\n*Requer avaliação manual!*")
        
        st.markdown("---")
        
        # Sistema
        with st.expander("⚙️ Sistema", expanded=False):
            st.caption(f"Cache consulta: {CACHE_TTL_SECONDS//60} min")
            st.caption(f"Cache ranking: 24h")
            
            if st.button("🧹 Limpar Cache", use_container_width=True):
                st.cache_data.clear()
                st.cache_resource.clear()
                for key in list(st.session_state.keys()):
                    del st.session_state[key]
                gc.collect()
                st.rerun()
            
            if st.session_state.get('consulta_dados'):
                df_mem = st.session_state.consulta_dados.get('df')
                if df_mem is not None:
                    mem_mb = df_mem.memory_usage(deep=True).sum() / 1024 / 1024
                    st.info(f"📊 {mem_mb:.1f} MB ({len(df_mem):,} linhas)")
        
        st.markdown("---")
        st.caption("Receita Estadual de SC")
    
    # Variáveis para compatibilidade
    nav_page = st.session_state.nav_page
    nivel = "ALTA"  # Valor padrão, será sobrescrito pelas páginas
    consultar = False
    identificador = ""
    
    # =========================================================================
    # ÁREA PRINCIPAL - RANKING, CONSULTA OU PRODUTOS
    # =========================================================================
    
    # Se está no modo Ranking
    if nav_page == "ranking":
        render_ranking(engine, nivel)
        st.stop()
    
    # Se está no modo Pesquisa de Produtos
    if nav_page == "produtos":
        render_pesquisa_produtos(engine)
        st.stop()
    
    # =========================================================================
    # EXIBE CONTEÚDO
    # =========================================================================
    
    # Obtém configuração do grupo para uso na página principal
    grupo_cfg_main = get_grupo_config()

    if st.session_state.consulta_dados is None:
        # Header compacto com botões de navegação
        col_title, col_btn1, col_btn2 = st.columns([3, 1, 1])
        with col_title:
            st.markdown(f"""
            <h2 style='color: #1565C0; margin: 0;'>🎯 Operação ARGOS</h2>
            <p style='color: #666; margin: 0; font-size: 0.9rem;'>{grupo_cfg_main['descricao']}</p>
            """, unsafe_allow_html=True)
        with col_btn1:
            if st.button("🏆 Ranking", use_container_width=True, type="secondary"):
                st.session_state.nav_page = "ranking"
                st.rerun()
        with col_btn2:
            if st.button("🔎 Produtos", use_container_width=True, type="secondary"):
                st.session_state.nav_page = "produtos"
                st.rerun()
        
        st.markdown("---")
        
        # =====================================================================
        # VERIFICAÇÃO DE DISPONIBILIDADE DAS TABELAS
        # =====================================================================
        # Limpa flag de tabela indisponível
        st.session_state.tabela_indisponivel = False
        
        if not check_tables_available(engine):
            st.warning(TABLE_UNAVAILABLE_MSG)
            st.stop()
        
        # =====================================================================
        # CAMPO DE CONSULTA COMPACTO E CENTRALIZADO
        # =====================================================================
        col_esq, col_form, col_dir = st.columns([1, 3, 1])
        
        with col_form:
            st.markdown("### 🔍 Consultar Empresa")
            
            # Campo e seletor lado a lado
            col_input, col_nivel = st.columns([2, 1])
            
            with col_input:
                # Verifica se há CNPJ pré-preenchido (vindo da página de produtos)
                valor_inicial = st.session_state.pop('cnpj_pre_preenchido', '')
                cnpj_ie_input = st.text_input(
                    "CNPJ ou IE",
                    value=valor_inicial,
                    placeholder="00.000.000/0000-00 ou 000000000",
                    key="cnpj_input_principal",
                    label_visibility="collapsed"
                )
            
            with col_nivel:
                nivel_consulta_principal = st.selectbox(
                    "Nível",
                    options=["ALTA", "MEDIA", "BAIXA"],
                    format_func=lambda x: {"BAIXA": "🔴 BAIXA", "MEDIA": "🟡 MÉDIA", "ALTA": "🟢 ALTA"}[x],
                    key="nivel_input_principal",
                    index=0,
                    label_visibility="collapsed"
                )
            
            # Botão de consulta
            if st.button("🔎 CONSULTAR", type="primary", use_container_width=True):
                if cnpj_ie_input:
                    ident_digits = sanitize_identificador(cnpj_ie_input)
                    if ident_digits:
                        # Limpa flag de tabela indisponível
                        st.session_state.tabela_indisponivel = False
                        
                        with st.status("🔄 Consultando...", expanded=True) as status:
                            st.write("🔍 Buscando contribuinte...")
                            progress_bar = st.progress(0)
                            contrib_info = get_contribuinte_info(engine, ident_digits)
                            progress_bar.progress(25)
                            
                            # Verifica se houve erro de tabela indisponível
                            if st.session_state.get('tabela_indisponivel', False):
                                status.update(label="⚠️ Tabelas indisponíveis", state="error", expanded=False)
                                st.warning(TABLE_UNAVAILABLE_MSG)
                            else:
                                if contrib_info:
                                    st.write(f"✅ **{contrib_info.get('razao_social', 'N/A')}**")
                                else:
                                    st.write("⚠️ Contribuinte não encontrado")
                                
                                st.write(f"📊 Carregando infrações ({nivel_consulta_principal})...")
                                progress_bar.progress(50)
                                df = get_base_df(engine, ident_digits, nivel_consulta_principal)
                                progress_bar.progress(100)
                                
                                # Verifica novamente se houve erro de tabela indisponível
                                if st.session_state.get('tabela_indisponivel', False):
                                    status.update(label="⚠️ Tabelas indisponíveis", state="error", expanded=False)
                                    st.warning(TABLE_UNAVAILABLE_MSG)
                                elif df.empty:
                                    status.update(label="❌ Nenhum registro", state="error", expanded=False)
                                    st.warning(f"⚠️ Nenhum registro para: {cnpj_ie_input}")
                                else:
                                    status.update(label=f"✅ {len(df):,} registros", state="complete", expanded=False)
                                    st.session_state.consulta_dados = {
                                        'df': df,
                                        'contrib_info': contrib_info,
                                        'ident_digits': ident_digits,
                                        'identificador': cnpj_ie_input,
                                        'nivel': nivel_consulta_principal
                                    }
                                    st.rerun()
                    else:
                        st.error("⚠️ CNPJ ou IE inválido.")
                else:
                    st.warning("⚠️ Digite um CNPJ ou IE.")
        
        # Cards de níveis compactos
        st.markdown("---")
        col1, col2, col3 = st.columns(3)
        with col1:
            st.success("**🟢 ALTA** - Consenso 3 IAs (1-2% erros)")
        with col2:
            st.warning("**🟡 MÉDIA** - Maioria 2x1 (até 5% erros)")
        with col3:
            st.error("**🔴 BAIXA** - IAs divergentes ⚠️")
        
    else:
        dados = st.session_state.consulta_dados
        df = dados['df']
        contrib_info = dados['contrib_info']
        ident_digits = dados['ident_digits']
        identificador_consulta = dados['identificador']
        nivel_consulta = dados['nivel']
        
        nivel_atual = nivel_consulta
        total_nivel, cfg, has_rows = calcular_totais(df, nivel_atual)
        
        if contrib_info:
            razao_social = contrib_info.get('razao_social', 'N/A')
            cnpj_formatado = contrib_info.get('cnpj', identificador_consulta)
            ie_formatado = contrib_info.get('ie', '')
            municipio = contrib_info.get('municipio', '')
            gerfe = contrib_info.get('gerfe', '')
        else:
            razao_social = df['razao_emitente'].iloc[0] if 'razao_emitente' in df.columns and not df['razao_emitente'].isna().all() else "N/A"
            cnpj_formatado = identificador_consulta
            ie_formatado = ''
            municipio = ''
            gerfe = ''
        
        # =====================================================================
        # HEADER FIXO COM BOTÕES
        # =====================================================================
        col_header, col_btn1, col_btn2, col_btn3 = st.columns([3, 1, 1, 1])
        with col_btn1:
            if st.button("🔍 Nova Consulta", use_container_width=True, type="secondary"):
                st.session_state.consulta_dados = None
                st.rerun()
        with col_btn2:
            if st.button("🏆 Ranking", use_container_width=True, type="secondary"):
                st.session_state.nav_page = "ranking"
                st.rerun()
        with col_btn3:
            if st.button("🔎 Produtos", use_container_width=True, type="secondary"):
                st.session_state.nav_page = "produtos"
                st.rerun()
        
        st.markdown(f"""
        <div style='background: linear-gradient(135deg, #1565C0 0%, #0D47A1 100%); 
                    padding: 1rem 1.5rem; border-radius: 10px; margin-bottom: 1rem; color: white;'>
            <div style='display: flex; justify-content: space-between; align-items: center; flex-wrap: wrap;'>
                <div>
                    <h3 style='margin: 0; color: white;'>{razao_social}</h3>
                    <p style='margin: 0.3rem 0 0 0; opacity: 0.9;'>
                        CNPJ: {cnpj_formatado} | IE: {ie_formatado} | {municipio} | GERFE: {gerfe}
                    </p>
                </div>
                <div style='text-align: right;'>
                    <span style='background: {cfg["cor"]}; padding: 0.3rem 0.8rem; border-radius: 20px; 
                                 font-weight: bold;'>{cfg["emoji"]} {cfg["label"]}</span>
                    <p style='margin: 0.3rem 0 0 0; opacity: 0.9;'>{format_number_br(len(df))} itens | {format_currency_br(total_nivel)}</p>
                </div>
            </div>
        </div>
        """, unsafe_allow_html=True)
        
        # =====================================================================
        # TABS DE NAVEGAÇÃO
        # =====================================================================
        tab_resumo, tab_exportar, tab_comparativo, tab_analise = st.tabs([
            "📊 Resumo", "📥 Exportar", "🎯 Comparativo", "📈 Análise"
        ])
        
        # -----------------------------------------------------------------
        # TAB 1: RESUMO (com comparativos)
        # -----------------------------------------------------------------
        with tab_resumo:
            # Busca estatísticas globais para comparação (cache 24h)
            global_stats = get_global_stats(engine, nivel_atual)
            
            # Calcula métricas comparativas
            if global_stats and global_stats['total_valor'] > 0:
                pct_valor_global = (total_nivel / global_stats['total_valor']) * 100
                pct_itens_global = (len(df) / global_stats['total_itens']) * 100 if global_stats['total_itens'] > 0 else 0
                media_global = global_stats['total_valor'] / global_stats['total_empresas'] if global_stats['total_empresas'] > 0 else 0
                diff_media = total_nivel - media_global
            else:
                pct_valor_global = 0
                pct_itens_global = 0
                media_global = 0
                diff_media = 0
            
            # Cards com comparativos
            st.markdown("### 📊 Resumo da Empresa")
            
            col1, col2, col3, col4 = st.columns(4)
            with col1:
                st.metric(
                    f"{cfg['emoji']} Total Infração",
                    format_currency_br(total_nivel),
                    delta=f"{pct_valor_global:.2f}% do total" if pct_valor_global > 0 else None,
                    delta_color="off"
                )
            with col2:
                st.metric(
                    "📦 Qtd. Itens",
                    format_number_br(len(df)),
                    delta=f"{pct_itens_global:.2f}% do total" if pct_itens_global > 0 else None,
                    delta_color="off"
                )
            with col3:
                periodos = df['periodo'].nunique() if 'periodo' in df.columns else 0
                st.metric("📅 Períodos", periodos)
            with col4:
                if 'data_emissao' in df.columns:
                    df_datas = pd.to_datetime(df['data_emissao'], errors='coerce')
                    if not df_datas.isna().all():
                        periodo_range = f"{df_datas.min().strftime('%m/%Y')} - {df_datas.max().strftime('%m/%Y')}"
                    else:
                        periodo_range = "N/A"
                else:
                    periodo_range = "N/A"
                st.metric("📆 Range", periodo_range)
            
            # Segunda linha de métricas comparativas
            col1, col2, col3, col4 = st.columns(4)
            with col1:
                media_empresa = total_nivel / len(df) if len(df) > 0 else 0
                st.metric(
                    "💵 Média/Item",
                    format_currency_br(media_empresa),
                    delta=f"{((media_empresa/global_stats['media_item'])-1)*100:.1f}% vs média geral" if global_stats and global_stats['media_item'] > 0 else None,
                    delta_color="inverse"
                )
            with col2:
                st.metric(
                    "📈 vs Média Empresas",
                    format_currency_br(diff_media),
                    delta="acima" if diff_media > 0 else "abaixo",
                    delta_color="inverse" if diff_media > 0 else "normal"
                )
            with col3:
                if global_stats and global_stats['total_empresas'] > 0:
                    # Estima posição no ranking (simplificado)
                    posicao_estimada = max(1, int(global_stats['total_empresas'] * (1 - pct_valor_global/100)))
                    st.metric("🏆 Ranking Estimado", f"Top {min(posicao_estimada, 100)}")
                else:
                    st.metric("🏆 Ranking", "N/A")
            with col4:
                if global_stats:
                    st.metric("🏢 Total Empresas", f"{global_stats['total_empresas']:,}")
                else:
                    st.metric("🏢 Total Empresas", "N/A")
            
            # Informações comparativas em texto
            if global_stats and pct_valor_global > 0:
                st.markdown("---")
                st.markdown(f"""
                <div style='background: #f8f9fa; padding: 1rem; border-radius: 8px; border-left: 4px solid {cfg["cor"]};'>
                    <p style='margin: 0;'>
                        <strong>📊 Comparativo:</strong> Esta empresa representa <strong>{pct_valor_global:.2f}%</strong> 
                        do valor total de infrações no nível {cfg['label']}. 
                        Com <strong>{format_number_br(len(df))}</strong> itens ({pct_itens_global:.2f}% do total).
                    </p>
                </div>
                """, unsafe_allow_html=True)
        
        # -----------------------------------------------------------------
        # TAB 2: EXPORTAR
        # -----------------------------------------------------------------
        with tab_exportar:
            # Obtém grupo atual para determinar modelos de exportação
            grupo_export = st.session_state.get('grupo_selecionado', GRUPO_PADRAO)
            grupo_cfg_export = get_grupo_config(grupo_export)
            modelos_export = grupo_cfg_export.get('modelos_exportacao', ['Anexo J'])

            # Seletor de modelo para grupos com múltiplos modelos (ex: GESMAC)
            modelo_selecionado = None
            if len(modelos_export) > 1:
                st.markdown("### 📋 Modelo de Exportação")
                modelo_selecionado = st.selectbox(
                    "Selecione o modelo de exportação",
                    options=modelos_export,
                    key="modelo_export_selector",
                    help="Selecione o tipo de documento para exportar"
                )
                st.markdown("---")

            df_export = build_export_df(df, nivel_atual, grupo=grupo_export, modelo_export=modelo_selecionado)

            if df_export is not None and not df_export.empty:
                total_rows = len(df_export)
                needs_split = total_rows > MAX_ROWS_PER_EXCEL
                is_large_file = total_rows > LARGE_FILE_WARNING
                
                filename_csv = get_export_filename(contrib_info, nivel_atual, "csv")
                cache_key = f"excel_data_{ident_digits}_{nivel_atual}"
                
                if cache_key not in st.session_state:
                    st.session_state[cache_key] = None
                
                if needs_split:
                    num_partes = math.ceil(total_rows / MAX_ROWS_PER_EXCEL)
                    st.warning(f"⚠️ {total_rows:,} linhas → {num_partes} arquivos")
                elif is_large_file:
                    st.info(f"📊 {total_rows:,} linhas")
                
                # Aviso sobre bloqueio para arquivos grandes
                if total_rows > 100000:
                    st.warning("""
                    ⚠️ **Atenção:** Arquivos Excel com muitas linhas podem levar **2-4 minutos** para gerar.
                    Durante esse tempo, a aplicação pode ficar lenta para outros usuários.
                    
                    **Recomendação:** Use **CSV** (gera em segundos) ou **Salvar na Rede** (mais rápido).
                    """)
                
                sub_tab_rede, sub_tab_download = st.tabs(["💾 Rede (Recomendado)", "📥 Download"])
                
                with sub_tab_rede:
                    st.markdown("**📁 Caminho de Rede:**")
                    # Usa text_input disabled para permitir seleção e cópia fácil
                    st.text_input(
                        "Caminho",
                        value=REDE_PATH,
                        disabled=True,
                        label_visibility="collapsed",
                        help="Selecione e copie com Ctrl+C"
                    )
                    st.caption("💡 Clique no campo acima, selecione tudo (Ctrl+A) e copie (Ctrl+C)")
                    
                    col1, col2 = st.columns(2)
                    with col1:
                        if st.button("🚀 Salvar CSV", use_container_width=True, type="primary"):
                            progress_bar = st.progress(0, text="Iniciando...")
                            progress_bar.progress(10, text="📊 Preparando dados (10%)...")
                            success, message, filepath, _ = save_csv_to_network(df_export, contrib_info, nivel_atual)
                            progress_bar.progress(100, text="✅ Concluído (100%)")
                            if success:
                                st.success(f"✅ {message}")
                                st.code(filepath)
                            else:
                                st.error(message)
                    with col2:
                        if st.button("💾 Salvar Excel", use_container_width=True):
                            progress_bar = st.progress(0, text="Iniciando exportação Excel...")
                            status_text = st.empty()
                            
                            status_text.info("⏳ Gerando arquivo Excel... Isso pode levar alguns minutos.")
                            
                            # Etapa 1: Preparação (0-5%)
                            progress_bar.progress(5, text="📋 Preparando estrutura (5%)...")
                            
                            # Etapa 2: Gera Excel com progresso simulado
                            progress_bar.progress(10, text="📊 Processando dados (10%)...")
                            
                            # Callback de progresso
                            def progress_callback(current, total, msg):
                                pct = int(10 + (current / max(total, 1)) * 80)  # 10% a 90%
                                progress_bar.progress(pct, text=f"{msg} ({pct}%)")
                            
                            success, message, file_paths, _ = save_to_network_fast(
                                df_export, contrib_info, nivel_atual, progress_callback
                            )
                            
                            progress_bar.progress(100, text="✅ Concluído (100%)")
                            status_text.empty()
                            
                            if success:
                                st.success(f"✅ {message}")
                                for fp in file_paths:
                                    st.code(fp)
                            else:
                                st.error(message)
                
                with sub_tab_download:
                    col1, col2 = st.columns(2)
                    with col1:
                        csv_data = export_to_csv(df_export, ident_digits, nivel_atual)
                        st.download_button("📥 CSV", csv_data, file_name=filename_csv, mime="text/csv", use_container_width=True)
                    with col2:
                        if st.button("📊 Gerar Excel", use_container_width=True):
                            progress_bar = st.progress(0, text="Iniciando geração do Excel...")
                            status_text = st.empty()
                            
                            total_rows_export = len(df_export)
                            tempo_estimado = max(60, total_rows_export // 5000)  # ~5000 linhas/seg
                            status_text.info(f"⏳ Gerando Excel ({total_rows_export:,} linhas)... Tempo estimado: ~{tempo_estimado//60} min {tempo_estimado%60} seg")
                            
                            # Callback de progresso real
                            def progress_callback_download(pct, msg):
                                progress_bar.progress(pct, text=f"{msg} ({pct}%)")
                            
                            # Chama a função de exportação com callback
                            excel_bytes = export_to_excel_template(
                                df_export, contrib_info, nivel_atual,
                                progress_callback=progress_callback_download
                            )
                            
                            st.session_state[cache_key] = excel_bytes
                            
                            progress_bar.progress(100, text="✅ Excel gerado com sucesso! (100%)")
                            
                            # Calcula tamanho do arquivo
                            tamanho_mb = len(excel_bytes) / (1024 * 1024)
                            status_text.success(f"✅ Arquivo pronto! **{tamanho_mb:.1f} MB** - Clique em '📥 Baixar Excel' abaixo.")
                            
                            # Força rerun para mostrar o botão de download
                            st.rerun()
                        
                        if st.session_state.get(cache_key):
                            excel_data = st.session_state[cache_key]
                            tamanho_mb = len(excel_data) / (1024 * 1024)
                            
                            st.info(f"📦 Arquivo pronto: **{tamanho_mb:.1f} MB** ({len(df_export):,} linhas)")
                            
                            # Aviso para arquivos grandes
                            if tamanho_mb > 50:
                                st.warning("⚠️ Arquivo grande! O download pode levar alguns segundos. Aguarde o navegador processar.")
                            
                            st.download_button(
                                "📥 Baixar Excel", 
                                excel_data,
                                file_name=get_export_filename(contrib_info, nivel_atual, "xlsx"),
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                use_container_width=True,
                                type="primary",
                                help="Clique e aguarde o navegador iniciar o download"
                            )
                            
                            st.caption("💡 Se o download não iniciar, verifique se seu navegador bloqueou popups.")
        
        # -----------------------------------------------------------------
        # TAB 3: COMPARATIVO
        # -----------------------------------------------------------------
        with tab_comparativo:
            # Passa períodos disponíveis para filtro correto
            periodos_disponiveis = df['periodo'].dropna().unique().tolist() if 'periodo' in df.columns else None
            render_comparativo_niveis(engine, ident_digits, len(df), periodos_disponiveis)
        
        # -----------------------------------------------------------------
        # TAB 4: ANÁLISE (OTIMIZADA - LAZY LOADING)
        # -----------------------------------------------------------------
        with tab_analise:
            col_infracao = 'infracao_ia' if 'infracao_ia' in df.columns else cfg['col_infracao']
            
            # Verifica se precisa filtrar por período (datasets grandes)
            total_rows = len(df)
            if total_rows > LARGE_DATASET_THRESHOLD:
                st.warning(f"⚠️ Dataset grande ({total_rows:,} linhas). Análise limitada aos **últimos 12 meses** para melhor performance.")
                
                # Filtra últimos 12 meses
                if 'periodo' in df.columns and df['periodo'].notna().any():
                    # Gera lista dos últimos 12 períodos no formato MM/AAAA
                    hoje = datetime.now()
                    ultimos_12_meses = []
                    for i in range(12):
                        mes = hoje.month - i
                        ano = hoje.year
                        while mes <= 0:
                            mes += 12
                            ano -= 1
                        ultimos_12_meses.append(f"{mes:02d}/{ano}")
                    
                    # Tenta filtrar pelos últimos 12 meses
                    df_filtrado = df[df['periodo'].isin(ultimos_12_meses)]
                    
                    if len(df_filtrado) > 0:
                        df_analise = df_filtrado.copy()
                        st.caption(f"📊 Analisando {len(df_analise):,} registros dos últimos 12 meses")
                    else:
                        # Se não encontrou, pega os períodos mais recentes disponíveis
                        periodos_disponiveis = df['periodo'].dropna().unique()
                        # Ordena períodos (formato MM/AAAA -> AAAA/MM para ordenar)
                        periodos_ordenados = sorted(
                            periodos_disponiveis, 
                            key=lambda x: f"{x[3:7]}/{x[0:2]}" if len(str(x)) >= 7 else x,
                            reverse=True
                        )[:12]  # Pega os 12 mais recentes
                        df_analise = df[df['periodo'].isin(periodos_ordenados)].copy()
                        st.caption(f"📊 Analisando {len(df_analise):,} registros dos 12 períodos mais recentes")
                else:
                    # Se não tem período, limita a 200k linhas
                    df_analise = df.head(LARGE_DATASET_THRESHOLD).copy()
                    st.caption(f"📊 Analisando primeiros {len(df_analise):,} registros")
            else:
                df_analise = df
                st.caption("💡 Clique nas seções para expandir")
            
            # Agregações são cacheadas no session_state
            agg_key = f"analise_agg_{ident_digits}_{nivel_atual}"
            
            # Verifica se df_analise tem dados
            if len(df_analise) == 0:
                st.error("❌ Nenhum dado encontrado para análise nos períodos selecionados.")
                st.info("💡 Verifique se há dados nos últimos 12 meses ou use a aba Resumo para ver todos os dados.")
            else:
                # ----- ESTATÍSTICAS (PRIMEIRO E EXPANDIDO) -----
                with st.expander("📊 Estatísticas", expanded=True):
                    # Força recálculo se não estiver no cache
                    cache_key_stats = f"{agg_key}_stats_{len(df_analise)}"
                    if cache_key_stats not in st.session_state:
                        valores = pd.to_numeric(df_analise[col_infracao], errors='coerce').fillna(0)
                        if len(valores) > 0 and valores.sum() > 0:
                            st.session_state[cache_key_stats] = valores.describe()
                        else:
                            st.session_state[cache_key_stats] = None
                    
                    stats = st.session_state.get(cache_key_stats)
                    
                    if stats is not None and not pd.isna(stats.get('mean', float('nan'))):
                        col1, col2, col3, col4 = st.columns(4)
                        col1.metric("Média", format_currency_br(stats['mean']))
                        col2.metric("Mediana", format_currency_br(stats['50%']))
                        col3.metric("Mínimo", format_currency_br(stats['min']))
                        col4.metric("Máximo", format_currency_br(stats['max']))
                        
                        col1, col2, col3, col4 = st.columns(4)
                        col1.metric("Total", format_currency_br(stats['mean'] * stats['count']))
                        col2.metric("Qtd. Itens", format_number_br(int(stats['count'])))
                        col3.metric("Desvio Padrão", format_currency_br(stats['std']))
                        col4.metric("75º Percentil", format_currency_br(stats['75%']))
                    else:
                        st.warning("⚠️ Não há dados numéricos válidos para calcular estatísticas.")
                
                # ----- VISUALIZAÇÕES TEMPORAIS -----
                with st.expander("📅 Evolução Temporal", expanded=False):
                    if 'periodo' in df_analise.columns:
                        # Seletor de tipo de visualização
                        tipo_viz = st.radio(
                            "Tipo de visualização:",
                            ["📅 Evolução Temporal", "📊 Agregado por Mês"],
                            horizontal=True,
                            key="tipo_viz_analise"
                        )
                        
                        # Cache para agregações
                        cache_key_periodo = f"agg_periodo_{ident_digits}_{nivel_atual}"
                        
                        if cache_key_periodo not in st.session_state:
                            df_temp = df_analise[['periodo', col_infracao]].copy()
                            df_temp['valor'] = pd.to_numeric(df_temp[col_infracao], errors='coerce').fillna(0)
                            
                            # Agrupa por período
                            df_agg = df_temp.groupby('periodo')['valor'].agg(['sum', 'count']).reset_index()
                            df_agg.columns = ['Período', 'Valor', 'Qtd']
                            
                            # Cria coluna para ordenação cronológica (AAAA-MM)
                            df_agg['ordem'] = df_agg['Período'].apply(
                                lambda x: f"{x[3:7]}-{x[0:2]}" if len(str(x)) >= 7 else x
                            )
                            
                            # Extrai mês para agregação mensal
                            df_agg['Mes'] = df_agg['Período'].apply(
                                lambda x: x[0:2] if len(str(x)) >= 2 else '00'
                            )
                            
                            st.session_state[cache_key_periodo] = df_agg
                        
                        df_periodo = st.session_state[cache_key_periodo].copy()
                        
                        # Mapeamento de mês para nome
                        meses_nome = {
                            '01': 'Jan', '02': 'Fev', '03': 'Mar',
                            '04': 'Abr', '05': 'Mai', '06': 'Jun',
                            '07': 'Jul', '08': 'Ago', '09': 'Set',
                            '10': 'Out', '11': 'Nov', '12': 'Dez'
                        }
                        
                        col1, col2 = st.columns(2)
                        
                        if tipo_viz == "📅 Evolução Temporal":
                            df_plot = df_periodo.sort_values('ordem')
                            
                            with col1:
                                fig = px.bar(df_plot, x='Período', y='Valor', 
                                            title="💰 Valor por Período",
                                            color_discrete_sequence=[cfg['cor']])
                                fig.update_layout(showlegend=False, xaxis_tickangle=-45, height=350)
                                st.plotly_chart(fig, use_container_width=True, key="bar_valor_temporal")
                            
                            with col2:
                                fig = px.line(df_plot, x='Período', y='Qtd', 
                                             title="📊 Quantidade por Período", 
                                             markers=True)
                                fig.update_layout(xaxis_tickangle=-45, height=350)
                                st.plotly_chart(fig, use_container_width=True, key="line_qtd_temporal")
                        
                        else:  # Agregado por Mês
                            df_mes = df_periodo.groupby('Mes').agg({
                                'Valor': 'sum',
                                'Qtd': 'sum'
                            }).reset_index()
                            df_mes['Mes_Nome'] = df_mes['Mes'].map(meses_nome)
                            df_mes = df_mes.sort_values('Mes')
                            
                            with col1:
                                fig = px.bar(df_mes, x='Mes_Nome', y='Valor', 
                                            title="💰 Valor Agregado por Mês",
                                            color_discrete_sequence=[cfg['cor']])
                                fig.update_layout(showlegend=False, xaxis_title="Mês", height=350)
                                st.plotly_chart(fig, use_container_width=True, key="bar_valor_mes")
                            
                            with col2:
                                fig = px.bar(df_mes, x='Mes_Nome', y='Qtd', 
                                            title="📊 Quantidade Agregada por Mês",
                                            color_discrete_sequence=['#1976D2'])
                                fig.update_layout(xaxis_title="Mês", height=350)
                                st.plotly_chart(fig, use_container_width=True, key="bar_qtd_mes")
                    else:
                        st.info("Coluna 'periodo' não disponível para visualização temporal.")
                
                # ----- NCM/CFOP (EM TEXTO) -----
                with st.expander("🏷️ Top 10 NCM / CFOP", expanded=False):
                    col1, col2 = st.columns(2)
                    
                    with col1:
                        st.markdown("##### 🏷️ Top 10 NCMs")
                        if 'ncm' in df_analise.columns:
                            if f"{agg_key}_ncm" not in st.session_state:
                                df_temp = df_analise[['ncm', col_infracao]].copy()
                                df_temp['valor'] = pd.to_numeric(df_temp[col_infracao], errors='coerce').fillna(0)
                                df_ncm = df_temp.groupby('ncm')['valor'].agg(['sum', 'count']).reset_index()
                                df_ncm.columns = ['NCM', 'Valor', 'Qtd']
                                df_ncm = df_ncm.nlargest(10, 'Valor')
                                df_ncm['Valor_fmt'] = df_ncm['Valor'].apply(format_currency_br)
                                df_ncm['Qtd_fmt'] = df_ncm['Qtd'].apply(lambda x: f"{x:,}".replace(',', '.'))
                                
                                # Busca descrições dos NCMs
                                ncm_desc = get_ncm_descricoes(engine, df_ncm['NCM'].tolist())
                                df_ncm['Descricao'] = df_ncm['NCM'].astype(str).map(ncm_desc).fillna('')
                                
                                st.session_state[f"{agg_key}_ncm"] = df_ncm
                            
                            df_ncm = st.session_state[f"{agg_key}_ncm"]
                            for i, row in df_ncm.iterrows():
                                ncm_code = row['NCM']
                                descricao = row.get('Descricao', '')
                                # Trunca descrição se muito longa
                                if descricao and len(descricao) > 80:
                                    descricao = descricao[:80] + "..."
                                
                                if descricao:
                                    st.markdown(f"**{ncm_code}** — {row['Valor_fmt']} ({row['Qtd_fmt']} itens)")
                                    st.caption(f"↳ {descricao}")
                                else:
                                    st.markdown(f"**{ncm_code}** — {row['Valor_fmt']} ({row['Qtd_fmt']} itens)")
                    
                    with col2:
                        st.markdown("##### 📋 Top 10 CFOPs")
                        if 'cfop' in df_analise.columns:
                            if f"{agg_key}_cfop" not in st.session_state:
                                df_temp = df_analise[['cfop', col_infracao]].copy()
                                df_temp['valor'] = pd.to_numeric(df_temp[col_infracao], errors='coerce').fillna(0)
                                df_cfop = df_temp.groupby('cfop')['valor'].agg(['sum', 'count']).reset_index()
                                df_cfop.columns = ['CFOP', 'Valor', 'Qtd']
                                df_cfop = df_cfop.nlargest(10, 'Valor')
                                df_cfop['Valor_fmt'] = df_cfop['Valor'].apply(format_currency_br)
                                df_cfop['Qtd_fmt'] = df_cfop['Qtd'].apply(lambda x: f"{x:,}".replace(',', '.'))
                                
                                # Busca descrições dos CFOPs
                                cfop_desc = get_cfop_descricoes(engine, df_cfop['CFOP'].tolist())
                                df_cfop['Descricao'] = df_cfop['CFOP'].astype(str).map(cfop_desc).fillna('')
                                
                                st.session_state[f"{agg_key}_cfop"] = df_cfop
                            
                            df_cfop = st.session_state[f"{agg_key}_cfop"]
                            for i, row in df_cfop.iterrows():
                                cfop_code = row['CFOP']
                                descricao = row.get('Descricao', '')
                                # Trunca descrição se muito longa
                                if descricao and len(descricao) > 60:
                                    descricao = descricao[:60] + "..."
                                
                                if descricao:
                                    st.markdown(f"**{cfop_code}** — {row['Valor_fmt']} ({row['Qtd_fmt']} itens)")
                                    st.caption(f"↳ {descricao}")
                                else:
                                    st.markdown(f"**{cfop_code}** — {row['Valor_fmt']} ({row['Qtd_fmt']} itens)")
                
                # ----- PRODUTOS (HEATMAP TOP 10) -----
                with st.expander("📦 Top 10 Produtos", expanded=False):
                    if 'descricao' in df_analise.columns:
                        if f"{agg_key}_prod" not in st.session_state:
                            df_temp = df_analise[['descricao', col_infracao]].copy()
                            df_temp['valor'] = pd.to_numeric(df_temp[col_infracao], errors='coerce').fillna(0)
                            df_prod = df_temp.groupby('descricao')['valor'].agg(['sum', 'count']).reset_index()
                            df_prod.columns = ['Produto', 'Valor', 'Qtd']
                            df_prod = df_prod.nlargest(10, 'Valor').reset_index(drop=True)
                            st.session_state[f"{agg_key}_prod"] = df_prod
                        
                        df_prod = st.session_state[f"{agg_key}_prod"]
                        
                        # Heatmap com Plotly
                        fig = px.imshow(
                            df_prod[['Valor']].T,
                            labels=dict(x="Produto", y="", color="Valor (R$)"),
                            x=df_prod['Produto'].apply(lambda x: x[:30] + '...' if len(str(x)) > 30 else x),
                            y=['Valor'],
                            color_continuous_scale='Blues',
                            aspect='auto'
                        )
                        fig.update_layout(
                            title="🔥 Heatmap - Top 10 Produtos por Valor",
                            height=150,
                            xaxis_tickangle=-45
                        )
                        st.plotly_chart(fig, use_container_width=True, key="heatmap_produtos")
                        
                        # Tabela detalhada
                        st.markdown("##### 📋 Detalhamento")
                        for i, row in df_prod.iterrows():
                            pct = (row['Valor'] / df_prod['Valor'].sum()) * 100
                            st.markdown(f"**{i+1}.** {row['Produto'][:50]}{'...' if len(str(row['Produto'])) > 50 else ''}")
                            st.caption(f"   💰 {format_currency_br(row['Valor'])} | 📦 {row['Qtd']:,} itens | {pct:.1f}%")
                
                # ----- DADOS -----
                with st.expander("📋 Visualizar Dados", expanded=False):
                    col1, col2 = st.columns(2)
                    with col1:
                        n_rows = st.selectbox("Linhas", [50, 100, 200, 500], index=1)
                    with col2:
                        if 'periodo' in df_analise.columns:
                            periodos = ['Todos'] + sorted(df_analise['periodo'].unique().tolist())
                            periodo_filter = st.selectbox("Período", periodos)
                        else:
                            periodo_filter = 'Todos'
                    
                    df_show = df_analise
                    if 'periodo' in df_analise.columns and periodo_filter != 'Todos':
                        df_show = df_analise[df_analise['periodo'] == periodo_filter]
                    
                    st.dataframe(df_show.head(n_rows), use_container_width=True)
                    st.caption(f"Exibindo {min(n_rows, len(df_show))} de {len(df_show)}")


# =============================================================================
# EXECUÇÃO
# =============================================================================

if __name__ == "__main__":
    main()